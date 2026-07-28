import streamlit as st
import pandas as pd
import os
import time
import io
import traceback
import threading
import uuid
import re
import sqlite3
import json
import random
import math  # [新增] 愛校服務 2.0
import unicodedata  # [Fix] 剝除 Notion rich_text 夾帶的隱藏格式字元
import concurrent.futures
from contextlib import closing
from datetime import datetime, date, timedelta
from datetime import timezone
import pytz
import gspread
# import fitz  # [移除] 消警告單已改用 Excel 產製，不再需要 PyMuPDF
from google.oauth2.service_account import Credentials as SACredentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
from streamlit.runtime.scriptrunner import add_script_run_ctx, get_script_run_ctx
from PIL import Image, ImageOps

try:
    from notion_client import Client
    NOTION_INSTALLED = True
except ImportError:
    NOTION_INSTALLED = False

# --- [DIAG v4] Module-level Worker log，用 cache_resource 保護，跨 rerun 不被清空 ---
import collections

@st.cache_resource
def _get_worker_log():
    return collections.deque(maxlen=100)

_WORKER_LOG = _get_worker_log()

# --- 1. 網頁設定 ---
# 透過 Streamlit Secrets 判斷目前是測試區還是正式區 (預設為正式區)
sys_env = st.secrets.get("ENV", "PROD")

# [V5.31 Patch 1] 確保 set_page_config 是第一個執行的指令
if sys_env == "DEV":
    st.set_page_config(page_title="🔧測試版-中壢家商，衛愛而生", layout="wide", page_icon="🧹")
    st.sidebar.info(f"🕵️‍♀️ 系統目前抓到的身分證是：[{sys_env}]")
    st.warning("🚧 **目前位於 DEV 測試環境！** 在這裡送出的資料僅供測試，不會影響正式成績。")
else:
    st.set_page_config(page_title="中壢家商，衛愛而生", layout="wide", page_icon="🧹")


# --- 2. 核心參數與全域設定 ---
try:
    TW_TZ = pytz.timezone('Asia/Taipei')
    MAX_IMAGE_BYTES = 20 * 1024 * 1024
    # [Fix #3-B] UPLOAD_SEM 與 IMG_DIR 已移除：照片改為同步上傳 Drive，不再暫存本機
    QUEUE_DB_PATH = "local_status_v5.db"  # 只保留 service_issued dedup + system_status 監控
    
    SHEET_URL = "https://docs.google.com/spreadsheets/d/11BXtN3aevJls6Q2IR_IbT80-9XvhBkjbTCgANmsxqkg/edit"
    SHEET_TABS = {
        "main": "main_data", "settings": "settings", "roster": "roster",
        "inspectors": "inspectors", "duty": "duty",
        "appeals": "appeals", "holidays": "holidays", "service_hours": "service_hours",
        "office_areas": "office_areas", "published_results": "published_results",
        "task_queue": "task_queue_v3",  # [Fix #3-B] 雲端佇列分頁，取代本機 SQLite task_queue 表
        "student_debts": "student_debts", "debt_history": "debt_history",  # [新增] 愛校服務 2.0
        "class_areas": "class_areas"  # [新增] 班級 → 外掃區域 對照表
    }

    EXPECTED_COLUMNS = [
        "日期", "週次", "班級", "評分項目", "檢查人員",
        "內掃原始分", "外掃原始分", "垃圾原始分", "垃圾內掃原始分", "垃圾外掃原始分", "晨間打掃原始分", "手機人數",
        "備註", "違規細項", "照片路徑", "登錄時間", "修正", "紀錄ID"
    ]
    APPEAL_COLUMNS = ["申訴日期", "班級", "違規日期", "違規項目", "原始扣分", "申訴理由", "佐證照片", "處理狀態", "登錄時間", "對應紀錄ID", "審核回覆"]

    # ==========================================
    # [V5.34] safe_cached：像 @st.cache_data 但「失敗結果不被 cache」
    # ----------------------------------------------------------------
    # 修掉「重新部署時尖峰流量打爆 API → 預設值被 cache 數小時」的雷。
    # 規則：
    #   - 成功（含資料 / 合法的空資料）→ 正常 cache，遵守 TTL
    #   - 函式 raise 例外 → 不會 cache → 下次呼叫會重試
    #   - 例外被本 wrapper 接住 → 印 log + 回傳 default_factory() 給呼叫端
    #
    # 用法：
    #   @safe_cached(ttl=300, default_factory=lambda: pd.DataFrame())
    #   def load_foo():
    #       ws = get_worksheet(...)
    #       if not ws: raise RuntimeError("worksheet unavailable")
    #       # 正常邏輯，例外就讓它往外丟，不要 catch 後 return default
    #       return result
    # ==========================================
    def safe_cached(ttl, default_factory):
        def decorator(inner):
            # [Fix] 直接把 @st.cache_data 套在 inner 身上，保留每個 loader 的唯一 qualname。
            # 之前用中間 wrapper "cached" 會讓 12 個 loader 共用同一個 cache key → 災難。
            cached = st.cache_data(ttl=ttl)(inner)

            def wrapper(*args, **kwargs):
                try:
                    return cached(*args, **kwargs)
                except Exception as e:
                    print(f"[safe_cached][{inner.__name__}] fallback to default: {e}")
                    return default_factory()

            wrapper.clear = cached.clear  # 保留 .clear() 介面與舊用法相容
            wrapper.__name__ = inner.__name__
            return wrapper
        return decorator

    # ==========================================
    # Notion API 輔助函式 
    # ==========================================
    @st.cache_resource
    def get_notion_client():
        if NOTION_INSTALLED:
            token = st.secrets.get("notion_token") or st.secrets.get("system_config", {}).get("notion_token")
            if token: return Client(auth=token)
        return None

    def fetch_available_notion_tasks():
        client = get_notion_client()
        db_id = st.secrets.get("notion_db_id") or st.secrets.get("system_config", {}).get("notion_db_id")
        if not client or not db_id: 
            return [], "系統尚未設定 Notion Token 或 Database ID"
        
        try:
            response = client.databases.query(
                database_id=db_id,
                filter={"property": "任務狀態", "status": {"equals": "等待認領中"}} # ⭐️ 改為無表情符號
            )
            tasks = []
            for page in response.get("results", []):
                props = page.get("properties", {})
                title = props.get("任務名稱", {}).get("title", [{}])
                title_text = title[0].get("text", {}).get("content", "未命名任務") if title else "未命名任務"
                
                date_obj = props.get("任務日期", {}).get("date", {})
                raw_date = date_obj.get("start", "未定") if date_obj else "未定"
                if raw_date != "未定":
                    try:
                        parsed_date = datetime.fromisoformat(raw_date.replace("Z", "+00:00"))
                        if len(raw_date) <= 10:
                            date_val = parsed_date.strftime("%Y-%m-%d")
                        else:
                            date_val = parsed_date.strftime("%Y-%m-%d %H:%M")
                    except Exception:
                        date_val = raw_date
                else:
                    date_val = "未定"
                
                area = props.get("任務內容", {}).get("rich_text", [{}])
                area_text = area[0].get("text", {}).get("content", "未填寫") if area else "未填寫"

                req_num_obj = props.get("需求人數", {}).get("number")
                req_num = req_num_obj if req_num_obj else 1  
                
                claimed_obj = props.get("認領學號", {}).get("rich_text", [])
                # [Fix] 合併所有 rich_text blocks，Notion 超過一定長度會自動拆成多個 block
                claimed_str = "".join(b.get("text", {}).get("content", "") for b in claimed_obj)
                current_claimants = [s.strip() for s in claimed_str.split(",") if s.strip()]
                current_count = len(current_claimants)
                
                tasks.append({
                    "id": page["id"], "title": title_text, "date": date_val, "area": area_text,
                    "req_num": req_num, "current_count": current_count
                })
            return tasks, None
        except Exception as e:
            return [], f"Notion API 讀取失敗詳細錯誤: {str(e)}"

    def claim_notion_task(page_id, student_id, purpose_tag=""):  
        client = get_notion_client()
        if not client:
            return False, "Notion 服務目前未啟用或連線失敗，請通知管理員檢查系統設定。"
            
        try:
            page = client.pages.retrieve(page_id=page_id)
            props = page.get("properties", {})
            
            req_num_obj = props.get("需求人數", {}).get("number")
            req_num = req_num_obj if req_num_obj else 1
            
            claimed_obj = props.get("認領學號", {}).get("rich_text", [])
            # [Fix] 合併所有 rich_text blocks，Notion 超過一定長度會自動拆成多個 block
            claimed_str = "".join(b.get("text", {}).get("content", "") for b in claimed_obj)
            current_claimants = [s.strip() for s in claimed_str.split(",") if s.strip()]
            
            if any(str(student_id) in c for c in current_claimants):  
                return False, f"學號 {student_id} 已經認領過此任務囉！"
                
            claim_label = f"{student_id}({purpose_tag})" if purpose_tag else str(student_id)  
            current_claimants.append(claim_label)
            new_claimed_str = ", ".join(current_claimants)
            
            is_full = len(current_claimants) >= req_num
            update_props = {
                "認領學號": {"rich_text": [{"text": {"content": new_claimed_str}}]}
            }
            if is_full:
                update_props["任務狀態"] = {"status": {"name": "被認領走了"}} # ⭐️ 改為新狀態

            client.pages.update(
                page_id=page_id,
                properties=update_props
            )
            return True, "滿團" if is_full else "未滿"
            
        except Exception as e:
            return False, str(e)

    def fetch_claimed_notion_tasks():
        """抓取 Notion 狀態為「被認領走了」或「任務完成囉」的任務，回傳待驗收列表"""
        client = get_notion_client()
        db_id = st.secrets.get("notion_db_id") or st.secrets.get("system_config", {}).get("notion_db_id")
        if not client or not db_id:
            return []
        try:
            response = client.databases.query(
                database_id=db_id,
                # ⭐️ 這裡使用 OR 邏輯，不管是在進行中還是學生標記完成了，組長都看得到！
                filter={
                    "or": [
                        {"property": "任務狀態", "status": {"equals": "被認領走了"}},
                        {"property": "任務狀態", "status": {"equals": "任務完成囉"}}
                    ]
                }
            )
            tasks = []
            for page in response.get("results", []):
                props = page.get("properties", {})
                title = props.get("任務名稱", {}).get("title", [{}])
                title_text = title[0].get("text", {}).get("content", "未命名任務") if title else "未命名任務"
                # [愛校2.0] 任務內容
                area = props.get("任務內容", {}).get("rich_text", [{}])
                area_text = area[0].get("text", {}).get("content", "未填寫") if area else "未填寫"
                date_obj = props.get("任務日期", {}).get("date", {})
                raw_date = date_obj.get("start", "未定") if date_obj else "未定"
                time_start = ""
                if raw_date != "未定":
                    try:
                        parsed_date = datetime.fromisoformat(raw_date.replace("Z", "+00:00"))
                        date_val = parsed_date.strftime("%Y-%m-%d")
                        # [愛校2.0] 若有時間資訊，保存起始時間供銷過單使用
                        if len(raw_date) > 10:
                            time_start = parsed_date.strftime("%H:%M")
                    except Exception:
                        date_val = raw_date
                else:
                    date_val = "未定"
                claimed_obj = props.get("認領學號", {}).get("rich_text", [])
                # [Fix] 合併所有 rich_text blocks，Notion 超過一定長度會自動拆成多個 block
                claimed_str = "".join(b.get("text", {}).get("content", "") for b in claimed_obj)
                claimants = [s.strip() for s in claimed_str.split(",") if s.strip()]
                tasks.append({
                    "id": page["id"], "title": title_text, "area": area_text,
                    "date": date_val, "time_start": time_start,
                    "claimants": claimants
                })
            return tasks
        except Exception as e:
            print(f"[fetch_claimed_notion_tasks] {e}")
            return []

    def update_notion_task_status(page_id, new_status):
        """將 Notion 任務的「任務狀態」改為指定值"""
        client = get_notion_client()
        if not client:
            return False
        try:
            client.pages.update(
                page_id=page_id,
                properties={"任務狀態": {"status": {"name": new_status}}}
            )
            return True
        except Exception as e:
            print(f"[update_notion_task_status] {e}")
            return False

    # ==========================================
    # SRE Utils: 重試機制
    # ==========================================
    def execute_with_retry(func, max_retries=5, base_delay=1.0, timeout=30):
        # [Patch 1] 首次呼叫零延遲；僅重試時 exponential backoff
        for attempt in range(max_retries):
            try:
                if attempt > 0:
                    # 重試才加延遲：exponential backoff + jitter
                    sleep_time = (base_delay * (2 ** (attempt - 1))) + random.uniform(0, 1)
                    time.sleep(sleep_time)
                with concurrent.futures.ThreadPoolExecutor(max_workers=1) as executor:
                    future = executor.submit(func)
                    return future.result(timeout=timeout)
            except concurrent.futures.TimeoutError:
                print(f"API Hard Timeout on attempt {attempt+1}")
                if attempt >= max_retries - 1: 
                    raise Exception("API 連線超時，請稍後再試")
            except Exception as e:
                error_str = str(e).lower()
                is_retryable = any(x in error_str for x in ['429', '500', '503', 'quota', 'rate limit', 'timed out', 'connection'])
                if is_retryable and attempt < max_retries - 1:
                    pass  # backoff 在下一輪 loop 頭部執行
                else: raise e

    # ==========================================
    # Google 連線與圖片壓縮
    # ==========================================
    @st.cache_resource
    def get_credentials():
        # [Fix #7] 改用 google-auth，取代已停止維護的 oauth2client
        # SACredentials 支援執行緒安全的自動 token 刷新
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        if "gcp_service_account" not in st.secrets:
            return None
        return SACredentials.from_service_account_info(
            dict(st.secrets["gcp_service_account"]), scopes=scope
        )

    def get_gspread_client():
        creds = get_credentials()
        # [Fix #7] gspread.Client(auth=creds) 是 google-auth 的正確用法
        return gspread.Client(auth=creds) if creds else None

    def get_drive_service():
        creds = get_credentials()
        return build('drive', 'v3', credentials=creds, cache_discovery=False) if creds else None

    @st.cache_resource
    def get_spreadsheet():
        # [Fix #9] 快取整個 Spreadsheet 物件，避免每次 get_worksheet 都重新 open_by_url
        client = get_gspread_client()
        if not client: return None
        try:
            return client.open_by_url(SHEET_URL)
        except Exception as e:
            print(f"[get_spreadsheet] 無法開啟試算表: {e}")
            return None

    def get_worksheet(tab_name):
        # [Fix #9] 改用快取的 Spreadsheet 物件，只在找不到分頁時才建立新分頁
        sheet = get_spreadsheet()
        if not sheet: return None
        for attempt in range(4):
            try:
                try:
                    return sheet.worksheet(tab_name)
                except gspread.WorksheetNotFound:
                    cols = 20 if tab_name != "appeals" else 15
                    init_rows = 2000 if tab_name == "task_queue" else 500
                    ws = sheet.add_worksheet(title=tab_name, rows=init_rows, cols=cols)
                    if tab_name == "appeals": ws.append_row(APPEAL_COLUMNS)
                    if tab_name == "service_hours": ws.append_row(["日期", "學號", "班級", "類別", "時數", "紀錄ID", "核發狀態"])  # [方案A] 加入核發狀態欄
                    if tab_name == "holidays": ws.append_row(["日期", "說明"])
                    if tab_name == "office_areas": ws.append_row(["區域名稱", "負責班級"])
                    if tab_name == "published_results": ws.append_row(["週次", "排名", "年級", "班級", "總扣分", "優良次數", "總成績", "評等", "排名模式", "發布時間"])
                    if tab_name == "task_queue": ws.append_row(["id", "task_type", "created_ts", "payload_json", "status", "attempts", "last_error"])
                    if tab_name == "student_debts": ws.append_row(["學號", "未完成時數", "備註"])  # [Fix 3] 補上備註欄
                    if tab_name == "debt_history": ws.append_row(["時間", "學號", "異動時數", "剩餘時數", "事由"])  # [新增] 愛校服務 2.0
                    if tab_name == "class_areas": ws.append_row(["班級", "外掃區域"])  # [新增] 班級 → 外掃區域 對照表
                    return ws
            except Exception as e:
                if "429" in str(e):
                    time.sleep(2 * (attempt + 1) + random.uniform(0, 1))
                    continue
                else: return None
        return None

    def compress_image_bytes(file_bytes, quality=60):
        # [V5.32] 壓縮參數調整：1200px + quality=60，減少 Drive 上傳時間
        # 對手機拍攝的現場照片而言畫質仍足夠辨認違規細節
        try:
            img = Image.open(io.BytesIO(file_bytes))
            img = ImageOps.exif_transpose(img)
            if img.mode != "RGB": img = img.convert("RGB")
            if img.width > 1200:
                ratio = 1200 / float(img.width)
                img = img.resize((1200, int(img.height * ratio)), Image.Resampling.LANCZOS)
            out_buffer = io.BytesIO()
            img.save(out_buffer, format="JPEG", quality=quality, optimize=True)
            out_buffer.seek(0)
            return out_buffer
        except Exception as e:
            print(f"[compress_image] {e}")
            return io.BytesIO(file_bytes)

    def upload_image_to_drive(file_obj, filename):
        def _upload_action():
            service = get_drive_service()
            folder_id = st.secrets["system_config"]["drive_folder_id"]
            file = service.files().create(
                body={'name': filename, 'parents': [folder_id]},
                media_body=MediaIoBaseUpload(file_obj, mimetype='image/jpeg', resumable=False), 
                fields='id', supportsAllDrives=True
            ).execute()
            # 資料夾已設定「知道連結的人都能檢視」，檔案自動繼承，不需逐檔設定權限
            return f"https://drive.google.com/thumbnail?id={file.get('id')}&sz=w1000"
        return execute_with_retry(_upload_action)

    def clean_id(val):
        # [V5.31 Patch 3] 改為保留字串型態，僅去除 Excel 尾端的 .0，保護學號前導 0
        s = str(val).strip()
        if re.fullmatch(r"\d+\.0", s):
            s = s[:-2]
        return s

    # ==========================================
    # 本機 SQLite（只保留 dedup + 監控狀態，不含 task_queue）
    # ==========================================
    def open_local_db():
        # [Fix #3-B] task_queue 已移至 Google Sheets，這裡只保留兩張輕量表
        conn = sqlite3.connect(QUEUE_DB_PATH, timeout=10.0, isolation_level=None)
        conn.execute("PRAGMA journal_mode=WAL;")
        # [Fix] 去重 key 加入 class_name，允許同學同日不同任務各別寫入
        # 先檢查舊 schema（3欄）是否存在，若是則遷移至新 schema（4欄）
        try:
            cols = [row[1] for row in conn.execute("PRAGMA table_info(service_issued)").fetchall()]
            if cols and "class_name" not in cols:
                # 舊 schema → 備份後重建
                conn.execute("ALTER TABLE service_issued RENAME TO service_issued_old")
                conn.execute("""CREATE TABLE service_issued (
                    date TEXT, sid TEXT, category TEXT, class_name TEXT,
                    PRIMARY KEY(date, sid, category, class_name))""")
                conn.execute("""INSERT OR IGNORE INTO service_issued
                    SELECT date, sid, category, '' FROM service_issued_old""")
                conn.execute("DROP TABLE service_issued_old")
                print("[db_migrate] service_issued 已從 3欄 key 遷移至 4欄 key")
            elif not cols:
                conn.execute("""CREATE TABLE IF NOT EXISTS service_issued (
                    date TEXT, sid TEXT, category TEXT, class_name TEXT,
                    PRIMARY KEY(date, sid, category, class_name))""")
        except Exception as e:
            print(f"[db_migrate] schema 遷移時發生錯誤，嘗試重建: {e}")
            try:
                conn.execute("DROP TABLE IF EXISTS service_issued")
                conn.execute("""CREATE TABLE service_issued (
                    date TEXT, sid TEXT, category TEXT, class_name TEXT,
                    PRIMARY KEY(date, sid, category, class_name))""")
            except Exception as e2:
                print(f"[db_migrate] 重建失敗: {e2}")
        conn.execute("CREATE TABLE IF NOT EXISTS system_status (key TEXT PRIMARY KEY, val TEXT)")
        # [Fix 1] 防止 campus_service_verify 重試時重複扣時：以 (task_id, sid) 為主鍵記錄已完成的 debt 操作
        conn.execute("""CREATE TABLE IF NOT EXISTS debt_processed (
            task_id TEXT, sid TEXT, PRIMARY KEY(task_id, sid))""")
        # [Patch C] 防止 update_student_debt 在 _write_history 成功、_update_debts 失敗時的重試重複寫入 debt_history
        #           _write_history 成功後立即寫入此表，重試時偵測到已寫則跳過 _write_history，直接重試 _update_debts
        conn.execute("""CREATE TABLE IF NOT EXISTS debt_history_written (
            task_id TEXT, sid TEXT, PRIMARY KEY(task_id, sid))""")
        # [Fix 4] enqueue_task_nb 的備援緩衝：當 Sheets append_row 在背景執行緒失敗時，
        #         暫存至此，Worker 下一輪會自動排空推入 Sheets
        conn.execute("""CREATE TABLE IF NOT EXISTS fallback_queue (
            task_id   TEXT PRIMARY KEY,
            task_type TEXT,
            payload_json TEXT,
            created_ts   TEXT)""")
        return conn

    def update_worker_heartbeat():
        try:
            with closing(open_local_db()) as conn:
                conn.execute("INSERT OR REPLACE INTO system_status VALUES ('worker_heartbeat', ?)", (str(time.time()),))
        except Exception as e: print(f"[heartbeat] {e}")

    def update_last_success_time():
        try:
            with closing(open_local_db()) as conn:
                conn.execute("INSERT OR REPLACE INTO system_status VALUES ('last_success_time', ?)", (str(time.time()),))
        except Exception as e: print(f"[last_success] {e}")

    def get_worker_heartbeat_sec():
        try:
            with closing(open_local_db()) as conn:
                cur = conn.cursor()
                cur.execute("SELECT val FROM system_status WHERE key='worker_heartbeat'")
                row = cur.fetchone()
                if row: return time.time() - float(row[0])
        except Exception as e: print(f"[heartbeat_sec] {e}")
        return 999999

    def get_last_success_sec():
        try:
            with closing(open_local_db()) as conn:
                cur = conn.cursor()
                cur.execute("SELECT val FROM system_status WHERE key='last_success_time'")
                row = cur.fetchone()
                if row: return time.time() - float(row[0])
        except Exception as e: print(f"[last_success_sec] {e}")
        return 999999

    # ==========================================
    # Google Sheets 雲端佇列（取代 SQLite task_queue）
    # ==========================================
    # 欄位索引（1-based）：id=1, task_type=2, created_ts=3, payload_json=4, status=5, attempts=6, last_error=7
    _QCOL_STATUS   = 5
    _QCOL_ATTEMPTS = 6
    _QCOL_ERROR    = 7

    def enqueue_task(task_type, payload):
        # [Fix #3-B] 寫入 Sheets task_queue 分頁，同步等待確認後返回
        # 成功：回傳 task_id（字串）
        # 失敗：回傳 None，呼叫端必須檢查並顯示錯誤，不可靜默假設成功
        task_id = str(uuid.uuid4())
        try:
            target_sheet_url = None
            def _action():
                nonlocal target_sheet_url
                ws = get_worksheet(SHEET_TABS["task_queue"])
                if not ws: raise Exception("無法取得 task_queue 工作表")
                # [DEBUG] 記錄實際寫入的試算表 URL，供診斷用
                try:
                    target_sheet_url = ws.spreadsheet.url
                except Exception:
                    target_sheet_url = "unknown"
                ws.append_row([
                    task_id, task_type,
                    datetime.now(timezone.utc).isoformat(),
                    json.dumps(payload, ensure_ascii=False),
                    "PENDING", 0, ""
                ], value_input_option="RAW")
            execute_with_retry(_action)
            print(f"[enqueue] 成功寫入 task_id={task_id[:8]} type={task_type} sheet={target_sheet_url}")
            return task_id  # 確認寫入成功才回傳
        except Exception as e:
            print(f"[enqueue] 加入佇列失敗: {e}")
            return None  # 失敗明確回傳 None，讓 UI 顯示錯誤

    def get_pending_count():
        try:
            ws = get_worksheet(SHEET_TABS["task_queue"])
            if not ws: return 0
            statuses = ws.col_values(_QCOL_STATUS)[1:]  # 跳過標題列
            return sum(1 for s in statuses if s in ("PENDING", "RETRY"))
        except Exception as e:
            print(f"[pending_count] {e}")
            return 0

    @safe_cached(ttl=15, default_factory=lambda: {"pending": 0, "retry": 0, "failed": 0, "oldest_pending_sec": 0, "recent_errors": []})  # 15秒快取，避免管理後台每次重繪都打 Sheets API
    def get_queue_metrics():
        metrics = {"pending": 0, "retry": 0, "failed": 0, "oldest_pending_sec": 0, "recent_errors": []}
        ws = get_worksheet(SHEET_TABS["task_queue"])
        if not ws:
            raise RuntimeError("task_queue worksheet unavailable")
        records = ws.get_all_records()
        for r in records:
            s = r.get("status", "")
            if s == "PENDING":  metrics["pending"]  += 1
            elif s == "RETRY":  metrics["retry"]    += 1
            elif s == "FAILED": metrics["failed"]   += 1
        pending_recs = [r for r in records if r.get("status") in ("PENDING", "RETRY") and r.get("created_ts")]
        if pending_recs:
            try:
                oldest = min(r["created_ts"] for r in pending_recs)
                metrics["oldest_pending_sec"] = (datetime.now(pytz.utc) - datetime.fromisoformat(oldest.replace("Z", "+00:00"))).total_seconds()
            except Exception: pass
        err_recs = sorted([r for r in records if r.get("status") in ("FAILED", "RETRY") and r.get("last_error")],
                          key=lambda x: x.get("created_ts", ""), reverse=True)[:5]
        metrics["recent_errors"] = [(r.get("last_error"), r.get("created_ts")) for r in err_recs]
        return metrics

    def fetch_next_task(max_attempts=6):
        # [Fix #3-B] 從 Sheets 讀取第一筆 PENDING 任務，標記為 IN_PROGRESS
        # 注意：background_worker 已改用 _extract_next_task(ws, records) 避免重複讀
        # 此函式保留供緊急 fallback 使用
        try:
            ws = get_worksheet(SHEET_TABS["task_queue"])
            if not ws: return None
            records = ws.get_all_records()
            return _extract_next_task(ws, records, max_attempts)
        except Exception as e:
            print(f"fetch_next_task error: {e}")
        return None

    def update_task_status(task_id, status, attempts, last_error, _row_idx=None):
        # DONE → 直接刪行，讓 task_queue 不積累；失敗/重試才保留並更新
        # [Debug Fix] delete_rows / batch_update 加上 execute_with_retry，
        # 避免單次 429/timeout 靜默失敗後任務永遠卡在 IN_PROGRESS
        try:
            ws = get_worksheet(SHEET_TABS["task_queue"])
            if not ws: return

            # 定位行號
            if _row_idx:
                ridx = _row_idx
            else:
                ids = ws.col_values(1)[1:]
                if task_id not in ids: return
                ridx = ids.index(task_id) + 2

            if status == "DONE":
                # 成功完成 → 直接刪除這行，task_queue 永遠不會積累歷史紀錄
                try:
                    execute_with_retry(lambda: ws.delete_rows(ridx))
                except Exception as e:
                    print(f"[task_queue] 刪行失敗（{task_id[:8]}，row={ridx}）: {e}")
            else:
                # FAILED 或 RETRY → 更新狀態保留供查閱
                def _status_update():
                    ws.batch_update([
                        {"range": f"E{ridx}", "values": [[status]]},
                        {"range": f"F{ridx}", "values": [[attempts]]},
                        {"range": f"G{ridx}", "values": [[str(last_error)[:200] if last_error else ""]]}
                    ])
                execute_with_retry(_status_update)
        except Exception as e:
            print(f"update_task_status error (task={task_id[:8] if task_id else '?'}, status={status}): {e}")

    # ==========================================
    # 背景處理邏輯
    # ==========================================

    def fetch_all_pending_service_tasks(max_attempts=6):
        # [Fix #3-B] 從 Sheets 批次撈出所有 service_hours_only PENDING 任務
        # 注意：此函式現在接受外部 ws + records，避免在同一 Worker 輪次重複讀 Sheets
        # 實際呼叫由 background_worker 統一管理（見下方）
        try:
            ws = get_worksheet(SHEET_TABS["task_queue"])
            if not ws: return []
            records = ws.get_all_records()
            return _extract_svc_tasks(ws, records, max_attempts)
        except Exception as e:
            print(f"[batch] 批次抓取 service_hours 任務失敗: {e}")
            return []

    def _extract_svc_tasks(ws, records, max_attempts=6):
        # 從已讀取的 records 中提取 service_hours_only 任務，避免重複讀 Sheets
        _svc_found = [r.get("id","?")[:8] for r in records if r.get("task_type")=="service_hours_only" and r.get("status") in ("PENDING","RETRY")]
        if _svc_found:
            _WORKER_LOG.append(f"[{datetime.now(TW_TZ).strftime('%H:%M:%S')}] _extract_svc_tasks: 找到 {len(_svc_found)} 筆 svc 任務: {_svc_found}")
        result = []
        batch_updates = []
        for i, r in enumerate(records):
            if (r.get("task_type") == "service_hours_only"
                    and r.get("status") in ("PENDING", "RETRY")
                    and int(r.get("attempts", 0)) < max_attempts):
                row_idx = i + 2
                att_new = int(r.get("attempts", 0)) + 1
                result.append({
                    "id":        r["id"],
                    "task_type": r["task_type"],
                    "payload":   json.loads(r.get("payload_json") or "{}"),
                    "attempts":  att_new,
                    "_row_idx":  row_idx
                })
                batch_updates.append({"range": f"E{row_idx}", "values": [["IN_PROGRESS"]]})
                batch_updates.append({"range": f"F{row_idx}", "values": [[att_new]]})
        if batch_updates:
            try:
                ws.batch_update(batch_updates)
            except Exception as e:
                print(f"[batch_update] 標記 IN_PROGRESS 失敗: {e}")
        return result

    def _extract_next_task(ws, records, max_attempts=6):
        # 從已讀取的 records 中取出第一筆非 service_hours_only 的 PENDING 任務
        _now_s = datetime.now(TW_TZ).strftime("%H:%M:%S")
        _all_statuses = [r.get("status","?") for r in records]
        _WORKER_LOG.append(f"[{_now_s}] _extract_next_task called, total_records={len(records)}, statuses={_all_statuses[:10]}")
        for i, r in enumerate(records):
            if (r.get("task_type") != "service_hours_only"
                    and r.get("status") in ("PENDING", "RETRY")
                    and int(r.get("attempts", 0)) < max_attempts):
                row_idx = i + 2
                attempts_new = int(r.get("attempts", 0)) + 1
                _WORKER_LOG.append(f"[{_now_s}] ✅ 找到任務！id={str(r.get('id',''))[:8]} type={r.get('task_type')} status={r.get('status')} row={row_idx}")
                try:
                    ws.batch_update([
                        {"range": f"E{row_idx}", "values": [["IN_PROGRESS"]]},
                        {"range": f"F{row_idx}", "values": [[attempts_new]]}
                    ])
                    _WORKER_LOG.append(f"[{_now_s}] ✅ 已標記 IN_PROGRESS row={row_idx}")
                except Exception as e:
                    _WORKER_LOG.append(f"[{_now_s}] ❌ 標記 IN_PROGRESS 失敗: {e}")
                    print(f"[fetch_next] 標記 IN_PROGRESS 失敗: {e}")
                    return None
                return {
                    "id":        r["id"],
                    "task_type": r["task_type"],
                    "payload":   json.loads(r.get("payload_json") or "{}"),
                    "attempts":  attempts_new,
                    "_row_idx":  row_idx
                }
        return None

    def process_service_tasks_batch(tasks):
        # [Fix #3] 把多個 service_hours_only 任務的所有學生合併，一次 append_rows 寫入
        # [Fix] 先用 SELECT 檢查 dedup，但 INSERT 移到 Sheets 寫入成功之後
        #       避免 Sheets 失敗 + SQLite 已標記 → 重試時被跳過的致命 bug
        is_dry_run = str(st.secrets.get("system_config", {}).get("dry_run", "false")).lower() in ["true", "1"]
        if is_dry_run:
            return True, "DRY_RUN_SUCCESS"
        rows_to_write = []
        dedup_keys = []  # [Fix] 收集 dedup keys，Sheets 成功後才寫 SQLite
        for task in tasks:
            payload = task["payload"]
            t_date     = payload.get("date", str(date.today()))
            t_cat      = payload.get("category", "")
            t_cls_name = payload.get("class_name", "")
            for sid in payload.get("student_list", []):
                # [Fix] 先 SELECT 檢查是否已發放，不做 INSERT
                is_dup = False
                try:
                    with closing(open_local_db()) as conn:
                        cur = conn.execute(
                            "SELECT 1 FROM service_issued WHERE date=? AND sid=? AND category=? AND class_name=?",
                            (t_date, str(sid), t_cat, t_cls_name)
                        )
                        is_dup = cur.fetchone() is not None
                except Exception:
                    pass  # 查詢失敗不阻擋寫入
                _dk = (t_date, str(sid), t_cat, t_cls_name)
                if not is_dup and _dk not in dedup_keys:  # [V6.2 Fix] 批內互查：同批兩個任務含同一學生時只發一次
                    rows_to_write.append([
                        t_date, str(sid),
                        t_cls_name, t_cat,
                        str(payload.get("hours", 0.5)),
                        uuid.uuid4().hex[:8],
                        ""   # [方案A] 核發狀態：空白 = 未核發
                    ])
                    dedup_keys.append(_dk)
        if not rows_to_write:
            return True, None  # 全部都是重複，視為成功
        try:
            def _batch_action():
                ws = get_worksheet(SHEET_TABS["service_hours"])
                if not ws: raise Exception("無法取得 service_hours 工作表")
                ws.append_rows(rows_to_write, value_input_option="RAW")
            execute_with_retry(_batch_action)
            # [Fix] Sheets 寫入成功後，才寫 SQLite dedup 記錄
            for key in dedup_keys:
                try:
                    with closing(open_local_db()) as conn:
                        conn.execute("INSERT OR IGNORE INTO service_issued VALUES (?, ?, ?, ?)", key)
                except Exception as e:
                    print(f"[dedup] SQLite dedup 寫入失敗（可忽略）: {e}")
            return True, None
        except Exception as e:
            return False, str(e)

    # [Patch 5] Worker 記憶體級去重快取，避免每筆任務都讀整欄紀錄ID
    @st.cache_resource
    def _get_written_ids_cache():
        return set()
    _WRITTEN_IDS_CACHE = _get_written_ids_cache()

    def _append_main_entry_row(entry):
        _rid = str(entry.get("紀錄ID", ""))
        def _action():
            ws = get_worksheet(SHEET_TABS["main"])
            if not ws: return
            # [Patch 5] 先查記憶體快取（O(1)），命中即跳過，不打 Sheets API
            if _rid and _rid in _WRITTEN_IDS_CACHE:
                print(f"[DEDUP-MEM] 紀錄ID {_rid} 在記憶體快取中，跳過寫入")
                return
            # 快取 miss → 才讀 Sheets 確認（防 Worker 重啟後重複）
            try:
                existing_ids = ws.col_values(EXPECTED_COLUMNS.index("紀錄ID") + 1)
                if _rid in existing_ids:
                    print(f"[DEDUP-SHEET] 紀錄ID {_rid} 已存在，跳過寫入")
                    if _rid: _WRITTEN_IDS_CACHE.add(_rid)
                    return
            except Exception as e:
                print(f"[DEDUP] 防重複檢查失敗，繼續寫入: {e}")
            row = [str(entry.get(col, "")).upper() if isinstance(entry.get(col, ""), bool) else str(entry.get(col, "")) for col in EXPECTED_COLUMNS]
            ws.append_row(row)
            if _rid: _WRITTEN_IDS_CACHE.add(_rid)
        execute_with_retry(_action)
    
    def _append_service_row_unique(entry):
        t_date     = str(entry.get("日期", ""))
        t_sid      = str(entry.get("學號", ""))
        t_cat      = str(entry.get("類別", ""))
        t_cls_name = str(entry.get("班級", ""))  # [Fix] 加入 class_name 至去重 key
        
        try:
            with closing(open_local_db()) as conn:
                conn.execute("INSERT INTO service_issued VALUES (?, ?, ?, ?)",
                             (t_date, t_sid, t_cat, t_cls_name))
        except sqlite3.IntegrityError:
            return
            
        def _action():
            ws = get_worksheet(SHEET_TABS["service_hours"])
            if not ws: return
            new_row = [t_date, t_sid, str(entry.get("班級", "")), t_cat, str(entry.get("時數", "")), str(entry.get("紀錄ID", "")), ""]  # [方案A] 核發狀態空白
            ws.append_row(new_row)
        execute_with_retry(_action)

    # =========================================================================
    # [Fix] 統一 tag 解析函式（全域定義，Worker 與 Admin UI 共用）
    # =========================================================================
    def _strip_notion_invisible(s):
        """剝除 Notion rich_text 可能夾帶的所有 Unicode 格式/控制字元。"""
        cleaned = ''.join(
            c for c in str(s)
            if unicodedata.category(c) not in ('Cf', 'Cc', 'Cs')
        )
        cleaned = re.sub(r'[\s\u3000\u2000-\u200a\u202f\u205f]+', ' ', cleaned)
        return cleaned.strip()

    def _parse_claimant_tag(claimant_str):
        """從認領學號字串解析出 (學號, 標籤)，支援半形/全形括號"""
        claimant_str = _strip_notion_invisible(claimant_str)
        sid_match = re.match(r"(\d+)", claimant_str)
        sid = sid_match.group(1) if sid_match else None
        tag_match = re.search(r"[\(\uff08](.*?)[\)\uff09]", claimant_str)
        if tag_match:
            tag = _strip_notion_invisible(tag_match.group(1))
        else:
            tag = "還時數"
        return sid, tag

    def _write_service_hours_direct(class_name, category, hours, student_list, svc_date):
        """直接寫入 service_hours 工作表（Worker 用，不經由佇列），含 dedup。
        [Fix] 回傳實際寫入筆數（int）：
              > 0 = 成功寫入幾筆
              = 0 = 全部被 dedup 過濾（不拋例外，由呼叫端決定如何處理）
        """
        rows = []
        dedup_keys = []
        for sid in student_list:
            is_dup = False
            try:
                with closing(open_local_db()) as conn:
                    cur = conn.execute(
                        "SELECT 1 FROM service_issued WHERE date=? AND sid=? AND category=? AND class_name=?",
                        (svc_date, str(sid), category, class_name)
                    )
                    is_dup = cur.fetchone() is not None
            except Exception:
                pass
            if is_dup:
                print(f"[service_hours] dedup 跳過 sid={sid} date={svc_date} cat={category}")
            else:
                rows.append([svc_date, str(sid), class_name, category, str(hours), uuid.uuid4().hex[:8], ""])
                dedup_keys.append((svc_date, str(sid), category, class_name))

        print(f"[service_hours] 準備寫入 {len(rows)} 筆（共{len(student_list)}人），"
              f"category={category}，date={svc_date}，class={class_name}")

        if not rows:
            print(f"[service_hours] ⚠️ 全部被 dedup 過濾，不寫入。student_list={student_list}")
            return 0  # [Fix] 改回傳 0，不再是 None，讓呼叫端可以判斷

        def _action():
            ws = get_worksheet(SHEET_TABS["service_hours"])
            if not ws: raise Exception("無法取得 service_hours 工作表")
            ws.append_rows(rows, value_input_option="RAW")

        execute_with_retry(_action)
        print(f"[service_hours] ✅ 寫入成功 {len(rows)} 筆")

        for key in dedup_keys:
            try:
                with closing(open_local_db()) as conn:
                    conn.execute("INSERT OR IGNORE INTO service_issued VALUES (?, ?, ?, ?)", key)
            except Exception:
                pass

        return len(rows)  # [Fix] 回傳實際寫入筆數


    def _mark_service_hours_issued(record_ids: list):
        """[方案A] 將 service_hours 中指定紀錄ID的「核發狀態」欄標記為「已核發」。
        record_ids：欲標記的紀錄ID清單（service_hours 第6欄）。"""
        if not record_ids:
            return
        try:
            def _action():
                ws = get_worksheet(SHEET_TABS["service_hours"])
                if not ws: raise Exception("無法取得 service_hours 工作表")
                all_vals = ws.get_all_values()
                if not all_vals:
                    return
                # 找出 header，確認欄位位置
                header = all_vals[0]
                try:
                    rid_col  = header.index("紀錄ID") + 1       # 1-based
                    stat_col = header.index("核發狀態") + 1      # 1-based
                except ValueError:
                    # header 找不到代表舊 Sheet 尚未加欄，直接用固定位置
                    rid_col  = 6
                    stat_col = 7
                # 建立 紀錄ID → row_idx 映射（跳過 header）
                id_set = set(str(i) for i in record_ids)
                batch = []
                for i, row in enumerate(all_vals[1:], start=2):  # 2-based
                    rid_val = row[rid_col - 1] if len(row) >= rid_col else ""
                    if str(rid_val).strip() in id_set:
                        batch.append({
                            "range": f"{chr(64 + stat_col)}{i}",
                            "values": [["已核發"]]
                        })
                if batch:
                    ws.batch_update(batch)
                    print(f"[mark_issued] 標記 {len(batch)} 筆已核發")
            execute_with_retry(_action)
        except Exception as e:
            print(f"[mark_issued] 標記失敗: {e}")
    def update_last_error_summary(err_msg):
        """記錄最近一次 Worker 發生的錯誤摘要至 SQLite system_status。"""
        try:
            with closing(open_local_db()) as conn:
                short_msg = str(err_msg)[:120]
                conn.execute("INSERT OR REPLACE INTO system_status VALUES ('last_error_summary', ?)", (short_msg,))
        except Exception as e:
            print(f"[error_summary] {e}")

    def get_last_error_summary():
        try:
            with closing(open_local_db()) as conn:
                cur = conn.cursor()
                cur.execute("SELECT val FROM system_status WHERE key='last_error_summary'")
                row = cur.fetchone()
                return row[0] if row else "無紀錄"
        except Exception as e:
            print(f"[error_summary_read] {e}")
            return "無紀錄"

    def _write_worker_diag(task_id, task_type, stage, detail=""):
        """[Patch 10] 改為純記憶體 + print 診斷，不再寫入 Sheets 浪費配額。"""
        now = datetime.now(TW_TZ).strftime("%H:%M:%S")
        diag_msg = f"[{now}] DIAG task={task_id[:8] if task_id else 'unk'} type={task_type} stage={stage} {detail}"
        print(diag_msg)
        _WORKER_LOG.append(diag_msg)

    def process_task(task):
        task_type, payload = task["task_type"], task["payload"]

        _task_id_diag = str(task.get("id", ""))

        is_dry_run = str(st.secrets.get("system_config", {}).get("dry_run", "false")).lower() in ["true", "1"]
        if is_dry_run:
            _write_worker_diag(_task_id_diag, task_type, "DRY_RUN")
            time.sleep(random.uniform(0.3, 0.6))
            return True, "DRY_RUN_SUCCESS"

        # [Fix #3] service_hours_only 現在由 Worker 批次處理（process_service_tasks_batch）
        # 保留此分支只作為 fallback：若任務因故繞過批次路徑，仍可單筆處理
        if task_type == "service_hours_only":
            return process_service_tasks_batch([task])

        entry = payload.get("entry", {})
        # [Fix #3-B] 照片已在 save_entry/save_appeal 同步上傳完畢，
        # entry["照片路徑"] 已包含 Drive 連結，Worker 直接寫 Sheets 即可
        try:
            if task_type in ["main_entry", "volunteer_report"]:
                _append_main_entry_row(entry)
                inspector_name = entry.get("檢查人員", "")
                # [V5.28] 根據參數決定是否發放時數 (預設發放，以防其他地方使用)
                if "學號:" in inspector_name and payload.get("award_inspector_hours", True):
                    sid = inspector_name.split("學號:")[1].strip()
                    _append_service_row_unique({"日期": entry.get("日期"), "學號": sid, "班級": "糾察隊", "類別": "整潔評分糾察", "時數": 0.25, "紀錄ID": uuid.uuid4().hex[:8]})

                if task_type == "volunteer_report":
                    # [Fix #3] volunteer_report 多名學生的時數改為批次寫入（1 次 append_rows）
                    svc_rows = []
                    t_date    = entry.get("日期", str(date.today()))
                    t_cat     = payload.get("custom_category", "晨掃志工")
                    t_cls_vol = entry.get("班級", "")  # [Fix] 加入 class_name 至去重 key
                    for sid in payload.get("student_list", []):
                        try:
                            with closing(open_local_db()) as conn:
                                conn.execute("INSERT INTO service_issued VALUES (?, ?, ?, ?)",
                                             (t_date, str(sid), t_cat, t_cls_vol))
                            svc_rows.append([
                                t_date, str(sid), t_cls_vol, t_cat,
                                str(payload.get("custom_hours", 0.5)), uuid.uuid4().hex[:8]
                            ])
                        except sqlite3.IntegrityError:
                            pass  # 已發放，跳過
                    if svc_rows:
                        def _svc_batch():
                            ws = get_worksheet(SHEET_TABS["service_hours"])
                            if not ws: raise Exception("無法取得 service_hours 工作表")
                            ws.append_rows(svc_rows, value_input_option="RAW")
                        execute_with_retry(_svc_batch)

            elif task_type == "appeal_entry":
                # [Fix #3-B] 佐證照片已在 save_appeal 同步上傳，entry["佐證照片"] 已設定
                execute_with_retry(lambda: get_worksheet(SHEET_TABS["appeals"]).append_row([str(entry.get(col, "")) for col in APPEAL_COLUMNS]))

            # ── [愛校2.0 Async] 背景驗收愛校任務 ──────────────────────
            elif task_type == "campus_service_verify":
                claimants      = payload.get("claimants", [])
                task_title     = payload.get("task_title", "")
                task_hours     = payload.get("task_hours", 1.0)
                task_date_str  = payload.get("task_date", str(date.today()))
                notion_page_id = payload.get("notion_page_id", "")
                task_area      = payload.get("task_area", "")
                time_start_v   = payload.get("time_start", "")
                # [Fix 1] 用 task_id 作為 debt_processed 的去重 key，
                #         確保重試時不會對同一學生重複扣時。
                _task_id_for_dedup = str(task.get("id", "")).strip()
                # [Patch B] task_id 為空時停用 dedup 保護（不應發生，但防禦極端情況）
                _dedup_enabled = bool(_task_id_for_dedup)

                print(f"[verify] ▶ 開始執行 task_id={_task_id_for_dedup[:8] if _task_id_for_dedup else '?'} "
                      f"claimants={claimants} area={task_area} date={task_date_str} hours={task_hours}")

                svc_debt_sids   = []  # 還時數
                svc_appeal_sids = []  # 消警告
                debt_failures   = []  # [Fix] 追蹤 update_student_debt 失敗的學號

                for clm in claimants:
                    _sid_w, _tag_w = _parse_claimant_tag(clm)
                    if not _sid_w:
                        print(f"[verify] ⚠️ 無法解析 sid，跳過: '{clm}'")
                        continue
                    _tag_n = _strip_notion_invisible(_tag_w)
                    print(f"[verify] 解析: '{clm}' → sid={_sid_w}, tag='{_tag_n}'")

                    if _tag_n in ("還時數", ""):
                        # [Fix 1] 先查 debt_processed，避免重試時重複扣時
                        _already_debted = False
                        if _dedup_enabled:
                            try:
                                with closing(open_local_db()) as _conn:
                                    _already_debted = _conn.execute(
                                        "SELECT 1 FROM debt_processed WHERE task_id=? AND sid=?",
                                        (_task_id_for_dedup, _sid_w)
                                    ).fetchone() is not None
                            except Exception as _de:
                                print(f"[verify] debt_processed 查詢失敗（將繼續執行）: {_de}")

                        if not _already_debted:
                            _debt_ok = update_student_debt(_sid_w, -task_hours, f"愛校驗收：{task_title}", _task_id=_task_id_for_dedup)
                            if _debt_ok:
                                if _dedup_enabled:
                                    try:
                                        with closing(open_local_db()) as _conn:
                                            _conn.execute(
                                                "INSERT OR IGNORE INTO debt_processed VALUES (?, ?)",
                                                (_task_id_for_dedup, _sid_w)
                                            )
                                    except Exception as _we:
                                        print(f"[verify] debt_processed 寫入失敗（可忽略）: {_we}")
                            else:
                                # [Fix] 失敗要記錄，之後拋例外讓 Worker RETRY
                                print(f"[verify] ❌ update_student_debt 失敗 sid={_sid_w}")
                                debt_failures.append(_sid_w)
                        else:
                            print(f"[verify] {_sid_w} 已在前次嘗試中完成扣時，略過（dedup）")

                        svc_debt_sids.append(_sid_w)
                        time.sleep(0.3)

                    elif _tag_n == "消警告":
                        svc_appeal_sids.append(_sid_w)

                    elif _tag_n == "糾察懲罰":
                        pass  # 不發時數、不消警告

                    else:
                        # 未知 tag → 視為還時數（同樣套用 dedup 保護）
                        print(f"[worker verify] 未知 tag '{_tag_w}' for {_sid_w}，視為還時數")
                        _already_debted = False
                        if _dedup_enabled:
                            try:
                                with closing(open_local_db()) as _conn:
                                    _already_debted = _conn.execute(
                                        "SELECT 1 FROM debt_processed WHERE task_id=? AND sid=?",
                                        (_task_id_for_dedup, _sid_w)
                                    ).fetchone() is not None
                            except Exception:
                                pass

                        if not _already_debted:
                            _debt_ok = update_student_debt(_sid_w, -task_hours, f"愛校驗收：{task_title}", _task_id=_task_id_for_dedup)
                            if _debt_ok:
                                if _dedup_enabled:
                                    try:
                                        with closing(open_local_db()) as _conn:
                                            _conn.execute(
                                                "INSERT OR IGNORE INTO debt_processed VALUES (?, ?)",
                                                (_task_id_for_dedup, _sid_w)
                                            )
                                    except Exception:
                                        pass
                            else:
                                print(f"[verify] ❌ update_student_debt 失敗（未知tag）sid={_sid_w}")
                                debt_failures.append(_sid_w)
                        else:
                            print(f"[verify] {_sid_w}（未知tag）已在前次嘗試中完成扣時，略過")

                        svc_debt_sids.append(_sid_w)
                        time.sleep(0.3)

                # [DIAG v6] 迴圈結束，印出解析結果
                _WORKER_LOG.append(
                    f"[{datetime.now(TW_TZ).strftime('%H:%M:%S')}] "
                    f"PARSE_DONE: debt_sids={svc_debt_sids} appeal_sids={svc_appeal_sids} "
                    f"punish={_punish_students if '_punish_students' in dir() else 'N/A'} "
                    f"debt_failures={debt_failures}"
                )

                # [Fix] 若有 update_student_debt 失敗，先拋例外讓 Worker RETRY，
                #       不繼續執行 service_hours 寫入，避免資料半成功
                if debt_failures:
                    raise Exception(f"[verify] update_student_debt 失敗，待 RETRY：sids={debt_failures}")

                # 寫入 service_hours：還時數（直接寫，已有 SQLite dedup 保護）
                if svc_debt_sids:
                    _written_debt = _write_service_hours_direct(
                        "愛校打掃", "返校打掃(補打掃)", task_hours, svc_debt_sids,
                        task_date_str if task_date_str != "未定" else str(date.today())
                    )
                    # [Fix] dedup 空轉（0筆）印出警告，但不 RETRY
                    # （service_issued 有記錄代表上一輪已成功寫入，不需重試）
                    _WORKER_LOG.append(f"[{datetime.now(TW_TZ).strftime('%H:%M:%S')}] _write_service_hours_direct(debt) 回傳: {_written_debt} 筆")
                    if _written_debt == 0:
                        print(f"[verify] ⚠️ 還時數 service_hours 寫入 0 筆（dedup 判定已存在），不重試。sids={svc_debt_sids}")

                # [Fix 2] 消警告：直接呼叫 _write_service_hours_direct（有 execute_with_retry + SQLite dedup）
                # ※ Worker 背景執行緒內同步寫入是安全的，不影響 UI
                # ※ 之前改寫 SQLite fallback_queue 的做法在 Streamlit Cloud 不可靠（ephemeral storage）
                if svc_appeal_sids:
                    _cf = f"{task_area}|{time_start_v}" if time_start_v else task_area
                    _ad = task_date_str if task_date_str != "未定" else str(date.today())
                    print(f"[worker] 消警告 service_hours 直接寫入，共 {len(svc_appeal_sids)} 人：{svc_appeal_sids}")
                    _written_appeal = _write_service_hours_direct(_cf, "愛校服務(消警告)", task_hours, svc_appeal_sids, _ad)
                    _WORKER_LOG.append(f"[{datetime.now(TW_TZ).strftime('%H:%M:%S')}] _write_service_hours_direct(appeal) 回傳: {_written_appeal} 筆")
                    if _written_appeal == 0:
                        print(f"[verify] ⚠️ 消警告 service_hours 寫入 0 筆（dedup 判定已存在），不重試。sids={svc_appeal_sids}")

                # 更新 Notion 狀態
                if notion_page_id:
                    _notion_ok = update_notion_task_status(notion_page_id, "任務已驗收")
                    _WORKER_LOG.append(f"[{datetime.now(TW_TZ).strftime('%H:%M:%S')}] Notion 更新: {'✅成功' if _notion_ok else '❌失敗'}")
                    print(f"[worker] Notion 狀態更新：{'成功' if _notion_ok else '失敗（已記錄，不影響任務完成）'} "
                          f"page_id={notion_page_id[:8] if notion_page_id else 'None'}")
                else:
                    print("[worker] notion_page_id 為空，跳過 Notion 更新")
                _done_diag = f"debt={svc_debt_sids}|appeal={svc_appeal_sids}"
                _write_worker_diag(_task_id_for_dedup, "campus_service_verify", "VERIFY_DONE", _done_diag)
                print(f"[worker] ✅ campus_service_verify 完成：debt={svc_debt_sids} appeal={svc_appeal_sids}")

            # ── [愛校2.0 Async] 背景晨掃審核 ──────────────────────────
            elif task_type == "morning_sweep_approve":
                _ms_rid      = payload.get("record_id", "")
                _ms_action   = payload.get("action", "approve")  # approve / reject
                _ms_score    = payload.get("score_val", 0)
                _ms_new_item = payload.get("new_item", "晨間打掃(學期加分)")
                _ms_new_note = payload.get("new_note", "")

                def _do_ms():
                    ws = get_worksheet(SHEET_TABS["main"])
                    if not ws: raise Exception("無法取得 main 工作表")
                    id_list = [str(v).strip() for v in ws.col_values(EXPECTED_COLUMNS.index("紀錄ID") + 1)]
                    rid_str = str(_ms_rid).strip()
                    if rid_str not in id_list:
                        print(f"[worker morning] record {rid_str} not found, skip")
                        return
                    ridx = id_list.index(rid_str) + 1
                    # 檢查是否已被審核（避免重複處理）
                    current_item = ws.cell(ridx, EXPECTED_COLUMNS.index("評分項目") + 1).value
                    if current_item and ("學期加分" in str(current_item) or "已駁回" in str(current_item)):
                        print(f"[worker morning] record {rid_str} 已審核 ({current_item}), skip")
                        return
                    if _ms_action == "approve":
                        ws.update_cell(ridx, EXPECTED_COLUMNS.index("晨間打掃原始分") + 1, _ms_score)
                    ws.update_cell(ridx, EXPECTED_COLUMNS.index("評分項目") + 1, _ms_new_item)
                    ws.update_cell(ridx, EXPECTED_COLUMNS.index("備註") + 1, _ms_new_note)
                execute_with_retry(_do_ms)
                try: load_main_data.clear()
                except Exception: pass
                print(f"[worker] morning_sweep_{_ms_action} 完成：{_ms_rid}")

            # ── [愛校2.0 Async] 背景申訴審核 ──────────────────────────
            elif task_type == "appeal_review":
                _ar_record_id = payload.get("record_id", "")
                _ar_status    = payload.get("status", "")
                _ar_reply     = payload.get("reply_text", "")

                def _do_ar():
                    ws_appeals = get_worksheet(SHEET_TABS["appeals"])
                    ws_main    = get_worksheet(SHEET_TABS["main"])
                    if not ws_appeals: raise Exception("無法取得 appeals 工作表")
                    data = ws_appeals.get_all_records()
                    t_row = next((i + 2 for i, r in enumerate(data)
                                  if str(r.get("對應紀錄ID")) == str(_ar_record_id)
                                  and str(r.get("處理狀態")) == "待處理"), None)
                    if not t_row:
                        print(f"[worker appeal] 找不到 record {_ar_record_id} 或已處理")
                        return
                    ws_appeals.update_cell(t_row, APPEAL_COLUMNS.index("處理狀態") + 1, _ar_status)
                    if "審核回覆" in APPEAL_COLUMNS:
                        ws_appeals.update_cell(t_row, APPEAL_COLUMNS.index("審核回覆") + 1, _ar_reply)
                    if _ar_status == "已核可" and ws_main:
                        m_data = ws_main.get_all_records()
                        m_row = next((j + 2 for j, mr in enumerate(m_data)
                                      if str(mr.get("紀錄ID")) == str(_ar_record_id)), None)
                        if m_row:
                            ws_main.update_cell(m_row, EXPECTED_COLUMNS.index("修正") + 1, "TRUE")
                execute_with_retry(_do_ar)
                try:
                    load_main_data.clear()
                    load_appeals.clear()
                except Exception: pass
                print(f"[worker] appeal_review 完成：{_ar_record_id} → {_ar_status}")

            elif task_type == "revoke_record":
                # [V6.1] 組長直接撤分：將主紀錄標記為修正(TRUE)，排名計算即自動排除
                _rv_id = payload.get("record_id", "")
                _rv_reason = payload.get("reason", "")

                def _do_rv():
                    ws_main = get_worksheet(SHEET_TABS["main"])
                    if not ws_main: raise Exception("無法取得 main 工作表")
                    m_data = ws_main.get_all_records()
                    m_row = next((j + 2 for j, mr in enumerate(m_data)
                                  if str(mr.get("紀錄ID")) == str(_rv_id)), None)
                    if not m_row:
                        print(f"[worker revoke] 找不到紀錄 {_rv_id}，可能已被處理")
                        return
                    ws_main.update_cell(m_row, EXPECTED_COLUMNS.index("修正") + 1, "TRUE")
                    try:
                        _old_note = str(m_data[m_row - 2].get("備註", "") or "")
                        _new_note = (_old_note + " | " if _old_note else "") + f"【組長撤分】{_rv_reason}"
                        ws_main.update_cell(m_row, EXPECTED_COLUMNS.index("備註") + 1, _new_note)
                    except Exception as _e:
                        print(f"[worker revoke] 備註更新失敗（撤分本身已完成）: {_e}")
                execute_with_retry(_do_rv)
                try: load_main_data.clear()
                except Exception: pass
                print(f"[worker] revoke_record 完成：{_rv_id}")

            return True, None
        except Exception as e: return False, str(e)

    def background_worker(stop_event=None):
        try: add_script_run_ctx(threading.current_thread(), get_script_run_ctx())
        except Exception: pass  # Streamlit context 在背景執行緒可能不存在，忽略

        # [V5.32] Stuck task recovery：追蹤連續空轉次數，每 60 輪（約 5 分鐘）掃一次卡住任務
        _idle_loops = 0
        STUCK_THRESHOLD_SEC = 300  # IN_PROGRESS 超過 5 分鐘視為卡住

        def _recover_stuck_tasks(ws, records):
            """將超過 5 分鐘仍為 IN_PROGRESS 的任務重置為 RETRY，並清理 DIAG/DONE 歷史行"""
            now_utc = datetime.now(pytz.utc)
            batch_updates = []
            recovered = 0
            # [Patch 8] 同時收集需要刪除的 DIAG 行（從後往前刪，避免行號偏移）
            rows_to_delete = []
            for i, r in enumerate(records):
                row_idx = i + 2
                status = r.get("status", "")
                # 清理 DIAG 行和已完成的 DONE 行（不應該殘留）
                if status in ("DIAG", "DONE"):
                    rows_to_delete.append(row_idx)
                    continue
                if status != "IN_PROGRESS": continue
                ts_raw = r.get("created_ts", "")
                if not ts_raw: continue
                try:
                    created = datetime.fromisoformat(ts_raw.replace("Z", "+00:00"))
                    if (now_utc - created).total_seconds() > STUCK_THRESHOLD_SEC:
                        attempts_new = int(r.get("attempts", 0))
                        batch_updates.append({"range": f"E{row_idx}", "values": [["RETRY"]]})
                        batch_updates.append({"range": f"F{row_idx}", "values": [[attempts_new]]})
                        batch_updates.append({"range": f"G{row_idx}", "values": [["[AUTO_RECOVERY] IN_PROGRESS timeout"]]})
                        recovered += 1
                except Exception: continue
            if batch_updates:
                try:
                    ws.batch_update(batch_updates)
                    print(f"[recovery] 重置了 {recovered} 個卡住的 IN_PROGRESS 任務")
                except Exception as e:
                    print(f"[recovery] batch_update 失敗: {e}")
            # 從後往前刪除 DIAG/DONE 行，避免行號偏移
            if rows_to_delete:
                try:
                    for ridx in sorted(rows_to_delete, reverse=True):
                        ws.delete_rows(ridx)
                    print(f"[cleanup] 清理了 {len(rows_to_delete)} 個 DIAG/DONE 殘留行")
                except Exception as e:
                    print(f"[cleanup] 清理 DIAG 行失敗（可忽略）: {e}")
            return recovered

        _WORKER_LOG.append(f"[BW] Worker 執行緒啟動")
        while True:
            if stop_event and stop_event.is_set(): break
            try: update_worker_heartbeat()
            except Exception as e:
                print(f"[worker] heartbeat 寫入失敗: {e}")
            try:
                # [Fix #3-C] 每輪只讀一次 Sheets task_queue，避免 rate limit
                ws = get_worksheet(SHEET_TABS["task_queue"])
                if not ws:
                    time.sleep(5.0)
                    continue

                # [Fix 4 / Patch A] 排空 fallback_queue：將 enqueue_task_nb 背景寫入失敗的任務推入 Sheets
                # [Patch A] 改用 append_rows 一次批次寫入，避免逐筆 append_row 造成 429
                try:
                    with closing(open_local_db()) as _fb_conn:
                        _fb_rows = _fb_conn.execute(
                            "SELECT task_id, task_type, payload_json, created_ts FROM fallback_queue LIMIT 20"
                        ).fetchall()
                    if _fb_rows:
                        _batch_rows = [
                            [r[0], r[1], r[3], r[2], "PENDING", 0, ""]
                            for r in _fb_rows
                        ]
                        try:
                            ws.append_rows(_batch_rows, value_input_option="RAW")
                            # 成功後才從 SQLite 刪除
                            with closing(open_local_db()) as _fb_conn:
                                _fb_conn.executemany(
                                    "DELETE FROM fallback_queue WHERE task_id=?",
                                    [(r[0],) for r in _fb_rows]
                                )
                            print(f"[drain_fallback] 已批次推送 {len(_fb_rows)} 筆 fallback 任務至 Sheets")
                        except Exception as _fe:
                            print(f"[drain_fallback] append_rows 失敗，下輪再試: {_fe}")
                except Exception as _drain_e:
                    print(f"[drain_fallback] 排空 fallback_queue 失敗（忽略）: {_drain_e}")

                try:
                    records = ws.get_all_records()
                    _pending_count_log = sum(1 for r in records if r.get("status") in ("PENDING","RETRY"))
                    _WORKER_LOG.append(f"[{datetime.now(TW_TZ).strftime('%H:%M:%S')}] POLL ok, records={len(records)}, pending/retry={_pending_count_log}")
                except Exception as e:
                    err_str = str(e)
                    _WORKER_LOG.append(f"[{datetime.now(TW_TZ).strftime('%H:%M:%S')}] get_all_records FAILED: {str(e)[:80]}")
                    print(f"[worker] get_all_records 失敗: {e}")
                    # [配額修復 v2] 429 → 等 180 秒（3 分鐘），確保配額完全恢復後再試
                    # Google 配額是以「每分鐘」為單位重置，等 3 分鐘有足夠緩衝
                    if "429" in err_str:
                        time.sleep(180.0)
                    else:
                        time.sleep(10.0)
                    continue

                # [V5.32] 每 60 輪空轉或每次有資料時，檢查並回收卡住的任務
                _idle_loops += 1
                if _idle_loops >= 60:
                    _idle_loops = 0
                    _recover_stuck_tasks(ws, records)
                    # [V5.33] 移除 stuck recovery 後的重複 get_all_records，節省配額
                    # recover 只修改 status 欄，不影響後續的 _extract 邏輯判斷

                # 優先批次清空所有 service_hours_only 任務
                svc_tasks = _extract_svc_tasks(ws, records)
                if svc_tasks:
                    _idle_loops = 0
                    ok, err = process_service_tasks_batch(svc_tasks)
                    if ok: update_last_success_time()
                    else:
                        if err and "DRY_RUN" not in err: update_last_error_summary(err)
                    final_status = "DONE" if ok else "FAILED"
                    for t in svc_tasks:
                        update_task_status(t["id"], final_status, t["attempts"], err, _row_idx=t.get("_row_idx"))
                    # [Patch 9] 尖峰時段也縮短批次後等待
                    time.sleep(0.5 if _is_peak_hour(datetime.now(TW_TZ)) else 2.0)
                    continue

                # 處理含照片的任務（同一批 records，不重複讀）
                task = _extract_next_task(ws, records)
                if not task:
                    # [Patch 7] 尖峰時段縮短空轉等待，讓新任務更快被接手
                    _idle_sleep = 5.0 if _is_peak_hour(datetime.now(TW_TZ)) else 30.0
                    time.sleep(_idle_sleep)
                    continue

                _idle_loops = 0
                # [DIAG v3] 寫入 module-level log，確保不依賴任何 API
                _now_str = datetime.now(TW_TZ).strftime("%H:%M:%S")
                _tid_short = str(task.get('id',''))[:8]
                _task_type_log = task.get("task_type","?")
                _payload_log = str(task.get('payload', {}))[:200]
                _WORKER_LOG.append(f"[{_now_str}] BEFORE_PROCESS task_id={_tid_short} type={_task_type_log}")
                _WORKER_LOG.append(f"[{_now_str}] PAYLOAD={_payload_log}")
                # 特別把 claimants 完整列出，不截斷
                _claimants_full = task.get('payload', {}).get('claimants', 'KEY_NOT_FOUND')
                _WORKER_LOG.append(f"[{_now_str}] CLAIMANTS_FULL={_claimants_full}")
                ok, err = process_task(task)
                _WORKER_LOG.append(f"[{datetime.now(TW_TZ).strftime('%H:%M:%S')}] AFTER_PROCESS ok={ok} err={str(err)[:100]}")
                if ok: update_last_success_time()
                else:
                    if err and "DRY_RUN" not in err: update_last_error_summary(err)

                # [DIAG → print] 不再寫入 Sheets，節省尖峰時段配額
                print(f"[DIAG] task_id={str(task.get('id',''))[:8]} type={task.get('task_type','?')} ok={ok} err={str(err)[:150]}")
                if not ok and err and "FILE_NOT_FOUND" in str(err): task["attempts"] = 999
                update_task_status(task["id"], "DONE" if ok else ("FAILED" if task["attempts"] >= 6 else "RETRY"), task["attempts"], err, _row_idx=task.get("_row_idx"))
                # [Patch 4] 尖峰時段縮短等待，加快佇列消化速度
                _worker_sleep = 0.5 if _is_peak_hour(datetime.now(TW_TZ)) else 2.0
                time.sleep(_worker_sleep)
            except Exception as e:
                print(f"[worker] 未預期例外: {e}")
                time.sleep(5.0)

    @st.cache_resource
    def _get_worker_state():
        """儲存 Worker 的 stop_event 與 thread 引用，跨 rerun 不被清空。"""
        return {"stop_event": None, "thread": None, "started_at": None}

    def _start_fresh_worker():
        """啟動新 Worker，記錄到 _get_worker_state()。先等舊 thread 結束再啟動新的。"""
        state = _get_worker_state()
        old_t = state.get("thread")
        # 送停止訊號給舊 Worker
        if state.get("stop_event") is not None:
            state["stop_event"].set()
            _WORKER_LOG.append(f"[{datetime.now(TW_TZ).strftime('%H:%M:%S')}] 已送出停止訊號給舊 Worker")
        # 等舊 thread 真的結束（最多等 35 秒，確保它從 sleep 醒來並看到 stop_event）
        if old_t is not None and old_t.is_alive():
            _WORKER_LOG.append(f"[{datetime.now(TW_TZ).strftime('%H:%M:%S')}] 等待舊 Worker 結束...")
            old_t.join(timeout=35)
            _WORKER_LOG.append(f"[{datetime.now(TW_TZ).strftime('%H:%M:%S')}] 舊 Worker 已結束: {not old_t.is_alive()}")
        # 啟動新 Worker
        stop_event = threading.Event()
        t = threading.Thread(target=background_worker, args=(stop_event,), daemon=True)
        try:
            add_script_run_ctx(t)
        except Exception:
            pass
        t.start()
        state["stop_event"] = stop_event
        state["thread"] = t
        state["started_at"] = datetime.now(TW_TZ).strftime("%H:%M:%S")
        _WORKER_LOG.clear()
        _WORKER_LOG.append(f"[{state['started_at']}] ✅ 新 Worker 啟動，is_alive={t.is_alive()}")
        return stop_event

    @st.cache_resource
    def ensure_worker_started():
        _WORKER_LOG.append(f"[{datetime.now(TW_TZ).strftime('%H:%M:%S')}] ensure_worker_started 被呼叫，sys_env={sys_env}")
        if sys_env == "DEV":
            _WORKER_LOG.append("[ensure] DEV 環境，Worker 停用")
            return threading.Event()
        # 若已有 alive thread（例如之前的呼叫已啟動），不重複啟動，避免競爭
        state = _get_worker_state()
        t = state.get("thread")
        if t is not None and t.is_alive():
            _WORKER_LOG.append(f"[ensure] Worker 已存活（{t.name}），跳過重複啟動")
            return state.get("stop_event", threading.Event())
        _WORKER_LOG.append("[ensure] 啟動全新 Worker...")
        return _start_fresh_worker()
    _ = ensure_worker_started()

    # ==========================================
    # 前端資料讀取 
    # ==========================================
    @safe_cached(ttl=21600, default_factory=list)
    def load_holidays():
        ws = get_worksheet(SHEET_TABS["holidays"])
        if not ws:
            raise RuntimeError("holidays worksheet unavailable")
        # ws.get_all_records() 失敗會直接 raise，safe_cached 會接住、不快取失敗結果
        return [pd.to_datetime(str(r.get("日期", "")).strip()).date() for r in ws.get_all_records() if str(r.get("日期", "")).strip()]

    def is_within_appeal_period(violation_date, appeal_days=3):
        vd = pd.to_datetime(violation_date).date() if isinstance(violation_date, str) else violation_date
        holidays, today, current_date, workdays = load_holidays(), date.today(), vd, 0
        for _ in range(14): 
            if workdays >= appeal_days: break
            current_date += timedelta(days=1)
            if current_date.weekday() < 5 and current_date not in holidays: workdays += 1
        return today <= current_date

    @safe_cached(ttl=300, default_factory=lambda: pd.DataFrame(columns=EXPECTED_COLUMNS))
    def load_main_data():
        # 讀取整學期 main_data，統一快取一份。
        # 需要近兩週過濾的地方在 UI 層自己用 df[df["週次"] >= now_week-2] 處理。
        # [V5.34] 改用 safe_cached：基礎建設失敗 (worksheet None / API 失敗) 會 raise 不被快取
        ws = get_worksheet(SHEET_TABS["main"])
        if not ws:
            raise RuntimeError("main worksheet unavailable")
        df = pd.DataFrame(ws.get_all_records())  # API 失敗會 raise，由 safe_cached 接手
        if df.empty:
            # 合法的空狀態（新學期、剛開站），允許快取
            return pd.DataFrame(columns=EXPECTED_COLUMNS)
        if "班級" in df.columns: df["班級"] = df["班級"].astype(str).str.strip()
        for col in EXPECTED_COLUMNS:
            if col not in df.columns: df[col] = ""
        if "紀錄ID" not in df.columns: df["紀錄ID"] = df.index.astype(str)
        for col in ["內掃原始分", "外掃原始分", "垃圾原始分", "垃圾內掃原始分", "垃圾外掃原始分", "晨間打掃原始分", "手機人數", "週次"]:
            if col in df.columns: df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
        if "修正" in df.columns: df["修正"] = df["修正"].astype(str).apply(lambda x: True if x.upper() == "TRUE" else False)
        return df[EXPECTED_COLUMNS]

    @safe_cached(ttl=21600, default_factory=dict)
    def load_roster_dict():
        ws = get_worksheet(SHEET_TABS["roster"])
        if not ws:
            raise RuntimeError("roster worksheet unavailable")
        df = pd.DataFrame(ws.get_all_records())
        id_c, cls_c = next((c for c in df.columns if "學號" in c), None), next((c for c in df.columns if "班級" in c), None)
        return {clean_id(row[id_c]): str(row[cls_c]).strip() for _, row in df.iterrows()} if id_c and cls_c else {}

    def parse_roster_upload(uploaded_file):
        """[V6.4 資安] 解析名冊檔為 {學號: [學號,班級,座號,姓名]}。
        只存在當次瀏覽階段記憶體（session_state），關閉頁面即消失，絕不寫入任何雲端。"""
        _sheets_all = pd.read_excel(uploaded_file, sheet_name=None, dtype=str)
        _fix0 = lambda s: s[:-2] if str(s).endswith(".0") else str(s)
        out = {}
        for _sn, _sdf in _sheets_all.items():
            _sdf.columns = [str(c) for c in _sdf.columns]
            _cid = next((c for c in _sdf.columns if "學號" in c), None)
            _ccl = next((c for c in _sdf.columns if "班級" in c), None)
            _cse = next((c for c in _sdf.columns if "座號" in c), None)
            _cnm = next((c for c in _sdf.columns if "姓名" in c), None)
            if not _cid or not _cnm: continue
            for _, _rr in _sdf.iterrows():
                _s = clean_id(_rr.get(_cid, ""))
                if not _s or len(_s) < 4: continue
                out[_s] = [_s,
                           _fix0(str(_rr.get(_ccl, "") or "").strip()) if _ccl else "",
                           _fix0(str(_rr.get(_cse, "") or "").strip()) if _cse else "",
                           str(_rr.get(_cnm, "") or "").strip()]
        return out

    def load_roster_name_map():
        """[V6.3] 學號→姓名對照（roster 分頁需含姓名欄；查無回空 dict，銷過單姓名留白手寫）"""
        try:
            ws = get_worksheet(SHEET_TABS["roster"])
            df = pd.DataFrame(ws.get_all_records()) if ws else pd.DataFrame()
            id_c = next((c for c in df.columns if "學號" in c), None)
            nm_c = next((c for c in df.columns if "姓名" in c), None)
            if not id_c or not nm_c: return {}
            return {clean_id(r[id_c]): str(r[nm_c]).strip() for _, r in df.iterrows()}
        except Exception:
            return {}
    
    @safe_cached(ttl=3600, default_factory=lambda: ([], []))
    def load_sorted_classes():
        ws = get_worksheet(SHEET_TABS["roster"])
        if not ws:
            raise RuntimeError("roster worksheet unavailable")
        records = ws.get_all_records()
        if not records:
            all_vals = ws.get_all_values()
            if len(all_vals) > 1: records = [dict(zip(all_vals[0], row)) for row in all_vals[1:]]
        df = pd.DataFrame(records)
        class_col = next((c for c in df.columns if "班級" in str(c).strip()), None)
        if not class_col: return [], []
        unique = [c for c in df[class_col].astype(str).str.strip().unique().tolist() if c]
        dept_order = {"商": 1, "英": 2, "資": 3, "家": 4, "服": 5}
        cls_order  = {"甲": 1, "乙": 2, "丙": 3, "丁": 4}
        def get_sort_key(n):
            g   = 1 if "一" in n or "1" in n else (2 if "二" in n or "2" in n else (3 if "三" in n or "3" in n else 99))
            dep = next((v for k, v in dept_order.items() if k in n), 99)
            cls = next((v for k, v in cls_order.items()  if k in n), 99)
            return (g, dep, cls)
        sorted_all = sorted(unique, key=get_sort_key)
        return sorted_all, [{"grade": f"{get_sort_key(c)[0]}年級" if get_sort_key(c)[0]!=99 else "其他", "name": c} for c in sorted_all]

    @safe_cached(ttl=300, default_factory=lambda: (pd.DataFrame(), "error"))   # [效能] 5分鐘，尖峰時段不必每分鐘重打
    def get_daily_duty(target_date):
        ws = get_worksheet(SHEET_TABS["duty"])
        if not ws:
            raise RuntimeError("duty worksheet unavailable")
        df = pd.DataFrame(ws.get_all_records())
        if df.empty: return pd.DataFrame(), "no_data"
        date_col = next((c for c in df.columns if "日期" in c), None)
        if date_col:
            df[date_col] = pd.to_datetime(df[date_col], errors='coerce').dt.date
            return df[df[date_col] == (target_date if isinstance(target_date, date) else target_date.date())], "success"
        return pd.DataFrame(), "missing_cols"

    @safe_cached(ttl=3600, default_factory=dict)
    def load_office_area_map():
        ws = get_worksheet(SHEET_TABS["office_areas"])
        if not ws:
            raise RuntimeError("office_areas worksheet unavailable")
        return {str(r.get("區域名稱", "")).strip(): str(r.get("負責班級", "")).strip() for r in ws.get_all_records() if str(r.get("區域名稱", "")).strip()}

    # [新增] 班級 → 外掃區域 對照表 (cache 1 小時，學期內幾乎不會改)
    @safe_cached(ttl=3600, default_factory=dict)
    def load_class_outer_area_map():
        """回傳 {班級: 外掃區域} 字典。空字串或缺欄都會被略過。"""
        ws = get_worksheet(SHEET_TABS["class_areas"])
        if not ws:
            raise RuntimeError("class_areas worksheet unavailable")
        records = ws.get_all_records()
        result = {}
        for r in records:
            cls = str(r.get("班級", "")).strip()
            area = str(r.get("外掃區域", "")).strip()
            if cls and area:
                result[cls] = area
        return result

    @safe_cached(ttl=21600, default_factory=lambda: {"semester_start": "2025-08-25", "standard_n": 4})
    def load_settings():
        ws = get_worksheet(SHEET_TABS["settings"])
        if not ws:
            raise RuntimeError("settings worksheet unavailable")
        config = {"semester_start": "2025-08-25", "standard_n": 4}
        for row in ws.get_all_values():
            if len(row)>=2: config[row[0]] = int(row[1]) if row[0] == "standard_n" else row[1]
        return config

    def save_setting(key, val):
        ws = get_worksheet(SHEET_TABS["settings"])
        if ws:
            try:
                cell = ws.find(key)
                if cell: ws.update_cell(cell.row, cell.col+1, val)
                else: ws.append_row([key, val])
                # [V5.32] 只清設定相關快取，不 clear() 所有快取
                # 避免在尖峰時段管理員操作設定時引發 load_main_data 雪崩重讀
                load_settings.clear()
                return True
            except Exception as e:
                print(f"[save_setting] {e}")
                return False
        return False

    @safe_cached(ttl=300, default_factory=lambda: pd.DataFrame(columns=APPEAL_COLUMNS))   # [效能] 5分鐘，申訴資料不需秒級更新
    def load_appeals():
        ws = get_worksheet(SHEET_TABS["appeals"])
        if not ws:
            raise RuntimeError("appeals worksheet unavailable")
        df = pd.DataFrame(ws.get_all_records())
        for col in APPEAL_COLUMNS:
            if col not in df.columns: df[col] = "待處理" if col == "處理狀態" else ""
        return df[APPEAL_COLUMNS]

    # [新增] 愛校服務 2.0：欠時資料存取函式 =====================
    def load_student_debts():
        """讀取 student_debts 工作表，回傳 {學號: 未完成時數} 字典。
        同一學號若有多列（例如多次欠時分批登錄），時數自動加總，不會被後列覆蓋。"""
        ws = get_worksheet(SHEET_TABS["student_debts"])
        if not ws:
            return {}
        try:
            records = ws.get_all_records()
            if not records:
                return {}
            result = {}
            for r in records:
                sid = clean_id(str(r.get("學號", "")).strip())
                if not sid:
                    continue
                try:
                    hours = float(r.get("未完成時數", 0))
                except (ValueError, TypeError):
                    hours = 0.0
                # [Fix] 加總，避免重複學號後列覆蓋前列
                result[sid] = round(result.get(sid, 0.0) + hours, 2)
            return result
        except Exception as e:
            print(f"[load_student_debts] {e}")
            return {}

    def load_debt_history(sid):
        """讀取 debt_history 工作表，回傳該學號的歷史異動 DataFrame"""
        ws = get_worksheet(SHEET_TABS["debt_history"])
        cols = ["時間", "學號", "異動時數", "剩餘時數", "事由"]
        if not ws:
            return pd.DataFrame(columns=cols)
        try:
            df = pd.DataFrame(ws.get_all_records())
            if df.empty:
                return pd.DataFrame(columns=cols)
            for c in cols:
                if c not in df.columns:
                    df[c] = ""
            return df[df["學號"].astype(str).str.strip() == str(sid).strip()][cols].reset_index(drop=True)
        except Exception as e:
            print(f"[load_debt_history] {e}")
            return pd.DataFrame(columns=cols)

    def load_student_debt_note(sid):
        """讀取 student_debts 工作表的備註欄位。
        同一學號若有多列，將所有備註合併顯示（以 ；分隔，去重）。"""
        ws = get_worksheet(SHEET_TABS["student_debts"])
        if not ws:
            return ""
        try:
            records = ws.get_all_records()
            notes = []
            for r in records:
                if clean_id(str(r.get("學號", ""))) == clean_id(str(sid)):
                    note = str(r.get("備註", "")).strip()
                    # [Fix] 收集全部備註並去重，不只抓第一筆
                    if note and note not in notes:
                        notes.append(note)
            return "；".join(notes) if notes else ""
        except Exception as e:
            print(f"[load_student_debt_note] {e}")
            return ""

    def update_student_debt(sid, change_hours, reason, _task_id=""):
        """寫入 debt_history 一筆紀錄並同步更新 student_debts 中的未完成時數。
        [Fix] 不再合併多列！每列欠時保留各自的備註。
        正數 = 新增欠時（append 新列）。
        負數 = 扣減欠時（FIFO，從最舊的列開始扣，扣完的列才刪除）。
        [Patch C] _task_id：傳入 campus_service_verify 的 task id，用於 debt_history_written dedup，
                  防止 _write_history 成功但 _update_debts 失敗重試時重複寫入 debt_history。"""
        sid = str(sid).strip()
        _dedup_id = str(_task_id).strip()  # 空字串代表停用 history dedup（不由 task 驅動的直接呼叫）
        try:
            debts = load_student_debts()
            current = debts.get(sid, 0.0)
            new_remaining = round(current + change_hours, 2)
            now_str = datetime.now(TW_TZ).strftime("%Y-%m-%d %H:%M:%S")

            # [Patch C] 檢查 _write_history 是否已在前次嘗試中完成，避免重試時重複寫入 debt_history
            _history_already_written = False
            if _dedup_id:
                try:
                    with closing(open_local_db()) as _dc:
                        _history_already_written = _dc.execute(
                            "SELECT 1 FROM debt_history_written WHERE task_id=? AND sid=?",
                            (_dedup_id, sid)
                        ).fetchone() is not None
                except Exception as _dce:
                    print(f"[debt_history dedup] 查詢失敗（繼續執行）: {_dce}")

            def _write_history():
                ws_h = get_worksheet(SHEET_TABS["debt_history"])
                if not ws_h:
                    raise Exception("無法取得 debt_history 工作表")
                ws_h.append_row([now_str, sid, change_hours, new_remaining, reason],
                                value_input_option="RAW")

            if not _history_already_written:
                execute_with_retry(_write_history)
                # [Patch C] _write_history 成功後立即寫入 dedup 標記
                if _dedup_id:
                    try:
                        with closing(open_local_db()) as _dc:
                            _dc.execute(
                                "INSERT OR IGNORE INTO debt_history_written VALUES (?, ?)",
                                (_dedup_id, sid)
                            )
                    except Exception as _dwe:
                        print(f"[debt_history dedup] 標記寫入失敗（可忽略）: {_dwe}")
            else:
                print(f"[debt_history dedup] {sid} 已寫入 debt_history（task={_dedup_id[:8]}），略過重複寫入")
            def _update_debts():
                ws_d = get_worksheet(SHEET_TABS["student_debts"])
                if not ws_d:
                    raise Exception("無法取得 student_debts 工作表")

                if change_hours >= 0:
                    # [Fix 3] 新增欠時 → append 新列，第三欄寫入 reason 作為備註，保留既有列不動
                    ws_d.append_row([sid, change_hours, reason], value_input_option="RAW")
                else:
                    # [Fix] FIFO 扣減：從最舊（最上面）的列開始扣
                    all_vals = ws_d.get_all_values()
                    remaining_deduction = abs(change_hours)
                    updates = []   # [(row_idx_1based, new_hours)]
                    deletes = []   # [row_idx_1based]

                    for i, row in enumerate(all_vals):
                        if i == 0:  # skip header
                            continue
                        if clean_id(str(row[0]).strip()) != sid:
                            continue
                        if remaining_deduction <= 0:
                            break
                        try:
                            row_hours = float(row[1])
                        except (ValueError, TypeError, IndexError):
                            row_hours = 0.0
                        if row_hours <= 0:
                            continue

                        row_idx = i + 1  # 1-based for Sheets API
                        if remaining_deduction >= row_hours:
                            # 此列完全扣完 → 標記刪除
                            remaining_deduction = round(remaining_deduction - row_hours, 2)
                            deletes.append(row_idx)
                        else:
                            # 此列部分扣減 → 更新剩餘時數
                            new_val = round(row_hours - remaining_deduction, 2)
                            remaining_deduction = 0
                            updates.append((row_idx, new_val))

                    # 先更新部分扣減的列
                    for row_idx, new_val in updates:
                        ws_d.update_cell(row_idx, 2, new_val)

                    # 從後往前刪除完全扣完的列（避免行號偏移）
                    for row_idx in sorted(deletes, reverse=True):
                        ws_d.delete_rows(row_idx)
                        print(f"[update_debt] FIFO 刪除 row {row_idx}（{sid} 該列欠時已歸零）")

            execute_with_retry(_update_debts)
            return True
        except Exception as e:
            print(f"[update_student_debt] {e}")
            return False

    # ── 愛校服務申請單 Excel 模板（base64 嵌入，避免外部檔案依賴）──
    _APPEAL_FORM_TEMPLATE_B64 = (
        "UEsDBBQABgAIAAAAIQBBN4LPbgEAAAQFAAATAAgCW0NvbnRlbnRfVHlwZXNdLnhtbCCiBAIooAACAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACsVMluwjAQvVfqP0S+Vomhh6qqCBy6HFsk6AeYeJJY"
        "JLblGSj8fSdmUVWxCMElUWzPWybzPBit2iZZQkDjbC76WU8kYAunja1y8T39SJ9FgqSsVo2zkIs1oBgN"
        "7+8G07UHTLjaYi5qIv8iJRY1tAoz58HyTulCq4g/QyW9KuaqAvnY6z3JwlkCSyl1GGI4eINSLRpK3le8"
        "vFEyM1Ykr5tzHVUulPeNKRSxULm0+h9J6srSFKBdsWgZOkMfQGmsAahtMh8MM4YJELExFPIgZ4AGLyPd"
        "usq4MgrD2nh8YOtHGLqd4662dV/8O4LRkIxVoE/Vsne5auSPC/OZc/PsNMilrYktylpl7E73Cf54GGV8"
        "9W8spPMXgc/oIJ4xkPF5vYQIc4YQad0A3rrtEfQcc60C6Anx9FY3F/AX+5QOjtQ4OI+c2gCXd2EXka46"
        "9QwEgQzsQ3Jo2PaMHPmr2w7dnaJBH+CW8Q4b/gIAAP//AwBQSwMEFAAGAAgAAAAhALVVMCP0AAAATAIA"
        "AAsACAJfcmVscy8ucmVscyCiBAIooAACAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACskk1P"
        "wzAMhu9I/IfI99XdkBBCS3dBSLshVH6ASdwPtY2jJBvdvyccEFQagwNHf71+/Mrb3TyN6sgh9uI0rIsS"
        "FDsjtnethpf6cXUHKiZylkZxrOHEEXbV9dX2mUdKeSh2vY8qq7iooUvJ3yNG0/FEsRDPLlcaCROlHIYW"
        "PZmBWsZNWd5i+K4B1UJT7a2GsLc3oOqTz5t/15am6Q0/iDlM7NKZFchzYmfZrnzIbCH1+RpVU2g5abBi"
        "nnI6InlfZGzA80SbvxP9fC1OnMhSIjQS+DLPR8cloPV/WrQ08cudecQ3CcOryPDJgosfqN4BAAD//wMA"
        "UEsDBBQABgAIAAAAIQBvQjIgSQMAAAgHAAAPAAAAeGwvd29ya2Jvb2sueG1spFVdb9s2FH0fsP9A8F0W"
        "aX3YEeIUthxjAdoh2Nb2caAlOiIiiRpJfwRF3wpsQFFgA9oCRQdsfenLsL503/s7s9P9i11KtpPUL1kq"
        "2KSoax+ee8+51P6tRZGjGVdayLKHaYtgxMtEpqI86eG7X4ycLkbasDJluSx5D59xjW8dfPzR/lyq07GU"
        "pwgASt3DmTFV5Lo6yXjBdEtWvITIRKqCGViqE1dXirNUZ5ybInfbhIRuwUSJG4RIXQdDTiYi4UOZTAte"
        "mgZE8ZwZoK8zUekNWpFcB65g6nRaOYksKoAYi1yYsxoUoyKJjk5Kqdg4h7QXNEALBZ8QvpTA0N7sBKGd"
        "rQqRKKnlxLQA2m1I7+RPiUvplRIsdmtwPSTfVXwmrIZbViq8IatwixVegFHywWgUrFV7JYLi3RAt2HJr"
        "44P9icj5vca6iFXVp6ywSuUY5Uybw1QYnvZwB5Zyzq88UNNqMBU5RNudPa+N3YOtnY8VLED7fm64Kpnh"
        "sSwNWG1N/UNtVWPHmQQTo8/4V1OhOPQOWAjSgZElERvrY2YyNFV5UyQNXZW2UpnoVi5mvFVy4wZhm/iE"
        "8m4w5gkJqLt6/vU/f/8GfvKXP/+x/POX5V+v3dU337179fL86Q/nvz7699nvy2evXdpdPXq5+vHV8tsn"
        "y+dv3EsWZrv98j9MzBJbQxfq1uTW3L9fQ0hRRRujHhuF4P5oeBvE+pzNQDowCKRZd/YRaNP98sHIi/f6"
        "oTdwBv2YOj4Zxs5g2N1zuqNu4NGR1+n344eQhQqjRLKpydZ2sJg97IP2O6E7bLGJUBJNRXqx/wOyvhw7"
        "vzdsYg9tpvbguyf4XF8Yxy7R4r4oUznvYYeSAKOzzRIym9eR+yI1GZyzfmjboXn2CRcnGdClQdCFP0F3"
        "WFo9fIXOsKEzgsuxwxU67iU+9fkKvOoZlXVPNKKvvn+yfPzi/Onbdz89BvXhXLdHsa00xUhFdkt1lNJa"
        "yQ1KwvLkWCE72R8SG+QLc1ubegaXCiA6pP4e8Q77jufFvuN3Rh1QiASO53f8OPAHh5R0rEz2TREt8nky"
        "u1n/t31389qJLx/Za9VtD1nwaP0+Q5qbdcjmaB0K3JuxzmCLdvAfAAAA//8DAFBLAwQUAAYACAAAACEA"
        "gT6Ul/MAAAC6AgAAGgAIAXhsL19yZWxzL3dvcmtib29rLnhtbC5yZWxzIKIEASigAAEAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAArFJNS8QwEL0L/ocwd5t2FRHZdC8i7FXrDwjJtCnbJiEzfvTfGyq6XVjWSy8Db4Z5783Hdvc1"
        "DuIDE/XBK6iKEgR6E2zvOwVvzfPNAwhi7a0egkcFExLs6uur7QsOmnMTuT6SyCyeFDjm+CglGYejpiJE"
        "9LnShjRqzjB1Mmpz0B3KTVney7TkgPqEU+ytgrS3tyCaKWbl/7lD2/YGn4J5H9HzGQlJPA15ANHo1CEr"
        "+MFF9gjyvPxmTXnOa8Gj+gzlHKtLHqo1PXyGdCCHyEcffymSc+WimbtV7+F0QvvKKb/b8izL9O9m5MnH"
        "1d8AAAD//wMAUEsDBBQABgAIAAAAIQDo3dDimgYAAOkZAAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDEu"
        "eG1snJTbbtswDIbvB+wdDN3Hp5yNOMWwImtvhmJpt2tFpmMhluVJymnD3n2UYjsFMgxug0SUZerTT4rM"
        "4u4kSu8ASnNZpSTyQ+JBxWTGq21KXp5XgxnxtKFVRktZQUrOoMnd8uOHxVGqnS4AjIeESqekMKZOgkCz"
        "AgTVvqyhwje5VIIafFTbQNcKaOY2iTKIw3ASCMorciEkqg9D5jlncC/ZXkBlLhAFJTWoXxe81i1NsD44"
        "QdVuXw+YFDUiNrzk5uygxBMsedxWUtFNiXGfohFl3knhN8bfsD3Grd+cJDhTUsvc+EgOLppvw58H84Cy"
        "jnQbfy9MNAoUHLi9wCsqfp+kaNyx4its+E7YpIPZdKlkz7OU/A6bzwBtZIfwOrTv/pDlwtXJk1ouarqF"
        "NZiX+kl5OTfP8gkXsFZJsFwEnVfGsSBsEjwFeUo+RcmXOLQuzuM7h6N+NfcM3ayhBGYANUXE+yWlWDNq"
        "r3qKPdA9frX1W14WbclvpNxZ2CNuC61KB7HHUmb4AT5Did4PY+yan04ITjuddmOr+bWilWsSDC+DnO5L"
        "800eH4BvC4PSxv4UYa7Mkux8D5ph3ePZfuzATJYYF46e4LaBsW7pydkjz0yBs5k/isfTWYT+2MlnGyF6"
        "sb02UvxofKzEjoE37xhoW0bkT6NwPrRKeiHwvh0CbYOY/4uwAW1W3Eb5X0Gjhoa2ocXDtwaFyp0itC0j"
        "9uPZOBpPeidm0jDQtomJ/WgU9idMGwLaa2rfqAL/kC9XPBnOrhmZ3wqxZefK4y8AAAD//wAAAP//rJjt"
        "btowFIZvBeUCCg7hqwKkJXa820AMrf2xdmpYt939jjmOfT4sTUH0T6unxwce27Ff2A8vl8vVnq6n4/7j"
        "/ffs41CZajb8PL0N8Nfzppq9XA/Vcve0Wy3Iz7KanX8N1/cfXy+v30MFDPpjmtP5+dtfexnOlzdgi6d6"
        "VR3359D0S+gKZatqBv8ZAH8eF/v553E/P8eSNpXMI+kUsYo4RXpFPCVz0EyudcF1slpocqiaanzbrQSd"
        "BFYCJ0EvgSeAKcBSyOWq109gPHGBQh9YIHiZtEBmJVYo1pik2pVGNXyUVaMcEroXzJIP6rEkT6ongPk3"
        "2n9123lTJyA0OlQ74r8T+lhhFnSKDK/pYg282zyNYj5SybhhXCTwKw8ST0dfqlnzzr5UU6caNm0w+XLb"
        "wLRN3zahT5i1tPkl6BCsU4WVwCFYpooeQZ2AJ4BZrB9kEfowCwk6BMRCAoeAWCAgFgQwCzhlH7IWoQ+z"
        "QJDfQYeAWEjgEBAL2cMTwCy2D7IIfZgFAmKBgFhI4BAQC9nDE8As4Pl/yFqEPswCAbFAQCwkcAiIhezh"
        "CWAW4Yx6iMatEfOIhIhEQkwUcZEQF9XHU8JtCpHkrsPKYBAgp1Uk1AZrqI0kLo6iNlhDHnbamdsUQsd9"
        "NhgJqA0SaoOE2kjiwp0Pe5XayD4+1tw6c5tS/thAXJyeEPGGh+MlXYL57rqFyNaMJTkhSmJVjVOkj2Sb"
        "LxhKuF8pX4RgOzkA4+VMYqKRpFPEKuIU6RXxlHCbwrVfm7tWC6/kMG9puUSGa0PaD5ky76xOI6uR06jX"
        "yDPENQu54E5NvLMNXBdJU+Tb1sQaGh1Fcu5yzbh17YhoaxHoXKlmI5Jy6eW3Ihiyl+dTVQgfTXNHEDQy"
        "ObSR5B3fKWLVKKdIr0Z5SrhNIYRsJ8jA4oyfgkUQ+M+nW7yIwzTnBC8WoTWpKB9fClld5TTqNfIMsVmp"
        "C2nA3LPGt0bwQNMz2oj92OaipKmR1chp1GvkGULNef4S4x8AAAD//wAAAP//bJJRbsIwDIavEuUAoym0"
        "pRFFYulgPEyaxC5QwG2jlSRKzabt9DMItof5Lc4X+/9tZ3GC2IGBYRjFwZ8dVjJN5XLxey0itJVcqVJv"
        "VCkn/0ma6E2aMORR5dqonCUFkYIha8rZsDlrVRDhcoxKda1Szpua6ic1ZXWmVI0jK0WArZURyNgJ0AA4"
        "DZLgFWZUacZkmELXbIdzXc+596WuuZ0YldBEuJ0YpYhw/ZmZrllPma65rk2u6+tuJ38/aLkIvXeA9vAa"
        "Resdbo+VzKXArwCVdN549wFxtN5dmglNBy9N7KwbxQAt/bzkoZAi2q6/n9GHyy2V2HtEf7oFPTRHiJcg"
        "k6Tj8R7cqu4Az0GEJkDc2W+SLqUYD81ApzkptBbf/DPcdKTw0YLDBslXJYOPGBuLZERbsh+3x+u4Jp8+"
        "vo89AC5/AAAA//8DAFBLAwQUAAYACAAAACEA6aYluGYGAABTGwAAEwAAAHhsL3RoZW1lL3RoZW1lMS54"
        "bWzsWc1uGzcQvhfoOxB7TyzZkmIZkQNLluI2cWLYSoocqV1qlxF3uSApO7oVybFAgaJp0UuB3noo2gZI"
        "gF7Sp3Gbok2BvEKH5EpaWlRsJwb6Fx1sLffj/M9whrp67UHK0CERkvKsFVQvVwJEspBHNItbwZ1+79J6"
        "gKTCWYQZz0grmBAZXNt8/72reEMlJCUI9mdyA7eCRKl8Y2VFhrCM5WWekwzeDblIsYJHEa9EAh8B3ZSt"
        "rFYqjZUU0yxAGU6B7O3hkIYE9TXJYHNKvMvgMVNSL4RMHGjSxNlhsNGoqhFyIjtMoEPMWgHwifhRnzxQ"
        "AWJYKnjRCirmE6xsXl3BG8UmppbsLe3rmU+xr9gQjVYNTxEPZkyrvVrzyvaMvgEwtYjrdrudbnVGzwBw"
        "GIKmVpYyzVpvvdqe0iyB7NdF2p1KvVJz8SX6awsyN9vtdr1ZyGKJGpD9WlvAr1cata1VB29AFl9fwNfa"
        "W51Ow8EbkMU3FvC9K81GzcUbUMJoNlpAa4f2egX1GWTI2Y4Xvg7w9UoBn6MgGmbRpVkMeaaWxVqK73PR"
        "A4AGMqxohtQkJ0McQhR3cDoQFGsGeIPg0hu7FMqFJc0LyVDQXLWCD3MMGTGn9+r596+eP0Wvnj85fvjs"
        "+OFPx48eHT/80dJyNu7gLC5vfPntZ39+/TH64+k3Lx9/4cfLMv7XHz755efP/UDIoLlEL7588tuzJy++"
        "+vT37x574FsCD8rwPk2JRLfIEdrnKehmDONKTgbifDv6CabODpwAbQ/prkoc4K0JZj5cm7jGuyugePiA"
        "18f3HVkPEjFW1MP5RpI6wF3OWZsLrwFuaF4lC/fHWexnLsZl3D7Ghz7eHZw5ru2Oc6ia06B0bN9JiCPm"
        "HsOZwjHJiEL6HR8R4tHuHqWOXXdpKLjkQ4XuUdTG1GuSPh04gTTftENT8MvEpzO42rHN7l3U5syn9TY5"
        "dJGQEJh5hO8T5pjxOh4rnPpI9nHKyga/iVXiE/JgIsIyrisVeDomjKNuRKT07bktQN+S029gqFdet++y"
        "SeoihaIjH82bmPMycpuPOglOc6/MNEvK2A/kCEIUoz2ufPBd7maIfgY/4Gypu+9S4rj79EJwh8aOSPMA"
        "0W/GoqjaTv1Nafa6YswoVON3xXh6Om3B0eRLiZ0TJXgZ7l9YeLfxONsjEOuLB8+7uvuu7gb/+bq7LJfP"
        "Wm3nBRaa5HlfbLrkdGmTPKSMHagJIzel6ZMlHBZRDxZNA2+muNnQlCfwtSjuDi4W2OxBgquPqEoOEpxD"
        "j101I18sC9KxRDmXMNuZZTN8khO0zThJoc02k2Fdzwy2Hkisdnlkl9fKs+GMjJkUYzN/ThmtaQJnZbZ2"
        "5e2YVa1US83mqlY1oplS56g2Uxl8uKgaLM6sCV0Igt4FrNyAEV3LDrMJZiTSdrdz89QtmvWFukgmOCKF"
        "j7Teiz6qGidNY2UaRh4f6TnvFB+VuDU12bfgdhYnldnVlrCbeu9tvDQdbude0nl7Ih1ZVk5OlqGjVtCs"
        "r9YDFOK8FQxhrIWvaQ5el7rxwyyGu6FQCRv2pyazCde5N5v+sKzCTYW1+4LCTh3IhVTbWCY2NMyrIgRY"
        "ZoZwI/9qHcx6UQrYSH8DKdbWIRj+NinAjq5ryXBIQlV2dmnF3FEYQFFK+VgRcZBER2jAxmIfg/t1qII+"
        "EZVwO2Eqgn6AqzRtbfPKLc5F0pUvsAzOrmOWJ7gotzpFp5ls4SaPZzKYJyutEQ9088pulDu/KiblL0iV"
        "chj/z1TR5wlcF6xF2gMh3OQKjHS+tgIuVMKhCuUJDXsCLrlM7YBogetYeA1BBffJ5r8gh/q/zTlLw6Q1"
        "TH1qn8ZIUDiPVCII2YOyZKLvFGLV4uyyJFlByERUSVyZW7EH5JCwvq6BDX22ByiBUDfVpCgDBncy/tzn"
        "IoMGsW5y/qmdj03m87YHujuwLZbdf8ZepFYq+qWjoOk9+0xPNSsHrznYz3nU2oq1oPFq/cxHbQ6XPkj/"
        "gfOPipDZHyf0gdrn+1BbEfzWYNsrBFF9yTYeSBdIWx4H0DjZRRtMmpRtWIru9sLbKLiRLjrdGV/I0jfp"
        "dM9p7Flz5rJzcvH13ef5jF1Y2LF1udP1mBqS9mSK6vZoOsgYx5hftco/PPHBfXD0Nlzxj5mS9mr/AVzx"
        "wZRhfySA5LfONVs3/wIAAP//AwBQSwMEFAAGAAgAAAAhAHhwoFUfBAAAMBYAAA0AAAB4bC9zdHlsZXMu"
        "eG1s5FjNiuNGEL4H8g6i7xr9WPLaRtKyXo9gYQOBmUCubbllN9vqFq32rLwhkGNOOYWQfYCFQHLYQyB5"
        "oDBJ3iLV+rE1rD1ee73JmDAwVrdaX33VVV1dVcHjMmPGDZEFFTxEzoWNDMITMaN8HqIvrmNzgIxCYT7D"
        "THASohUp0OPo00+CQq0YuVoQogyA4EWIFkrlI8sqkgXJcHEhcsLhTSpkhhUM5dwqcknwrNAfZcxybbtv"
        "ZZhyVCOMsuR9QDIsXyxzMxFZjhWdUkbVqsJCRpaMns25kHjKgGrpeDgxSqcvXaOUrZBq9h05GU2kKESq"
        "LgDXEmlKE/Iu3aE1tHCyQQLk45Ac37LdO7qX8kgkz5LkhmrzoShIBVeFkYglVyEaAlG9BaMXXLzksX4F"
        "Fm5WRUHxyrjBDGYcZEVBIpiQhgLTwc5VMxxnpF5x+8PbP399e/vjd3///L1em+KMslX9ztUTlcmbxRkF"
        "A+hJS5OpKR0k7KfXt29+2yLJq2gusCzA5Wrmvf4eQR0dTgRra4knhx1+FNTBxrByPg1RHPdt/fevqHCP"
        "z/R2WfIAP6r87uR2cE+1N5XzF+D9lLH1gXT12YOJKIDIpYjkMQyM5vl6lcPJ4xBka5+u1u1ZPZd45bh+"
        "5wOrEhgFUyFnENTbUKAl11NRwEiqwIclnS/0rxI5/J8KpSDwRcGM4rngmOkD3H7R/RIuA4j7IVILiNtt"
        "2MBLJZqoYWn4Bn3v2opDRWHvUqDZsty7tlZmuy6NUmCahDB2pZX5Ml3vk46OZWrwZRZn6tksRHAd6jDW"
        "PoJRmsd6T+qB3qsuWo3dgQW3OgbXKNO1gF2sHCC4ndX6awPnOVvp+N/YaBdWbwcWyGiZ3MWqR+PK1zR2"
        "PX7C6JxnpBYXBbgdGgsh6SugoW+dBN4TuJQh9VA06c68lDi/JmVL1irT3Xu3S/v3ZgymO3RnP+JuaEe6"
        "T133IRhoH8nz86Kz2NZH/7/DeYpwss9bPziCnF/MO5zxWcQl45CL48gjrxO2zUX3H158584fugpb05aH"
        "m2p4hyZae47ZJvGB1BYd5Lz+iamcPi87Mgd4MMfrXPhXZQcUGp1q5k4ts65KDF0qh+iP37/569tfOun8"
        "dEmZonxLHQOYs3JTGVWFsdKNtapmWkuBAmlGUrxk6nr9MkSb58/IjC4zaEU1qz6nN0JVECHaPD/XxahT"
        "dXMg839eQAUJv8ZS0hB9dTl+NJxcxq45sMcD0+sR3xz644npe0/Hk0k8tF376ded9t4HNPeqbiSUBI43"
        "Khi0AGWjbEP+ajMXos6gpl+V4UC7y33o9u0nvmObcc92TK+PB+ag3/PN2HfcSd8bX/qx3+HuH9kEtC3H"
        "qduJmrw/UjQjjPLWVq2FurNgJBjeo4TVWsLatHqjfwAAAP//AwBQSwMEFAAGAAgAAAAhAMJWolR1AgAA"
        "dgUAABQAAAB4bC9zaGFyZWRTdHJpbmdzLnhtbKSUz1PaQBTH78zwP2Ryag812IPTdgIenOlMbz20fwAD"
        "UZiBDSWhU28pYqVSQGsoRTID6FCxCGL5FQX0jyG7G078C32AtRV6Mjlkf7y3331v32eXX/0QDDDvhbDk"
        "F5GTXV5ysIyAPKLXjzac7Ns3L588YxlJdiOvOyAiwcluChK76rLbeEmSGViLJCfrk+XQC46TPD4h6JaW"
        "xJCAwLIuhoNuGYbhDU4KhQW3V/IJghwMcE8djhUu6PYjlvGIESTDvs9ZJoL87yLC2t2Ei5f8Ll52kdIW"
        "1r5iPUqrCUOv4eMjXO/gzCdc06laGO12Rx9TQyVFYnlSLBEtiRO5oZKmatOsJsxShedkF89NpGZyuFs2"
        "BhrePsH1ywXbXtysxG+VclGS0Rc8zk+HikIKZfjTxmDSLy46NVJYr84vhWDNvkrbsVGmO28z9J7R65FM"
        "Dte/L2xZ0yGlmce8DTL+j9osAeMqQdWL6YqQD2on+z2vw8y6iORXXie7wjLyZggKisQ1Ed0CwHL3juru"
        "IEbfDuw2s901b2JW9GgyRXZ+2W20f0XyRZItE61gtz3Cly2ixWH42JJ46xxqZ7eRv3V7YNoAFtTVzBWs"
        "hINPDkAE7yUt5ZSuTThrHVqKRK8CI6PTLFE7cPaNAa0WreiN+3GincHNArrMUh6wBqZps0fbF+bOT5zI"
        "UPXIPB6M+1/w1rZxnSeduFn7gfd3jcENVSswb+hJfJ39c3OjOJ7FycYMh3H/kDT2sZZg4AMwJg3AMW2y"
        "5XH/syWemxUSS0+2+OetmD0UMAnxzOJklhlyVhppyj23KV5gwo00IDZUog+LhIN30/UbAAD//wMAUEsD"
        "BBQABgAIAAAAIQA7bTJLwQAAAEIBAAAjAAAAeGwvd29ya3NoZWV0cy9fcmVscy9zaGVldDEueG1sLnJl"
        "bHOEj8GKwjAURfcD/kN4e5PWhQxDUzciuFXnA2L62gbbl5D3FP17sxxlwOXlcM/lNpv7PKkbZg6RLNS6"
        "AoXkYxdosPB72i2/QbE46twUCS08kGHTLr6aA05OSonHkFgVC7GFUST9GMN+xNmxjgmpkD7m2UmJeTDJ"
        "+Ysb0Kyqam3yXwe0L0617yzkfVeDOj1SWf7sjn0fPG6jv85I8s+ESTmQYD6iSDnIRe3ygGJB63f2nmt9"
        "DgSmbczL8/YJAAD//wMAUEsDBBQABgAIAAAAIQA/mcwOIwEAAOQGAAAnAAAAeGwvcHJpbnRlclNldHRp"
        "bmdzL3ByaW50ZXJTZXR0aW5nczEuYmlucmNwY1Bg8GYwZTBkMGIoALLTGPIZioC0I5CXCmQXMwSARUoY"
        "dBnCGDyBUIHBGajelAEEGFmY2e4wcLAx/29gZ2TgZJjFbcKRwsDIwM8QwcTEwAQkmYE8RwYTsGrqEIxQ"
        "Y0A0DP8HAnTTXTz9QpUYIhi3MIcwOc2uPYDPdm64JAvUVIjZVHT2qFGDPARg6YoYZ0YAFQf7hniB1Aow"
        "eJCidTQd4AmBECYGBreACDcGBo4cZPYfMYim//XImlFKAma0SACaBM7KJhyz4LkbLAaRGI0GpBBwRC+f"
        "mRlYgdJM4PIVWBIyMjGkMDAxWupBSn1wwIIKdkQBDDEsBWYmC5wFUaPHQB4E2sHIieRQRoYTQNcwTEON"
        "PaAihABGXgQAAAD//wMAUEsDBBQABgAIAAAAIQCRHyoEaAEAAJgCAAARAAgBZG9jUHJvcHMvY29yZS54"
        "bWwgogQBKKAAAQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACMkk1OwzAUhPdI3CHyPnV+lKqyklQC1BWVKlEEYmfZryUi"
        "cSzbkOYk7FghdlyA8wDnwEmakAILlsnMfJp5cjzfFbnzAEpnpUiQP/GQA4KVPBPbBF2uF+4MOdpQwWle"
        "CkhQDRrN0+OjmEnCSgUrVUpQJgPtWJLQhMkE3RojCcaa3UJB9cQ6hBU3pSqosZ9qiyVld3QLOPC8KS7A"
        "UE4NxQ3QlQMR7ZGcDUh5r/IWwBmGHAoQRmN/4uNvrwFV6D8DrTJyFpmppd20rztmc9aJg3uns8FYVdWk"
        "Ctsatr+Pr5fnF+1UNxPNrRigNOaMMAXUlCpt9st6l8d49LM5YE61WdpbbzLgJ3X6/vb68fLsfD49xvi3"
        "2gdWKhMGeBp4wdT1Qjf0115A/IhEwc2Q6022Rru66wLcsTtIt7pXrsLTs/UCHfB8EkXEiyzvR77Z1QGL"
        "fe9/Em3DKQlnI2IPSNvSh28p/QIAAP//AwBQSwMEFAAGAAgAAAAhAA1ANB6tAQAADAMAABAACAFkb2NQ"
        "cm9wcy9hcHAueG1sIKIEASigAAEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAnJLBahsxEIbvhb7DonusdRpCMFqFkjTk"
        "0FKDndxV7awtKkuLNFnsnnsLvYVAaA7NoVAo9NJDT32b2s1jdHaXOOskp95m5h9+ffolsT+f2aSCEI13"
        "Gev3UpaA0z43bpKxk/HR1h5LIiqXK+sdZGwBke3L58/EMPgSAhqICVm4mLEpYjngPOopzFTskexIKXyY"
        "KaQ2TLgvCqPh0OuzGTjk22m6y2GO4HLIt8q1IWsdBxX+r2nudc0XT8eLkoCleFmW1miFdEv5xujgoy8w"
        "eTXXYAXvioLoRqDPgsGFTAXvtmKklYUDMpaFshEEvx+IY1B1aENlQpSiwkEFGn1IovlAsW2z5J2KUONk"
        "rFLBKIeEVa+1TVPbMmKQy19f//y+vr35Jjjp7awpu6vd2uzIfrNAxeZibdBykLBJODZoIb4thirgE8D9"
        "LnDD0OK2OKuPn1dfblbXn5bnV38vft5+P19e/niE2wRABz846rVx7+NJOfaHCuEuyc2hGE1VgJzCXye9"
        "HohjCjHY2uRgqtwE8rudx0L97qft55b93V76IqUn7cwEv//G8h8AAAD//wMAUEsBAi0AFAAGAAgAAAAh"
        "AEE3gs9uAQAABAUAABMAAAAAAAAAAAAAAAAAAAAAAFtDb250ZW50X1R5cGVzXS54bWxQSwECLQAUAAYA"
        "CAAAACEAtVUwI/QAAABMAgAACwAAAAAAAAAAAAAAAACnAwAAX3JlbHMvLnJlbHNQSwECLQAUAAYACAAA"
        "ACEAb0IyIEkDAAAIBwAADwAAAAAAAAAAAAAAAADMBgAAeGwvd29ya2Jvb2sueG1sUEsBAi0AFAAGAAgA"
        "AAAhAIE+lJfzAAAAugIAABoAAAAAAAAAAAAAAAAAQgoAAHhsL19yZWxzL3dvcmtib29rLnhtbC5yZWxz"
        "UEsBAi0AFAAGAAgAAAAhAOjd0OKaBgAA6RkAABgAAAAAAAAAAAAAAAAAdQwAAHhsL3dvcmtzaGVldHMv"
        "c2hlZXQxLnhtbFBLAQItABQABgAIAAAAIQDppiW4ZgYAAFMbAAATAAAAAAAAAAAAAAAAAEUTAAB4bC90"
        "aGVtZS90aGVtZTEueG1sUEsBAi0AFAAGAAgAAAAhAHhwoFUfBAAAMBYAAA0AAAAAAAAAAAAAAAAA3BkA"
        "AHhsL3N0eWxlcy54bWxQSwECLQAUAAYACAAAACEAwlaiVHUCAAB2BQAAFAAAAAAAAAAAAAAAAAAmHgAA"
        "eGwvc2hhcmVkU3RyaW5ncy54bWxQSwECLQAUAAYACAAAACEAO20yS8EAAABCAQAAIwAAAAAAAAAAAAAA"
        "AADNIAAAeGwvd29ya3NoZWV0cy9fcmVscy9zaGVldDEueG1sLnJlbHNQSwECLQAUAAYACAAAACEAP5nM"
        "DiMBAADkBgAAJwAAAAAAAAAAAAAAAADPIQAAeGwvcHJpbnRlclNldHRpbmdzL3ByaW50ZXJTZXR0aW5n"
        "czEuYmluUEsBAi0AFAAGAAgAAAAhAJEfKgRoAQAAmAIAABEAAAAAAAAAAAAAAAAANyMAAGRvY1Byb3Bz"
        "L2NvcmUueG1sUEsBAi0AFAAGAAgAAAAhAA1ANB6tAQAADAMAABAAAAAAAAAAAAAAAAAA1iUAAGRvY1By"
        "b3BzL2FwcC54bWxQSwUGAAAAAAwADAAmAwAAuSgAAAAA"
    )

    def generate_appeal_form_excel(student_id, cls_name, records, student_name=""):
        """[愛校2.0] 生成消警告申請單 Excel，直接套用學校官方模板，100% 保留原版格式。
        [V6.3] 姓名自動帶入、時數拆成每列 1 小時、時間欄置中換行、簽核區改四欄（移除主任教官）。"""
        import base64
        from copy import copy as _cpy
        from openpyxl import load_workbook

        # 從嵌入模板載入（保留所有框線、合併、字型、列印設定）
        tmpl_bytes = base64.b64decode(_APPEAL_FORM_TEMPLATE_B64)
        wb = load_workbook(io.BytesIO(tmpl_bytes))
        ws = wb.active

        # ── Row 3：填入班級 / 姓名 / 學號 ──
        from openpyxl.styles import Alignment
        _ac = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws['B3'] = cls_name
        ws['B3'].alignment = _ac          # [Fix] 置中
        ws['D3'] = str(student_name or '')   # [V6.3 #5] 姓名自動帶入（名冊查無則留空手寫）
        ws['D3'].alignment = _ac
        ws['F3'] = str(student_id)        # F3:G3 merged，填左上角
        ws['F3'].alignment = _ac          # [Fix] 置中

        # ── [V6.3 #2] 時數拆列：每 1 小時一列（1 小時銷 1 支警告），餘數自成一列 ──
        _flat = []
        for rec in records:
            try: _h = float(rec.get('hours', 0))
            except (ValueError, TypeError): _h = 0.0
            _n = int(_h)
            _parts = [1.0] * _n + ([round(_h - _n, 2)] if round(_h - _n, 2) > 0 else [])
            if not _parts: _parts = [0.0]
            for _p in _parts:
                _flat.append({**rec, 'hours': _p})
        if len(_flat) > 8:
            # 超過 8 列裝不下：前 7 列各 1 小時，第 8 列吸收剩餘時數
            _rest = round(sum(x['hours'] for x in _flat[7:]), 2)
            _flat = _flat[:7] + [{**_flat[7], 'hours': _rest}]
        records = _flat

        # ── Row 5-12：填入服務紀錄（最多 8 列）──
        # 欄對應：A=愛校事由, B=缺曠日期(留空), C:D(merged)=工作內容, E=時間起迄(含日期), F=師長驗收(留空), G=累計時數
        total_hours = 0.0
        for i in range(8):
            r = 5 + i
            if i < len(records):
                rec = records[i]
                h = 0.0
                try: h = float(rec.get('hours', 0))
                except (ValueError, TypeError): pass
                st       = rec.get('start_time', '')
                et       = rec.get('end_time', '')
                rec_date = rec.get('date', '')
                # [Fix] 時間起迄加上日期，格式：YYYY-MM-DD\nHH:MM~HH:MM
                if rec_date and (st or et):
                    time_str = f"{rec_date}\n{st}~{et}" if (st and et) else f"{rec_date}\n{st}"
                elif st or et:
                    time_str = f"{st}~{et}" if (st and et) else st
                else:
                    time_str = rec_date
                ws.cell(row=r, column=1).value = '消警告'
                ws.cell(row=r, column=2).value = None          # [Fix] 缺曠日期留空，手填
                ws.cell(row=r, column=3).value = rec.get('work_content', '')  # C:D merged，填C
                _e = ws.cell(row=r, column=5); _e.value = time_str
                _e.alignment = _ac                              # [V6.3 #4] 起訖時間：自動換行＋置中
                total_hours += h
                _g = ws.cell(row=r, column=7)
                _g.value = round(total_hours, 2) if h else None  # [V6.3 #2] 累計時數逐列遞增
                _g.alignment = _ac
            else:
                # 空白列：清除可能殘留的樣板文字
                for col in (1, 2, 3, 5, 7):
                    ws.cell(row=r, column=col).value = None

        # ── [V6.3 #3] 審查簽核改四欄：導師｜生輔組長｜學務主任｜校長（移除主任教官）──
        try:
            for _rng in ("B16:C16", "B17:C17"):
                if _rng in [str(x) for x in ws.merged_cells.ranges]:
                    ws.unmerge_cells(_rng)
            _lbl_font, _lbl_align = _cpy(ws['A16'].font), _cpy(ws['A16'].alignment)
            for _c in ("B16", "D16", "E16"):   # G16 為 F16:G16 合併從屬格（唯讀），不可觸碰
                ws[_c].value = None
            ws['A16'] = '導師'; ws['C16'] = '生輔組長'; ws['E16'] = '學務主任'; ws['F16'] = '校長'
            for _c in ("A16", "C16", "E16", "F16"):
                ws[_c].font = _lbl_font; ws[_c].alignment = _lbl_align
            for _rng in ("A16:B16", "C16:D16", "A17:B17", "C17:D17"):
                ws.merge_cells(_rng)
        except Exception as _sig_e:
            print(f"[appeal_form] 簽核區改版失敗（沿用原版）: {_sig_e}")

        # ── Row 13：合計愛校時數 → F13（F13:G13 merged）──
        ws['F13'] = round(total_hours, 2)

        # ── Row 20：列印日期 ──
        now_tw = datetime.now(TW_TZ)
        roc_y  = now_tw.year - 1911
        ws['A20'] = (
            f'（本表由衛生組系統自動產製，僅供消警告使用，不得銷過。'
            f'列印日期：民國 {roc_y} 年 {now_tw.month} 月 {now_tw.day} 日）'
        )

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    PUBLISHED_COLS = ["週次", "排名", "年級", "班級", "總扣分", "優良次數", "總成績", "評等", "排名模式", "發布時間"]

    @safe_cached(ttl=300, default_factory=lambda: pd.DataFrame(columns=PUBLISHED_COLS))   # [效能] 5分鐘快取，發布後學生很快就看得到
    def load_published_results():
        ws = get_worksheet(SHEET_TABS["published_results"])
        if not ws:
            raise RuntimeError("published_results worksheet unavailable")
        df = pd.DataFrame(ws.get_all_records())
        if df.empty: return pd.DataFrame(columns=PUBLISHED_COLS)
        for col in PUBLISHED_COLS:
            if col not in df.columns: df[col] = ""
        for col in ["週次", "排名", "總扣分", "優良次數", "總成績"]:
            if col in df.columns: df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)
        return df

    def publish_week_results(week_num, fin_ranked_df, rank_mode="全校"):
        """將計算好的週次排名寫入 published_results sheet，同一週再發布會覆蓋舊資料"""
        ws = get_worksheet(SHEET_TABS["published_results"])
        if not ws: return False, "無法連線至 Google Sheets"
        try:
            # 讀取現有資料，刪除同一週次的舊資料
            existing = ws.get_all_values()
            if len(existing) > 1:
                rows_to_delete = [i + 2 for i, row in enumerate(existing[1:])
                                  if row and str(row[0]) == str(week_num)]
                for ridx in sorted(rows_to_delete, reverse=True):
                    ws.delete_rows(ridx)

            # 寫入新資料
            now_str = datetime.now(TW_TZ).strftime("%Y-%m-%d %H:%M")
            for _, row in fin_ranked_df.iterrows():
                ws.append_row([
                    int(week_num),
                    int(row.get("排名", 0)),
                    str(row.get("年級", "")),
                    str(row.get("班級", "")),
                    int(row.get("總扣分", 0)),
                    int(row.get("優良次數", 0)),
                    int(row.get("總成績", 0)),
                    str(row.get("評等", "")),
                    rank_mode,
                    now_str
                ])
            load_published_results.clear()
            return True, f"第 {week_num} 週成績已發布！學生現在可以查詢。"
        except Exception as e:
            return False, str(e)

    def save_appeal(entry, proof_file=None):
        # [Fix #3-B] 佐證照片改為同步上傳 Drive，不再寫本機磁碟
        if proof_file:
            try:
                data = proof_file.read()
                if len(data) > MAX_IMAGE_BYTES:
                    st.error("❌ 照片過大（上限 20MB），請壓縮後再上傳")
                    return False
                fname = f"Appeal_{entry.get('班級', '')}_{datetime.now(TW_TZ).strftime('%H%M%S')}.jpg"
                with st.spinner("📤 上傳佐證照片中..."):
                    link = execute_with_retry(
                        lambda d=data, n=fname: upload_image_to_drive(compress_image_bytes(d), n)
                    )
                entry["佐證照片"] = link or ""
            except Exception as e:
                st.warning(f"⚠️ 佐證照片上傳失敗，將不含照片送出申訴：{e}")
                entry["佐證照片"] = ""

        entry.update({
            "申訴日期": entry.get("申訴日期", datetime.now(TW_TZ).strftime("%Y-%m-%d")),
            "處理狀態": entry.get("處理狀態", "待處理"),
            "登錄時間": entry.get("登錄時間", datetime.now(TW_TZ).strftime("%Y-%m-%d %H:%M:%S")),
            "申訴ID":   entry.get("申訴ID", datetime.now(TW_TZ).strftime("%Y%m%d%H%M%S") + "_" + uuid.uuid4().hex[:4]),
            "佐證照片": entry.get("佐證照片", "")
        })
        enqueue_task("appeal_entry", {"entry": entry})
        st.success("📩 申訴已送出，請耐心等候審核")
        return True
    
    def update_appeal_status(idx, status, record_id, reply_text=""):
        ws_appeals, ws_main = get_worksheet(SHEET_TABS["appeals"]), get_worksheet(SHEET_TABS["main"])
        try:
            data = ws_appeals.get_all_records()
            t_row = next((i + 2 for i, r in enumerate(data) if str(r.get("對應紀錄ID")) == str(record_id) and str(r.get("處理狀態")) == "待處理"), None)
            if t_row:
                ws_appeals.update_cell(t_row, APPEAL_COLUMNS.index("處理狀態") + 1, status)
                if "審核回覆" in APPEAL_COLUMNS:
                    ws_appeals.update_cell(t_row, APPEAL_COLUMNS.index("審核回覆") + 1, reply_text)
                    
                if status == "已核可":
                    m_data = ws_main.get_all_records()
                    m_row = next((j + 2 for j, mr in enumerate(m_data) if str(mr.get("紀錄ID")) == str(record_id)), None)
                    if m_row: ws_main.update_cell(m_row, EXPECTED_COLUMNS.index("修正") + 1, "TRUE")
                load_main_data.clear()
                load_appeals.clear()
                return True, "更新成功"
            return False, "找不到對應的申訴列"
        except Exception as e: return False, str(e)

    def delete_rows_by_ids(ids):
        ws = get_worksheet(SHEET_TABS["main"])
        if not ws: return False
        try:
            rows = sorted([i + 2 for i, r in enumerate(ws.get_all_records()) if str(r.get("紀錄ID")) in ids], reverse=True)
            for r in rows: ws.delete_rows(r)
            time.sleep(0.8); load_main_data.clear()
            return True
        except Exception as e: st.error(f"刪除失敗: {e}"); return False

    _INSPECTOR_DEFAULT = [{"label": "測試人員", "allowed_roles": ["內掃檢查"], "assigned_classes": [], "id_prefix": "測", "raw_role": "內掃"}]

    @safe_cached(ttl=21600, default_factory=lambda: list(_INSPECTOR_DEFAULT))
    def load_inspector_list():
        ws = get_worksheet(SHEET_TABS["inspectors"])
        if not ws:
            raise RuntimeError("inspectors worksheet unavailable")
        df = pd.DataFrame(ws.get_all_records())
        if df.empty: return list(_INSPECTOR_DEFAULT)
        inspectors, id_c, r_c, s_c = [], next((c for c in df.columns if "學號" in c or "編號" in c), None), next((c for c in df.columns if "負責" in c or "項目" in c), None), next((c for c in df.columns if "班級" in c or "範圍" in c), None)
        if id_c:
            for _, row in df.iterrows():
                sid, s_role = clean_id(row[id_c]), str(row[r_c]).strip() if r_c else ""
                
                allowed = []
                if "組長" in s_role:
                    allowed = ["內掃檢查", "外掃檢查", "垃圾/回收檢查", "晨間打掃"]
                else:
                    if "外掃" in s_role: allowed.append("外掃檢查")
                    if "垃圾" in s_role or "回收" in s_role: allowed.append("垃圾/回收檢查")
                    if "晨" in s_role: allowed.append("晨間打掃")
                    if "內掃" in s_role: allowed.append("內掃檢查")
                    
                    if "衛生糾察隊長" in s_role or "機動" in s_role:
                        allowed = [r for r in allowed if r != "垃圾/回收檢查"]
                        if not allowed: allowed = ["內掃檢查", "外掃檢查"]
                    elif "環保糾察隊長" in s_role:
                        allowed = [r for r in allowed if r not in ["內掃檢查", "外掃檢查"]]
                        if "垃圾/回收檢查" not in allowed: allowed.append("垃圾/回收檢查")
                        
                    if not allowed: allowed = ["內掃檢查"]

                s_classes = [c.strip() for c in str(row[s_c]).replace("、", ";").replace(",", ";").split(";") if c.strip()] if s_c and str(row[s_c]) else []
                
                inspectors.append({
                    "label": f"學號: {sid}", "allowed_roles": allowed, 
                    "assigned_classes": s_classes, "id_prefix": sid[0] if sid else "X",
                    "raw_role": s_role
                })
        return inspectors or list(_INSPECTOR_DEFAULT)

    def check_duplicate_record(df, check_date, inspector, role, target_class=None):
        if df.empty: return False
        try:
            # 同時比對原始 role、以及加了(優良)/(普通) 的變體，避免重複評分
            mask = (df["日期"].astype(str) == str(check_date)) & \
                   (df["檢查人員"] == inspector) & \
                   (df["評分項目"].astype(str).str.startswith(role))
            if target_class: mask &= (df["班級"] == target_class)
            return not df[mask].empty
        except Exception as e:
            print(f"[check_duplicate] {e}")
            return False

    # [V5.28] 加入 award_inspector_hours 控制是否發放時數
    # [V5.32] 尖峰時段 jitter 常數（秒）
    # 15:50~16:30 為尖峰時段，42 人同時送出時加入隨機延遲分散 API 壓力
    PEAK_START_H, PEAK_START_M = 15, 50
    PEAK_END_H,   PEAK_END_M   = 16, 30
    PEAK_JITTER_MAX = 8  # 最多延遲 8 秒

    def _is_peak_hour(now_dt):
        """判斷目前是否在尖峰時段"""
        t = now_dt.time()
        from datetime import time as dtime
        return dtime(PEAK_START_H, PEAK_START_M) <= t <= dtime(PEAK_END_H, PEAK_END_M)

    def save_entry(new_entry, uploaded_files=None, student_list=None, custom_hours=0.5, custom_category="晨掃志工", award_inspector_hours=True, skip_jitter=False):
        # [Fix #3-B] 照片改為同步上傳 Drive，不再寫本機磁碟，移除 Semaphore
        new_entry["日期"] = str(new_entry.get("日期", str(date.today())))
        new_entry["紀錄ID"] = new_entry.get("紀錄ID", f"{datetime.now(TW_TZ).strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:6]}")
        if "登錄時間" not in new_entry or not new_entry["登錄時間"]:
            new_entry["登錄時間"] = datetime.now(TW_TZ).strftime("%Y-%m-%d %H:%M:%S")

        # [V5.32] 尖峰時段 jitter：分散 42 人同時提交的 API 請求
        # [V5.34] skip_jitter=True 時略過，避免批次迴圈（垃圾登記）累積成數十秒卡頓
        if not skip_jitter and _is_peak_hour(datetime.now(TW_TZ)):
            jitter = random.uniform(0, PEAK_JITTER_MAX)
            with st.spinner(f"📶 排隊上傳中（{jitter:.1f}s）…"):
                time.sleep(jitter)

        # 同步上傳照片到 Drive（使用 ThreadPoolExecutor 並發，縮短等待時間）
        if uploaded_files:
            valid_files = []
            for i, up_file in enumerate(uploaded_files):
                if not up_file: continue
                try:
                    data = up_file.getvalue()
                    if len(data) > MAX_IMAGE_BYTES:
                        st.warning(f"檔案過大略過: {up_file.name}")
                        continue
                    valid_files.append((i, data, f"{new_entry['紀錄ID']}_{i}.jpg"))
                except Exception as e:
                    print(f"讀取檔案失敗: {e}")

            if valid_files:
                # [V5.32] 每張照片獨立顯示進度
                upload_status_area = st.empty()
                completed_links = []
                failed_count = 0

                def _upload_one_with_progress(args):
                    idx, data, fname = args
                    try:
                        link = execute_with_retry(
                            lambda d=data, n=fname: upload_image_to_drive(compress_image_bytes(d), n)
                        )
                        return link
                    except Exception as e:
                        print(f"Drive 上傳失敗 ({fname}): {e}")
                        return None

                upload_status_area.info(f"📤 準備上傳 {len(valid_files)} 張照片…")
                with concurrent.futures.ThreadPoolExecutor(max_workers=min(len(valid_files), 3)) as pool:
                    futures = {pool.submit(_upload_one_with_progress, args): i for i, args in enumerate(valid_files)}
                    for done_count, future in enumerate(concurrent.futures.as_completed(futures), 1):
                        result = future.result()
                        if result:
                            completed_links.append(result)
                            upload_status_area.info(f"📤 上傳中… {done_count}/{len(valid_files)} 張完成")
                        else:
                            failed_count += 1
                            upload_status_area.warning(f"⚠️ 第 {done_count} 張上傳失敗，繼續處理其餘照片…")

                if completed_links:
                    new_entry["照片路徑"] = ";".join(completed_links)
                    upload_status_area.success(f"✅ {len(completed_links)} 張照片上傳完成！正在寫入紀錄…")
                else:
                    upload_status_area.empty()

                if failed_count > 0:
                    st.warning(f"⚠️ {failed_count} 張照片上傳失敗，資料已送出但不含這些照片。請截圖後告知組長補傳。")

        payload = {
            "entry": new_entry,
            "student_list": student_list or [],
            "custom_hours": custom_hours,
            "custom_category": custom_category,
            "award_inspector_hours": award_inspector_hours
        }
        # [Patch 2] 檢查 enqueue 回傳值，防止靜默丟失資料
        task_id = enqueue_task("volunteer_report" if student_list is not None else "main_entry", payload)
        if not task_id:
            st.error("❌ 系統繁忙，資料未送出！請等待 10 秒後重新點擊送出。")
            return False
        return True

    def load_full_semester_data_for_export():
        # 直接重用 load_main_data 的快取，不重複讀 Sheets，記憶體只存一份
        return load_main_data()

    # ==========================================
    # 3. 主程式 UI 啟動前準備
    # ==========================================
    now_tw = datetime.now(TW_TZ)
    today_tw = now_tw.date()

    # [V5.35] 未來日期防呆 helper
    # 解決 Streamlit date_input + key 會把選過的日期記在 session_state，
    # 導致使用者下次回到該頁時，畫面上日期還停在舊值（可能是未來日期）就送出 → 寫入錯誤日期
    def _block_future_date(d, label="日期"):
        """檢查 d 是否大於今天（台灣時區）。是 → st.error + 回傳 True；否 → 回傳 False。
        呼叫端用 `if _block_future_date(d, '出勤日期'): return` 來阻擋送出。"""
        if d and d > today_tw:
            st.error(f"⛔ 「{label}」不能是未來日期（{d}）！請改回今天（{today_tw}）或更早，否則無法送出。")
            return True
        return False

    if "last_action_time" not in st.session_state:
        st.session_state.last_action_time = 0
    
    SYSTEM_CONFIG, ROSTER_DICT, INSPECTOR_LIST = load_settings(), load_roster_dict(), load_inspector_list()
    all_classes, structured_classes = load_sorted_classes()
    if not all_classes: all_classes, structured_classes = ["測試班級"], [{"grade": "其他", "name": "測試班級"}]
    grades = sorted(list(set([c["grade"] for c in structured_classes])))
    
    def get_week_num(d):
        try:
            if isinstance(d, datetime): d = d.date()
            # [週次對照表] 優先使用 settings 裡的 week_map 手動對照
            # 格式：2025-01-23:1,2025-02-23:2（只需填「錨點」，後面正常週數自動累計）
            week_map_str = SYSTEM_CONFIG.get("week_map", "")
            if week_map_str.strip():
                entries = []
                for item in week_map_str.split(","):
                    item = item.strip()
                    if ":" in item:
                        date_str, wn = item.rsplit(":", 1)
                        try:
                            entries.append((datetime.strptime(date_str.strip(), "%Y-%m-%d").date(), int(wn.strip())))
                        except Exception: pass
                if entries:
                    entries.sort(key=lambda x: x[0])
                    # 找到最後一個「錨點日期 <= d」的錨點，從那個錨點往後正常累計週數
                    anchor_date, anchor_week = None, 0
                    for start_date, wn in entries:
                        if d >= start_date:
                            anchor_date, anchor_week = start_date, wn
                        else:
                            break
                    if anchor_date is not None:
                        # 從錨點往後每7天加一週
                        return anchor_week + (d - anchor_date).days // 7
                    return 0
            # fallback：純數學計算（未設定 week_map 時使用）
            start = datetime.strptime(SYSTEM_CONFIG["semester_start"], "%Y-%m-%d").date()
            return max(0, ((d - start).days // 7) + 1)
        except Exception as e:
            print(f"[mode3_check] {e}")
            return 0

    # ── 全域樣式注入 ─────────────────────────────────────────
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+TC:wght@400;500;700&display=swap');

    /* ── 全域字體 ── */
    html, body, [class*="css"] {
        font-family: 'Noto Sans TC', sans-serif !important;
    }

    /* ── 修正手機上 radio / checkbox / selectbox 文字顏色（僅主內容區，不影響側邊欄）── */
    .main .stRadio label p,
    .main .stRadio label span,
    .main [data-testid="stRadio"] label,
    .main [data-testid="stRadio"] p,
    .main .stCheckbox label p,
    .main .stCheckbox label span,
    .main [data-testid="stCheckbox"] label,
    .main [data-testid="stSelectbox"] span,
    .main [data-testid="stSelectbox"] p,
    .main .stSelectbox label,
    .main .stMarkdown p,
    .main .stText p {
        color: #1a2a4a !important;
    }

    /* ── 主內容區背景 ── */
    .stApp {
        background: #f0f4f9;
    }
    [data-testid="stAppViewContainer"] > .main {
        background: #f0f4f9;
    }

    /* ── 主內容區塊加陰影卡片感 ── */
    [data-testid="stVerticalBlock"] > div > [data-testid="stVerticalBlock"] {
        background: white;
        border-radius: 16px;
        padding: 4px 0;
    }

    /* ── 頁面標題 h1 ── */
    h1 {
        color: #1a2a4a !important;
        font-weight: 700 !important;
        font-size: 26px !important;
        border-left: 5px solid #3182ce;
        padding-left: 12px !important;
        margin-bottom: 20px !important;
    }
    h2 { color: #2c3e6b !important; font-weight: 600 !important; }
    h3 { color: #2c4a8a !important; }

    /* ── 所有按鈕 ── */
    .stButton > button {
        border-radius: 10px !important;
        font-weight: 600 !important;
        font-size: 14px !important;
        padding: 8px 18px !important;
        border: none !important;
        background: linear-gradient(135deg, #2b6cb0, #3182ce) !important;
        color: white !important;
        box-shadow: 0 2px 8px rgba(49,130,206,0.35) !important;
        transition: all 0.2s ease !important;
    }
    .stButton > button:hover {
        background: linear-gradient(135deg, #1a4a8a, #2b6cb0) !important;
        box-shadow: 0 4px 14px rgba(49,130,206,0.45) !important;
        transform: translateY(-1px) !important;
    }
    .stButton > button:active { transform: translateY(0px) !important; }

    /* ── 下載按鈕特別用綠色 ── */
    [data-testid="stDownloadButton"] > button {
        background: linear-gradient(135deg, #276749, #38a169) !important;
        box-shadow: 0 2px 8px rgba(56,161,105,0.35) !important;
    }
    [data-testid="stDownloadButton"] > button:hover {
        background: linear-gradient(135deg, #1c4f37, #276749) !important;
    }

    /* ── 表單送出按鈕 ── */
    [data-testid="stFormSubmitButton"] > button {
        width: 100% !important;
        background: linear-gradient(135deg, #553c9a, #6b46c1) !important;
        box-shadow: 0 2px 8px rgba(107,70,193,0.35) !important;
        font-size: 16px !important;
        padding: 12px !important;
    }
    [data-testid="stFormSubmitButton"] > button:hover {
        background: linear-gradient(135deg, #44337a, #553c9a) !important;
    }

    /* ── input / selectbox / text_area ── */
    .stTextInput > div > div > input,
    .stTextArea > div > div > textarea,
    .stNumberInput > div > div > input {
        border-radius: 8px !important;
        border: 1.5px solid #c9d7e8 !important;
        background: #fafcff !important;
        font-size: 14px !important;
        transition: border 0.2s;
    }
    .stTextInput > div > div > input:focus,
    .stTextArea > div > div > textarea:focus {
        border-color: #3182ce !important;
        box-shadow: 0 0 0 3px rgba(49,130,206,0.15) !important;
    }

    /* ── selectbox ── */
    [data-testid="stSelectbox"] > div > div {
        border-radius: 8px !important;
        border: 1.5px solid #c9d7e8 !important;
        background: #fafcff !important;
    }

    /* ── info / warning / error / success 提示框 ── */
    [data-testid="stAlert"] {
        border-radius: 10px !important;
        border-left-width: 5px !important;
        font-size: 14px !important;
    }

    /* ── expander ── */
    [data-testid="stExpander"] {
        border: 1.5px solid #d8e4f0 !important;
        border-radius: 12px !important;
        background: white !important;
        box-shadow: 0 1px 6px rgba(0,0,0,0.06) !important;
    }
    [data-testid="stExpander"] summary {
        font-weight: 600 !important;
        font-size: 15px !important;
        color: #1a2a4a !important;
        padding: 12px 16px !important;
    }

    /* ── container border ── */
    [data-testid="stVerticalBlockBorderWrapper"] {
        border: 1.5px solid #d8e4f0 !important;
        border-radius: 14px !important;
        background: white !important;
        box-shadow: 0 2px 10px rgba(0,0,0,0.05) !important;
        padding: 12px !important;
    }

    /* ── dataframe / table ── */
    [data-testid="stDataFrame"] {
        border-radius: 10px !important;
        overflow: hidden !important;
        box-shadow: 0 1px 8px rgba(0,0,0,0.07) !important;
    }

    /* ── tabs ── */
    [data-testid="stTabs"] [role="tab"] {
        font-weight: 600 !important;
        font-size: 14px !important;
        border-radius: 8px 8px 0 0 !important;
        padding: 8px 16px !important;
    }
    [data-testid="stTabs"] [role="tab"][aria-selected="true"] {
        color: #2b6cb0 !important;
        border-bottom: 3px solid #3182ce !important;
    }

    /* ── divider ── */
    hr { border-color: #d8e4f0 !important; margin: 16px 0 !important; }

    /* ══════════════════════════════
       側邊欄
    ══════════════════════════════ */
    [data-testid="stSidebar"] {
        background: linear-gradient(160deg, #1a2a4a 0%, #0d1b35 100%) !important;
    }
    [data-testid="stSidebar"] * { color: #e8edf5 !important; }
    [data-testid="stSidebar"] .stRadio > label { display: none; }
    [data-testid="stSidebar"] .stRadio div[role="radiogroup"] {
        display: flex; flex-direction: column; gap: 6px; margin-top: 4px;
    }
    [data-testid="stSidebar"] .stRadio div[role="radiogroup"] label {
        display: flex !important; align-items: center;
        background: rgba(255,255,255,0.07);
        border: 1px solid rgba(255,255,255,0.12);
        border-radius: 10px;
        padding: 10px 14px;
        cursor: pointer;
        transition: all 0.2s ease;
        font-size: 15px !important;
        font-weight: 500 !important;
    }
    [data-testid="stSidebar"] .stRadio div[role="radiogroup"] label:hover {
        background: rgba(255,255,255,0.15);
        border-color: rgba(255,255,255,0.3);
        transform: translateX(3px);
    }
    [data-testid="stSidebar"] .stRadio div[role="radiogroup"] label:has(input:checked) {
        background: rgba(99,179,237,0.25);
        border-color: #63b3ed;
        color: #ffffff !important;
    }
    [data-testid="stSidebar"] .stRadio input[type="radio"] { display: none; }
    [data-testid="stSidebar"] h1 {
        font-size: 20px !important; font-weight: 700 !important;
        color: #ffffff !important;
        padding-bottom: 8px;
        border-bottom: 1px solid rgba(255,255,255,0.15);
        margin-bottom: 14px !important;
        border-left: none !important;
        padding-left: 0 !important;
    }
    .sidebar-footer {
        position: fixed; bottom: 20px; left: 0;
        width: 240px; text-align: center;
        font-size: 11px; color: rgba(255,255,255,0.3);
    }
    </style>
    """, unsafe_allow_html=True)

    st.sidebar.title("🏫 衛愛而生")
    st.sidebar.markdown("<div style='font-size:12px;color:rgba(255,255,255,0.4);margin-top:-12px;margin-bottom:16px;'>中壢家商 衛生組管理系統</div>", unsafe_allow_html=True)

    menu_options = ["糾察底家👀", "班級負責人🥸", "組長ㄉ窩💃"]  # [V6.1] 愛校任務認領、晨掃志工隊已冷凍（程式碼保留，恢復時加回選單即可）
    app_mode = st.sidebar.radio("請選擇模式", menu_options)

    st.sidebar.markdown("---")
    st.sidebar.markdown("📅 [衛生組行事曆](https://www.notion.so/312b7f229eea80c584a1e794c7b955a4)")
    st.sidebar.markdown("📸 [衛生組 Instagram](https://www.instagram.com/clvs_captain.h/)")
    st.sidebar.markdown("📂 [衛生組公開資料區](https://drive.google.com/drive/folders/14QcUILCmHKnKhDx2X1dIUl_6PNRndCub)")
    st.sidebar.markdown("<div class='sidebar-footer'>衛生組長林ㄊㄩ製作@2025</div>", unsafe_allow_html=True)

    # --- Mode: 愛校任務認領 🤝 ---
    if app_mode == "愛校任務認領 🤝":
        st.title("🤝 愛校服務認領區")
        st.info("💡 這裡的任務清單與 Notion 行事曆即時同步！成功認領後，任務會自動標記並更新。")
        
        # [新增] 愛校服務 2.0：欠時查詢
        with st.expander("🔍 查詢我的欠時與明細"):
            _query_sid = st.text_input("請輸入你的學號", placeholder="例如：112001", key="debt_query_sid")
            if st.button("🔎 查詢", key="debt_query_btn"):
                if not _query_sid:
                    st.error("請先輸入學號！")
                else:
                    _debts_map = load_student_debts()
                    _my_hours = _debts_map.get(clean_id(_query_sid), 0.0)
                    if _my_hours > 0:
                        st.warning(f"⚠️ 你目前共有 **{_my_hours}** 小時的欠時尚未歸還。")
                        # [新增] 顯示 Google Sheet 中的備註說明
                        _debt_note = load_student_debt_note(clean_id(_query_sid))
                        if _debt_note:
                            st.info(f"📝 備註說明：{_debt_note}")
                        _hist_df = load_debt_history(clean_id(_query_sid))
                        if not _hist_df.empty:
                            st.dataframe(_hist_df, hide_index=True)
                        else:
                            st.caption("暫無歷史異動紀錄。")
                    else:
                        st.success("🎉 你目前沒有欠時紀錄，繼續保持！")

        n_token = st.secrets.get("notion_token") or st.secrets.get("system_config", {}).get("notion_token")
        if not NOTION_INSTALLED:
            st.error("⚠️ 系統偵測到未安裝 `notion-client` 套件，請通知管理員檢查系統設定。")
        elif not n_token:
            st.warning("⚠️ Notion 金鑰尚未設定，請通知管理員至後台設定 `notion_token`。")
        else:
            with st.spinner("正在向 Notion 獲取最新任務..."):
                tasks, error_msg = fetch_available_notion_tasks()
                
            if error_msg:
                st.error(f"⚠️ 讀取 Notion 發生錯誤！請檢查以下錯誤訊息：\n\n{error_msg}")
            elif not tasks:
                st.success("🎉 娃，目前沒有任務！")
                st.balloons()
            else:
                st.write(f"目前共有 **{len(tasks)}** 項待認領的任務：")
                
                for t in tasks:
                    with st.container(border=True):
                        col1, col2 = st.columns([2, 1])
                        with col1:
                            st.subheader(f"📌 {t['title']} (進度: {t['current_count']} / {t['req_num']} 人)")
                            st.write(f"📅 **執行日期:** {t['date']}")
                            st.write(f"🧹 **任務內容:** {t['area']}")
                        
                        with col2:
                            with st.form(f"claim_form_{t['id']}"):
                                s_id = st.text_input("請輸入您的【學號】來認領：", placeholder="例如：112001", key=f"claim_sid_{t['id']}")
                                # [新增] 愛校服務 2.0：服務目的選擇
                                purpose_choice = st.selectbox("本次愛校服務目的", ["還時數", "消警告", "糾察懲罰"], key=f"claim_purpose_{t['id']}")
                                if st.form_submit_button("🚀 確認認領", use_container_width=True):
                                    if time.time() - st.session_state.last_action_time < 3:
                                        st.warning("⚠️ 系統處理中，請勿連續點擊！")
                                    elif not s_id:
                                        st.error("學號不能為空！")
                                    else:
                                        st.session_state.last_action_time = time.time()
                                        # [新增] 愛校服務 2.0：欠時防呆
                                        _purpose = purpose_choice
                                        _debts_check = load_student_debts()
                                        _my_debt = _debts_check.get(clean_id(s_id), 0.0)
                                        if _my_debt > 0 and _purpose != "還時數":
                                            _purpose = "還時數"
                                            st.warning(f"⚠️ 你目前仍有 {_my_debt} 小時欠時未還，本次目的已自動改為「還時數」！")
                                        # [Debug] 記錄實際寫入 Notion 的 purpose_tag
                                        print(f"[claim] sid={s_id}, purpose_tag='{_purpose}', label will be: {s_id}({_purpose})")
                                        with st.spinner("連線至 Notion 更新看板中..."):
                                            success, msg = claim_notion_task(t['id'], s_id, purpose_tag=_purpose)
                                        if success:
                                            if msg == "滿團":
                                                st.success(f"✅ 學號 {s_id} 認領成功！此任務已額滿，自動從看板隱藏。")
                                            else:
                                                st.success(f"✅ 學號 {s_id} 認領成功！目前還缺人，趕緊揪同學來認領！")
                                            time.sleep(2)
                                            st.rerun()
                                        else:
                                            st.error(f"認領失敗：{msg}")

    # --- Mode 1: 糾察評分 ---
    elif app_mode == "糾察底家👀":
        st.title("📝 衛生糾察評分系統")
        if "team_logged_in" not in st.session_state: st.session_state["team_logged_in"] = False

        daily_hygiene = SYSTEM_CONFIG.get("daily_hygiene_task", "")
        if daily_hygiene:
            formatted_hygiene = daily_hygiene.replace('\n', '<br>')
            mascot_url = "https://drive.google.com/thumbnail?id=128ITPXtpGNuI-wLIt6p-qd4ZNNhCGbhd" 
            
            bubble_html_h = f"""
            <style>
            .mascot-container-h {{ display: flex; align-items: flex-start; margin-bottom: 20px; gap: 15px; }}
            .mascot-img-h {{ width: 200px; flex-shrink: 0; }}
            .speech-bubble-h {{
                position: relative; background: #D0E8F2;
                border-radius: 15px; padding: 15px 20px; color: #05445E; font-size: 16px;
                box-shadow: 2px 4px 10px rgba(0,0,0,0.1); border: 2px solid #189AB4; flex-grow: 1;
            }}
            .speech-bubble-h::before {{ content: ''; position: absolute; left: -20px; top: 30px; border: 10px solid transparent; border-right-color: #189AB4; }}
            .speech-bubble-h::after {{ content: ''; position: absolute; left: -16px; top: 30px; border: 10px solid transparent; border-right-color: #D0E8F2; }}
            @media (max-width: 500px) {{
                .mascot-img-h {{ width: 120px; }}
                .speech-bubble-h {{ font-size: 14px; padding: 10px 15px; }}
            }}
            </style>
            <div class="mascot-container-h">
                <img src="{mascot_url}" class="mascot-img-h" />
                <div class="speech-bubble-h">
                    <strong>📢 組長廣播 / 糾察重點：</strong><br>
                    {formatted_hygiene}
                </div>
            </div>
            """
            st.markdown(bubble_html_h, unsafe_allow_html=True)
        
        if not st.session_state["team_logged_in"]:
            with st.expander("🔐 身份驗證", expanded=True):
                pwd_input = st.text_input("請輸入隊伍通行碼", type="password", key="m1_login_pwd")
                if pwd_input:
                    if pwd_input == st.secrets["system_config"]["team_password"]:
                        st.session_state["team_logged_in"] = True
                        st.rerun()
                    else:
                        st.error("通行碼錯誤")
        
        if st.session_state["team_logged_in"]:
            prefixes = sorted(list(set([p["id_prefix"] for p in INSPECTOR_LIST])))
            if not prefixes: st.warning("找不到糾察名單")
            else:
                sel_p = st.radio("步驟 1：選擇開頭", [f"{p}開頭" for p in prefixes], horizontal=True, key="m1_p_radio")[0]
                inspector_name = st.radio("步驟 2：點選身份", [p["label"] for p in INSPECTOR_LIST if p["id_prefix"] == sel_p], key="m1_name_radio")
                curr_inspector = next((p for p in INSPECTOR_LIST if p["label"] == inspector_name), {})
                allowed_roles = [r for r in curr_inspector.get("allowed_roles", ["內掃檢查"]) if r != "晨間打掃"] or ["內掃檢查"]
                
                st.markdown("---")
                c_d, c_r = st.columns(2)
                input_date = c_d.date_input("檢查日期", today_tw)
                if _block_future_date(input_date, "檢查日期"):
                    st.stop()
                role = c_r.radio("檢查項目", allowed_roles, horizontal=True, key="m1_role_radio") if len(allowed_roles)>1 else allowed_roles[0]
                week_num = get_week_num(input_date)
                main_df = load_main_data()
                # 糾察評分只需近兩週資料（查重複用），在 UI 層過濾
                now_week = get_week_num(today_tw)
                if now_week >= 3:
                    main_df = main_df[main_df["週次"] >= now_week - 2]

                if role == "垃圾/回收檢查":
                    st.info("🗑️ 資源回收與垃圾檢查 (每日每班此項目總扣分上限2分將於結算時自動卡控)")
                    
                    step_a = st.radio("步驟 A: 選擇垃圾類別", ["一般垃圾", "紙類", "網袋aka塑膠鐵鋁", "其他"], horizontal=True, key="m1_trash_a")
                    sel_filter = st.radio("步驟 B: 篩選檢查對象", ["各處室 (外掃)"] + grades, horizontal=True, key="m1_trash_b")
                    
                    today_records = main_df[(main_df["日期"].astype(str) == str(input_date)) & (main_df["評分項目"] == "垃圾/回收檢查") & (main_df["違規細項"] == step_a)] if not main_df.empty else pd.DataFrame()
                    rows = []
                    
                    if sel_filter == "各處室 (外掃)":
                        office_map = load_office_area_map()
                        target_list = list(office_map.keys()) or ["教務處", "學務處", "總務處", "輔導室", "圖書館"]
                        for off_name in target_list:
                            cls_name = office_map.get(off_name, "未設定")
                            
                            is_dump_bad = any(f"外掃({off_name})" in str(r["備註"]) and "未倒垃圾" in str(r["備註"]) for _, r in today_records.iterrows()) if not today_records.empty else False
                            is_sort_bad = any(f"外掃({off_name})" in str(r["備註"]) and "未做好分類" in str(r["備註"]) for _, r in today_records.iterrows()) if not today_records.empty else False
                            
                            row_data = {"處室/區域": off_name, "負責班級": cls_name, "未倒垃圾": is_dump_bad, "未做好分類": is_sort_bad}
                            rows.append(row_data)
                            
                        col_config = {"處室/區域": st.column_config.TextColumn(disabled=True), "負責班級": st.column_config.TextColumn(disabled=True)}
                        col_config["未倒垃圾"] = st.column_config.CheckboxColumn("🗑️ 未倒垃圾", help="扣1分")
                        col_config["未做好分類"] = st.column_config.CheckboxColumn("♻️ 未做好分類", help="扣1分")
                        
                        edited_df = st.data_editor(pd.DataFrame(rows), column_config=col_config, hide_index=True, width="stretch", key="ed_offices")
                        
                        if st.button(f"💾 登記違規 ({step_a} - 各處室)"):
                            if time.time() - st.session_state.last_action_time < 3:
                                st.warning("⚠️ 系統處理中，請勿連續點擊！")
                            else:
                                st.session_state.last_action_time = time.time()
                                cnt = 0
                                for _, row in edited_df.iterrows():
                                    off, cls = row["處室/區域"], row["負責班級"]
                                    b_sort = row.get("未做好分類", False)
                                    b_dump = row.get("未倒垃圾", False)
                                    
                                    orig = next((x for x in rows if x["處室/區域"] == off), None)
                                    v_list = []
                                    if b_dump and not orig.get("未倒垃圾", False): v_list.append("未倒垃圾")
                                    if b_sort and not orig.get("未做好分類", False): v_list.append("未做好分類")
                                    
                                    if v_list:
                                        score = len(v_list)
                                        base = {"日期": input_date, "週次": week_num, "檢查人員": inspector_name, "登錄時間": now_tw.strftime("%Y-%m-%d %H:%M:%S"), "班級": cls, "評分項目": role, "垃圾內掃原始分": 0, "垃圾外掃原始分": score}
                                        if save_entry({**base, "備註": f"外掃({off})-{step_a}({','.join(v_list)})", "違規細項": step_a}, skip_jitter=True):
                                            cnt += 1
                                if cnt: st.success(f"✅ 已登記 {cnt} 筆違規！"); time.sleep(1.5); st.rerun()

                    else:
                        for cls_name in [c["name"] for c in structured_classes if c["grade"] == sel_filter]:
                            cls_rec = today_records[today_records["班級"] == cls_name] if not today_records.empty else pd.DataFrame()
                            
                            is_dump_bad = any("內掃" in str(r["備註"]) and "未倒垃圾" in str(r["備註"]) for _, r in cls_rec.iterrows()) if not cls_rec.empty else False
                            is_sort_bad = any("內掃" in str(r["備註"]) and "未做好分類" in str(r["備註"]) for _, r in cls_rec.iterrows()) if not cls_rec.empty else False
                            
                            row_data = {"班級": cls_name, "未倒垃圾": is_dump_bad, "未做好分類": is_sort_bad}
                            rows.append(row_data)
                            
                        col_config = {"班級": st.column_config.TextColumn(disabled=True)}
                        col_config["未倒垃圾"] = st.column_config.CheckboxColumn("🗑️ 未倒垃圾", help="扣1分")
                        col_config["未做好分類"] = st.column_config.CheckboxColumn("♻️ 未做好分類", help="扣1分")
                            
                        edited_df = st.data_editor(pd.DataFrame(rows), column_config=col_config, hide_index=True, width="stretch", key=f"ed_{sel_filter}")
                        
                        if st.button(f"💾 登記違規 ({step_a} - {sel_filter})"):
                            if time.time() - st.session_state.last_action_time < 3:
                                st.warning("⚠️ 系統處理中，請勿連續點擊！")
                            else:
                                st.session_state.last_action_time = time.time()
                                cnt = 0
                                for _, row in edited_df.iterrows():
                                    cls = row["班級"]
                                    b_sort = row.get("未做好分類", False)
                                    b_dump = row.get("未倒垃圾", False)
                                    
                                    orig = next((x for x in rows if x["班級"] == cls), None)
                                    v_list = []
                                    if b_dump and not orig.get("未倒垃圾", False): v_list.append("未倒垃圾")
                                    if b_sort and not orig.get("未做好分類", False): v_list.append("未做好分類")
                                    
                                    if v_list:
                                        score = len(v_list)
                                        base = {"日期": input_date, "週次": week_num, "檢查人員": inspector_name, "登錄時間": now_tw.strftime("%Y-%m-%d %H:%M:%S"), "班級": cls, "評分項目": role, "垃圾內掃原始分": score, "垃圾外掃原始分": 0}
                                        if save_entry({**base, "備註": f"內掃-{step_a}({','.join(v_list)})", "違規細項": step_a}, skip_jitter=True):
                                            cnt += 1
                                if cnt: st.success(f"✅ 已登記 {cnt} 筆違規！"); time.sleep(1.5); st.rerun()

                else:
                    assigned_classes = curr_inspector.get("assigned_classes", [])
                    is_last_task = True
                    pending_classes = []

                    # [初始化] 確保所有糾察（含機動/組長）都能使用此 key
                    if "submitted_inspections" not in st.session_state:
                        st.session_state.submitted_inspections = set()

                    if assigned_classes:

                        # 從 Google Sheet 讀到的已完成班級（字串，直接是班級名稱）
                        sheet_done = set(
                            main_df[
                                (main_df["日期"].astype(str) == str(input_date)) &
                                (main_df["檢查人員"] == inspector_name)
                            ]["班級"].astype(str).tolist()
                        )
                        # 從 session_state 讀到的本地已送出班級（格式：日期__糾察名__班級）
                        local_done = set(
                            key.split("__")[2]
                            for key in st.session_state.submitted_inspections
                            if key.startswith(f"{input_date}__{inspector_name}__")
                        )
                        completed_class_names = sheet_done | local_done
                        pending_classes = [c for c in assigned_classes if c not in completed_class_names]

                        # 合併「任務類型 + 進度」為同一個框
                        role_label = "🏫 內掃檢查" if role == "內掃檢查" else "🏢 外掃檢查"
                        progress_text = f"今日任務：{role_label}　|　進度：{len(completed_class_names)}/{len(assigned_classes)}"
                        if pending_classes:
                            progress_text += f"　|　尚缺：{', '.join(pending_classes)}"
                        else:
                            progress_text += "　|　✅ 今日任務全數完成！"
                        st.info(progress_text)

                        # [新增] 外掃模式：顯示「班級 → 外掃區域」對照表
                        _class_area_map = {}
                        if role == "外掃檢查":
                            _class_area_map = load_class_outer_area_map()
                            _table_rows = []
                            for _c in assigned_classes:
                                _done = _c in completed_class_names
                                _area = _class_area_map.get(_c, "（未設定區域，請通知管理員）")
                                _table_rows.append({
                                    "狀態": "✅ 已檢查" if _done else "⏳ 待檢查",
                                    "班級": _c,
                                    "外掃區域": _area,
                                })
                            if _table_rows:
                                with st.expander("📋 今日待檢查班級 + 外掃區域", expanded=True):
                                    st.dataframe(
                                        pd.DataFrame(_table_rows),
                                        hide_index=True,
                                        width="stretch"
                                    )
                                    st.caption("💡 若區域顯示「未設定」，請通知管理員到 Google Sheet 的 class_areas 分頁補資料。")

                        sel_cls = st.radio("選擇負責班級", assigned_classes, key="m1_cls_assigned")

                        # [新增] 外掃模式：點選班級後，再次提示該班外掃區域
                        if role == "外掃檢查" and sel_cls:
                            _sel_area = _class_area_map.get(sel_cls, "（未設定區域）")
                            st.info(f"📍 **{sel_cls}** 外掃區域：**{_sel_area}**")

                        # 判斷這是不是最後一個缺少的班級
                        if sel_cls in pending_classes and len(pending_classes) == 1:
                            is_last_task = True
                        elif sel_cls in pending_classes:
                            is_last_task = False
                        else:
                            is_last_task = False
                    else:
                        st.info("📍 今日任務：機動/隊長/組長自由巡查")
                        temp_g = st.radio("步驟 A: 選擇年級", grades, horizontal=True, key="m1_grade_select")
                        f_cls_list = [c["name"] for c in structured_classes if c["grade"] == temp_g]
                        sel_cls = st.radio("步驟 B: 選擇班級", f_cls_list, horizontal=True, key="m1_cls_select") if f_cls_list else None

                    if sel_cls:
                        st.divider()
                        # [Fix #6] 雙重防呆：先查 session_state（即時），再查快取 df（補強）
                        _submit_key_check = f"{input_date}__{inspector_name}__{sel_cls}"
                        _already_in_session = _submit_key_check in st.session_state.submitted_inspections
                        _already_in_sheet   = check_duplicate_record(main_df, input_date, inspector_name, role, sel_cls)
                        if _already_in_session or _already_in_sheet:
                            st.warning(f"⚠️ 今日已評過 {sel_cls}！")

                        # [關鍵] check_result 放在 form 外面，才能即時顯示/隱藏違規欄位
                        check_result = st.radio("檢查結果", ["⭐ 優良", "✅ 普通", "❌ 違規(需扣分)"], horizontal=True, key="m1_check_result")

                        # 違規細節區塊（在 form 外，隨 radio 即時顯示）
                        in_s, out_s, ph_c, note, sel_violations = 0, 0, 0, "", []

                        if check_result == "⭐ 優良":
                            # [需求2] 優良備註欄 - 放在 form 外顯示說明
                            st.caption("⏳ 優良紀錄將送交組長審核，審核通過前學生端顯示為「普通」。")

                        elif check_result == "✅ 普通":
                            st.caption("✅ 無扣分，表現普通。如有需要可在下方填寫備註（選填）。")

                        elif check_result == "❌ 違規(需扣分)":
                            if role == "內掃檢查":
                                in_s = st.number_input("內掃扣分", min_value=0, step=1, key="m1_in_s")
                                st.markdown("**📍 違規位置（可複選）**")
                                INNER_AREA_OPTIONS = ["走廊", "黑板", "地板", "窗戶(窗溝)", "陽台"]
                                sel_areas = st.multiselect("違規位置", INNER_AREA_OPTIONS, key="inner_areas")
                                st.markdown("**⚠️ 違規狀況（可複選）**")
                                INNER_STATUS_OPTIONS = ["髒亂", "沒拖地", "沒擦拭", "酒精未補", "掃具壞掉未換", "懸掛垃圾未清除", "人工垃圾", "蜘蛛網", "頭髮圈圈", "打掃玩手機"]
                                sel_violations = st.multiselect("違規狀況", INNER_STATUS_OPTIONS, key="inner_status")
                                extra_note = st.text_input("📝 其他補充（找不到對應選項時請在此輸入）", key="m1_extra_note")
                                note_parts = []
                                if sel_areas: note_parts.append("位置：" + "、".join(sel_areas))
                                if sel_violations: note_parts.append("狀況：" + "、".join(sel_violations))
                                if extra_note: note_parts.append(extra_note)
                                note = " | ".join(note_parts)

                            elif role == "外掃檢查":
                                out_s = st.number_input("外掃扣分", min_value=0, step=1, key="m1_out_s")
                                st.markdown("**📍 違規位置**")
                                loc_col1, loc_col2 = st.columns(2)
                                BUILDING_OPTIONS = ["", "誠信樓A棟(各處室)", "誠信樓B棟", "樸實樓(合作社)", "勤學樓(烘焙縫紉)", "敬業樓(圖書館)"]
                                FLOOR_MAP = {
                                    "誠信樓A棟(各處室)": ["", "1F", "2F", "3F", "4F", "5F", "6F"],
                                    "誠信樓B棟":         ["", "1F", "2F", "3F", "4F", "5F", "6F"],
                                    "樸實樓(合作社)":    ["", "1F", "2F", "3F", "4F", "5F"],
                                    "勤學樓(烘焙縫紉)": ["", "1F", "2F", "3F", "4F", "5F"],
                                    "敬業樓(圖書館)":   ["", "1F", "2F", "3F"],
                                }
                                sel_building = loc_col1.selectbox("大樓", BUILDING_OPTIONS, key="m1_building")
                                floor_opts = FLOOR_MAP.get(sel_building, ["", "1F", "2F", "3F", "4F", "5F", "6F"])
                                sel_floor = loc_col2.selectbox("樓層", floor_opts, key="m1_floor")
                                st.markdown("**⚠️ 違規項目（可複選）**")
                                OUTER_AREA_OPTIONS = ["男廁", "女廁", "茶水間", "無障礙廁所", "樓梯間", "洗手台", "天花板", "走廊", "地板", "陽台"]
                                sel_outer_areas = st.multiselect("違規項目", OUTER_AREA_OPTIONS, key="outer_areas")
                                st.markdown("**⚠️ 違規狀況（可複選）**")
                                OUTER_STATUS_OPTIONS = ["髒亂", "沒拖地", "沒掃地", "沒擦拭", "酒精未補", "掃具壞掉未換", "人工垃圾", "蜘蛛網", "頭髮圈圈", "打掃玩手機"]
                                sel_status = st.multiselect("違規狀況", OUTER_STATUS_OPTIONS, key="outer_status")
                                extra_note = st.text_input("📝 其他補充（找不到對應選項時請在此輸入）", key="m1_extra_note")
                                note_parts = []
                                if sel_building: note_parts.append(sel_building)
                                if sel_floor: note_parts.append(sel_floor)
                                if sel_outer_areas: note_parts.append("項目：" + "、".join(sel_outer_areas))
                                if sel_status: note_parts.append("狀況：" + "、".join(sel_status))
                                if extra_note: note_parts.append(extra_note)
                                note = " | ".join(note_parts)
                                sel_violations = sel_outer_areas + sel_status

                        # form 只負責：修正單勾選 + 照片上傳 + 送出按鈕
                        excellent_note_form = ""  # 預設值，避免未定義
                        normal_note_form = ""     # 普通備註預設值
                        with st.form("score_form", clear_on_submit=True):
                            is_fix = st.checkbox("🚩 這是修正單")
                            # 優良備註在 form 內，避免 session_state key 問題
                            if check_result == "⭐ 優良":
                                st.markdown("**📝 優良原因（選填）**")
                                excellent_note_form = st.text_input("請簡單描述優良原因，例如：地板乾淨、掃具整齊", key="m1_excellent_note_form")
                            # 普通備註在 form 內
                            elif check_result == "✅ 普通":
                                st.markdown("**📝 備註（選填）**")
                                normal_note_form = st.text_input("可填寫備註，例如：掃得不夠乾淨但未達扣分標準，下次請加強", key="m1_normal_note_form")
                            # [V6 照片脫鉤] 照片改上傳至糾察社群群組，系統端不再收照片，送出秒完成
                            st.info("📸 照片請上傳至【糾察社群群組】：無論髒或乾淨都要拍照（註明掃區＋日期），以示負責。系統端不需附照，送出更快！")
                            files = None  # [V6] 系統端不再上傳照片，save_entry 收到 None 即略過 Drive 上傳

                            if st.form_submit_button("送出"):
                                # [Patch 11] 三層防重複提交：
                                #   Layer 1: session_state submit_key（最快，防 Streamlit 重複 rerun）
                                #   Layer 2: last_action_time 5秒防抖（防手動連點）
                                #   Layer 3: Worker 端 _append_main_entry_row dedup（最終防線）
                                _submit_key = f"{input_date}__{inspector_name}__{sel_cls}"
                                if _submit_key in st.session_state.submitted_inspections:
                                    st.warning("⚠️ 此筆已送出，請勿重複提交！若需修正請使用修正單。")
                                elif time.time() - st.session_state.last_action_time < 5:
                                    st.warning("⚠️ 系統處理中，請稍候 5 秒再試！")
                                else:
                                    # [關鍵] 先標記、再執行 — 防止 Streamlit 雙重 rerun 競爭
                                    st.session_state.submitted_inspections.add(_submit_key)
                                    st.session_state.last_action_time = time.time()

                                    _save_ok = False
                                    if check_result == "⭐ 優良":
                                        try:
                                            _exc_note = excellent_note_form.strip() if excellent_note_form else ""
                                        except Exception:
                                            _exc_note = ""
                                        _note_text = f"優良原因：{_exc_note}" if _exc_note else "本次檢查表現優良，無扣分項目（待組長審核）"
                                        _save_ok = save_entry({"日期": input_date, "週次": week_num, "檢查人員": inspector_name, "登錄時間": now_tw.strftime("%Y-%m-%d %H:%M:%S"), "修正": is_fix, "班級": sel_cls, "評分項目": role + "(待審優良)", "內掃原始分": 0, "外掃原始分": 0, "垃圾原始分": 0, "垃圾內掃原始分": 0, "垃圾外掃原始分": 0, "手機人數": 0, "備註": _note_text}, uploaded_files=files, award_inspector_hours=False)  # [V6] 時數改由「衛生點名」發放，評分不再自動給時數
                                    elif check_result == "✅ 普通":
                                        try:
                                            _norm_note = normal_note_form.strip() if normal_note_form else ""
                                        except Exception:
                                            _norm_note = ""
                                        _norm_note_text = _norm_note if _norm_note else "本次檢查無扣分，表現普通"
                                        _save_ok = save_entry({"日期": input_date, "週次": week_num, "檢查人員": inspector_name, "登錄時間": now_tw.strftime("%Y-%m-%d %H:%M:%S"), "修正": is_fix, "班級": sel_cls, "評分項目": role + "(普通)", "內掃原始分": 0, "外掃原始分": 0, "垃圾原始分": 0, "垃圾內掃原始分": 0, "垃圾外掃原始分": 0, "手機人數": 0, "備註": _norm_note_text}, uploaded_files=files, award_inspector_hours=False)  # [V6] 時數改由「衛生點名」發放，評分不再自動給時數
                                    elif check_result == "❌ 違規(需扣分)":
                                        _deduct_note = (f"【警告，扣0分】{note}".strip()) if (in_s + out_s) == 0 else note
                                        _save_ok = save_entry({"日期": input_date, "週次": week_num, "檢查人員": inspector_name, "登錄時間": now_tw.strftime("%Y-%m-%d %H:%M:%S"), "修正": is_fix, "班級": sel_cls, "評分項目": role, "內掃原始分": in_s, "外掃原始分": out_s, "手機人數": ph_c, "備註": _deduct_note, "違規細項": "、".join(sel_violations) if sel_violations else ""}, uploaded_files=files, award_inspector_hours=False)  # [V6] 時數改由「衛生點名」發放，評分不再自動給時數

                                    if _save_ok:
                                        if check_result == "⭐ 優良":
                                            st.success("⭐ 優良紀錄已送出！等待組長審核中...")
                                        elif check_result == "✅ 普通":
                                            st.success("✅ 普通紀錄已登記！")
                                        elif check_result == "❌ 違規(需扣分)":
                                            if assigned_classes:
                                                if is_last_task:
                                                    st.success("✅ 送出成功！今日任務已全數完成，系統將自動核發 0.25 小時！")
                                                else:
                                                    st.success(f"✅ 送出成功！尚缺 {len(pending_classes)-1} 個班級，請繼續努力！")
                                            else:
                                                st.success("✅ 送出成功！系統將自動排程發放本日 0.25 小時。")
                                            st.caption("📡 若送出後畫面沒有反應，請稍候 30 秒再試一次，不要連續按多次。")
                                        time.sleep(1.5)
                                        st.rerun()
                                    else:
                                        # 送出失敗 → 移除標記，允許重試
                                        st.session_state.submitted_inspections.discard(_submit_key)

    # --- Mode 2: 班級負責人 ---
    elif app_mode == "班級負責人🥸":
        st.title("🔎 班級成績查詢")
        df, appeals_df = load_full_semester_data_for_export(), load_appeals()  # 班級負責人需要整學期資料才能正確顯示學期累計扣分
        appeal_map = {str(r.get("對應紀錄ID")): {"status": str(r.get("處理狀態", "")), "reply": str(r.get("審核回覆", ""))} for _, r in appeals_df.iterrows()} if not appeals_df.empty else {}

        sel_grade_m2 = st.radio("選擇年級", grades, horizontal=True, key="m2_grade_select")
        cls_opts = [c["name"] for c in structured_classes if c["grade"] == sel_grade_m2]

        if cls_opts:
            cls = st.selectbox("選擇班級", cls_opts, key="m2_cls_select")
            if cls and not df.empty:
                cls_df = df[df["班級"] == cls].copy()

                # ── [效能] 預計算申訴期限，不在每筆 loop 裡重複呼叫 ──
                holidays = load_holidays()
                def _appeal_deadline(vd):
                    try:
                        current_date, workdays = (pd.to_datetime(str(vd)).date() if isinstance(vd, str) else vd), 0
                        for _ in range(14):
                            if workdays >= 3: break
                            current_date += timedelta(days=1)
                            if current_date.weekday() < 5 and current_date not in holidays: workdays += 1
                        return current_date
                    except Exception: return date.today()

                unique_dates = cls_df["日期"].astype(str).unique()
                appeal_deadline_map = {d: _appeal_deadline(d) for d in unique_dates}

                # ── 摘要卡片 ──
                now_week = get_week_num(today_tw)
                week_df = cls_df[cls_df["週次"] == now_week] if "週次" in cls_df.columns else pd.DataFrame()

                def _calc_tot(r):
                    tr = r['垃圾內掃原始分'] + r['垃圾外掃原始分']
                    if tr == 0: tr = r['垃圾原始分']
                    return r['內掃原始分'] + r['外掃原始分'] + tr + r['晨間打掃原始分']

                def _is_real_deduct(r):
                    item = str(r['評分項目'])
                    # 排除優良、普通、學期加分（晨掃），以及申訴已核可的修正紀錄
                    # 注意：load_main_data 將修正欄位轉為布林值 True/False
                    is_corrected = r['修正'] is True or str(r['修正']).upper() == "TRUE"
                    return (not any(x in item for x in ["優良", "普通", "學期加分"])
                            and not is_corrected)

                def _is_bonus(r):
                    # 學期加分紀錄（晨間打掃(學期加分)，分數為負）
                    return "學期加分" in str(r['評分項目']) and _calc_tot(r) < 0

                # 本週：只算正數扣分
                week_deduct = sum(max(_calc_tot(r), 0) for _, r in week_df.iterrows()
                                  if _is_real_deduct(r)) if not week_df.empty else 0
                # 學期：扣分與加分分開
                total_deduct = max(sum(max(_calc_tot(r), 0) for _, r in cls_df.iterrows() if _is_real_deduct(r)), 0)
                total_bonus  = sum(abs(_calc_tot(r)) for _, r in cls_df.iterrows() if _is_bonus(r))

                pending_appeals = sum(1 for rid in [str(r['紀錄ID']) for _, r in cls_df.iterrows()]
                                      if appeal_map.get(rid, {}).get("status") == "待處理")

                mc1, mc2, mc3 = st.columns(3)
                mc1.markdown(
                    f"<div style='background:#f0f7ff;border-radius:12px;padding:14px 16px;border-left:4px solid #3182ce'>"
                    f"<div style='font-size:12px;color:#555;margin-bottom:4px'>本週扣分</div>"
                    f"<div style='font-size:26px;font-weight:700;color:{'#e53e3e' if week_deduct>0 else '#38a169'}'>{week_deduct} 分</div>"
                    f"</div>", unsafe_allow_html=True)
                mc2.markdown(
                    f"<div style='background:#f0f7ff;border-radius:12px;padding:14px 16px;border-left:4px solid #805ad5'>"
                    f"<div style='font-size:12px;color:#555;margin-bottom:4px'>待處理申訴</div>"
                    f"<div style='font-size:26px;font-weight:700;color:{'#d69e2e' if pending_appeals>0 else '#38a169'}'>{pending_appeals} 件</div>"
                    f"</div>", unsafe_allow_html=True)
                # 第三張卡片：有加分時同時顯示扣分與加分
                if total_bonus > 0:
                    mc3_content = (
                        f"<div style='font-size:12px;color:#555;margin-bottom:4px'>學期扣分 / 加分</div>"
                        f"<div style='font-size:22px;font-weight:700;color:#e53e3e'>{total_deduct} 分</div>"
                        f"<div style='font-size:14px;font-weight:600;color:#38a169;margin-top:2px'>🌟 加分 {total_bonus} 分</div>"
                    )
                else:
                    mc3_content = (
                        f"<div style='font-size:12px;color:#555;margin-bottom:4px'>學期累計扣分</div>"
                        f"<div style='font-size:26px;font-weight:700;color:{'#e53e3e' if total_deduct>0 else '#38a169'}'>{total_deduct} 分</div>"
                    )
                mc3.markdown(
                    f"<div style='background:#f0f7ff;border-radius:12px;padding:14px 16px;border-left:4px solid #dd6b20'>"
                    f"{mc3_content}</div>", unsafe_allow_html=True)

                st.markdown("<div style='margin-top:16px'></div>", unsafe_allow_html=True)

                # ── 分頁 ──
                tab_records, tab_ranking = st.tabs(["📋 扣分明細", "🏆 各週排名"])

                with tab_records:
                    records = cls_df.sort_values("登錄時間", ascending=False)
                    if records.empty:
                        st.success("🎉 目前沒有任何評分紀錄！")
                    else:
                        import html as _html

                        # [修正] 同班+同日+同評分項目合併成一張卡片，避免多張照片變多筆
                        # 先把評分項目正規化（把 待審優良 統一視為 普通 顯示用）
                        def _display_item(item):
                            return str(item).replace("(待審優良)", "(普通)")

                        seen_groups = {}  # key=(日期, 評分項目顯示名) → 代表row index
                        group_photos = {}  # key → 所有照片 list
                        group_rids = {}   # key → 所有 rid list

                        for idx, r in records.iterrows():
                            date_str = str(r['日期'])
                            item_disp = _display_item(r['評分項目'])
                            gkey = (date_str, item_disp)
                            if gkey not in seen_groups:
                                seen_groups[gkey] = idx
                                group_photos[gkey] = []
                                group_rids[gkey] = []
                            # 合併照片
                            if str(r.get('照片路徑', '')).strip() and "http" in str(r['照片路徑']):
                                group_photos[gkey] += [p for p in str(r['照片路徑']).split(";") if "http" in p]
                            group_rids[gkey].append(str(r['紀錄ID']))

                        for gkey, rep_idx in seen_groups.items():
                            r = records.loc[rep_idx]
                            date_str, item_disp = gkey
                            rid = str(r['紀錄ID'])
                            all_photos = group_photos[gkey]
                            all_rids   = group_rids[gkey]

                            # 申訴狀態：只要有任何一筆 rid 有申訴，就顯示
                            ap_info = {}
                            for _rid in all_rids:
                                if _rid in appeal_map:
                                    ap_info = appeal_map[_rid]
                                    rid = _rid  # 用有申訴紀錄的那筆 rid 做申訴對應
                                    break
                            ap_st    = ap_info.get("status")
                            ap_reply = ap_info.get("reply")

                            trash_score = r['垃圾內掃原始分'] + r['垃圾外掃原始分']
                            if trash_score == 0: trash_score = r['垃圾原始分']
                            tot = r['內掃原始分'] + r['外掃原始分'] + trash_score + r['晨間打掃原始分']

                            is_excellent = "優良" in str(r['評分項目']) and "待審" not in str(r['評分項目'])
                            is_pending_excellent = "待審優良" in str(r['評分項目'])
                            is_normal = "普通" in str(r['評分項目']) or is_pending_excellent
                            is_corrected = r['修正'] is True or str(r['修正']).upper() == "TRUE"

                            # 決定卡片顏色
                            if ap_st == "已核可" or is_corrected:
                                card_color, border_color, tag_bg, tag_color = "#f0fff4","#38a169","#c6f6d5","#276749"
                                tag_text = "✅ 申訴成功" if ap_st == "已核可" else "🛠️ 已修正"
                            elif ap_st == "已駁回":
                                card_color, border_color, tag_bg, tag_color, tag_text = "#fff5f5","#fc8181","#fed7d7","#9b2c2c","🚫 申訴駁回"
                            elif ap_st == "待處理":
                                card_color, border_color, tag_bg, tag_color, tag_text = "#fffbeb","#f6ad55","#fefcbf","#744210","⏳ 申訴中"
                            elif is_excellent:
                                card_color, border_color, tag_bg, tag_color, tag_text = "#f0fff4","#68d391","#c6f6d5","#276749","⭐ 優良"
                            elif is_normal:
                                card_color, border_color, tag_bg, tag_color, tag_text = "#f7fafc","#a0aec0","#edf2f7","#4a5568","✅ 普通"
                            elif "學期加分" in str(r['評分項目']):
                                card_color, border_color, tag_bg, tag_color = "#f0fff4","#38a169","#c6f6d5","#276749"
                                tag_text = f"🌟 學期加 {abs(tot)} 分"
                            else:
                                card_color, border_color, tag_bg, tag_color = "#fff5f5","#fc8181","#fed7d7","#9b2c2c"
                                if tot < 0:
                                    card_color, border_color, tag_bg, tag_color = "#f0fff4","#38a169","#c6f6d5","#276749"
                                    tag_text = f"🌟 學期加 {abs(tot)} 分"
                                else:
                                    tag_text = f"❌ 扣 {tot} 分"

                            week_str = f"第{r.get('週次','')}週"
                            disp_time = str(r.get('登錄時間', ''))
                            inspector_safe = _html.escape(str(r.get('檢查人員', '未知')))
                            item_safe      = _html.escape(item_disp)
                            note_raw       = str(r.get('備註', '')).strip()
                            note_part      = f"&nbsp;|&nbsp; 📝 {_html.escape(note_raw)}" if note_raw else ""
                            time_part      = f'<div style="font-size:12px;color:#718096;margin-top:4px">登錄：{_html.escape(disp_time)}</div>' if disp_time else ""

                            st.markdown(
                                f"<div style='background:{card_color};border:1.5px solid {border_color};border-radius:12px;padding:12px 16px;margin-bottom:10px'>"
                                f"<div style='display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:6px'>"
                                f"<div style='font-weight:600;font-size:15px'>{date_str} <span style='color:#718096;font-size:13px'>{week_str}</span></div>"
                                f"<span style='background:{tag_bg};color:{tag_color};border-radius:20px;padding:3px 12px;font-size:13px;font-weight:600'>{tag_text}</span>"
                                f"</div>"
                                f"<div style='font-size:13px;color:#4a5568;margin-top:6px'>🧑‍✈️ {inspector_safe} &nbsp;|&nbsp; 📌 {item_safe}{note_part}</div>"
                                f"{time_part}"
                                f"</div>",
                                unsafe_allow_html=True
                            )

                            # 申訴回覆
                            if ap_st == "已核可":
                                st.success(f"✅ 申訴成功。組長回覆：{ap_reply if ap_reply else '無'}")
                            elif ap_st == "已駁回":
                                st.error(f"🚫 申訴駁回。組長回覆：{ap_reply if ap_reply else '無'}")
                            elif ap_st == "待處理":
                                st.info("⏳ 申訴審核中，請耐心等候...")

                            # 照片與申訴整合（合併所有照片）
                            has_photo = len(all_photos) > 0
                            deadline = appeal_deadline_map.get(date_str, date.today())
                            # [V6.1] 線上申訴管道已撤銷：改為現場申訴（3日內中午至衛生組），撤分由組長後台操作
                            can_appeal = False  # 冷凍線上申訴，保留程式碼以備未來恢復

                            if has_photo or can_appeal or (tot > 0):
                                with st.expander("📋 查看詳情" + ("　|　📣 可申訴（截止 " + deadline.strftime('%m/%d') + "）" if can_appeal else "")):
                                    if has_photo:
                                        st.markdown("**📷 評分照片**")
                                        st.image(all_photos, width=200)
                                    if tot > 0 and not can_appeal:
                                        st.info("📣 對扣分不服者，請於 **3 日內**之中午時間，由衛生股長至衛生組提出申訴。")
                                    if can_appeal:
                                        st.markdown("---")
                                        st.markdown("**📣 提出申訴**")
                                        with st.form(f"ap_{rid}"):
                                            rsn = st.text_area("申訴理由（必填）", key=f"rsn_{rid}")
                                            pf = st.file_uploader("佐證照片（選填）", type=['jpg','png'], key=f"pf_{rid}")
                                            if st.form_submit_button("送出申訴"):
                                                if time.time() - st.session_state.last_action_time < 3:
                                                    st.warning("⚠️ 系統處理中，請勿連續點擊！")
                                                elif not rsn:
                                                    st.error("請填寫申訴理由")
                                                else:
                                                    st.session_state.last_action_time = time.time()
                                                    if save_appeal({"班級": cls, "違規日期": date_str, "違規項目": r['評分項目'], "原始扣分": str(tot), "申訴理由": rsn, "對應紀錄ID": rid}, pf if pf else None):
                                                        time.sleep(1.5)
                                                        st.rerun()

                with tab_ranking:
                    pub_df = load_published_results()
                    if pub_df.empty:
                        st.info("📭 組長尚未發布任何週次的成績，請耐心等候！")
                    else:
                        available_pub_weeks = sorted(pub_df["週次"].unique(), reverse=True)
                        sel_pub_week = st.selectbox("選擇查詢週次", available_pub_weeks,
                                                    format_func=lambda w: f"第 {w} 週", key="m2_pub_week")
                        week_pub = pub_df[pub_df["週次"] == sel_pub_week]
                        cls_row = week_pub[week_pub["班級"] == cls]

                        pub_time = week_pub["發布時間"].iloc[0] if not week_pub.empty else ""

                        if not cls_row.empty:
                            cr = cls_row.iloc[0]
                            rank_val = int(cr["排名"])
                            score_val = int(cr["總成績"])
                            deduct_val = int(cr["總扣分"])
                            exc_val = int(cr["優良次數"])

                            # 用扣分決定卡片顏色，不顯示評等避免與糾察優良混淆
                            color = "#f0fff4" if deduct_val == 0 else ("#fffbeb" if deduct_val <= 3 else "#fff5f5")
                            border = "#38a169" if deduct_val == 0 else ("#f6ad55" if deduct_val <= 3 else "#fc8181")
                            st.markdown(
                                f"<div style='background:{color};border:2px solid {border};border-radius:14px;padding:20px 24px;text-align:center;margin-bottom:16px'>"
                                f"<div style='font-size:14px;color:#718096;margin-bottom:4px'>第 {sel_pub_week} 週 &nbsp;·&nbsp; {cls}</div>"
                                f"<div style='font-size:48px;font-weight:700;color:#2d3748'>#{rank_val}</div>"
                                f"<div style='font-size:15px;color:#4a5568;margin-top:6px'>"
                                f"總成績 {score_val} 分 &nbsp;|&nbsp; 扣分 {deduct_val} 分 &nbsp;|&nbsp; ⭐ 優良評分 {exc_val} 次"
                                f"</div></div>",
                                unsafe_allow_html=True
                            )
                        else:
                            st.warning(f"找不到 {cls} 在第 {sel_pub_week} 週的排名資料。")

                        # 直接讀取發布時儲存的排名模式，不再靠偵測推斷
                        pub_rank_mode = str(week_pub["排名模式"].iloc[0]) if "排名模式" in week_pub.columns and not week_pub.empty else "年級"

                        if pub_rank_mode == "全校":
                            st.markdown("##### 全校完整排名")
                            st.dataframe(
                                week_pub[["排名","年級","班級","總扣分","優良次數","總成績"]].sort_values("排名").reset_index(drop=True),
                                hide_index=True
                            )
                        else:
                            cls_grade = next((c["grade"] for c in structured_classes if c["name"] == cls), "")
                            grade_pub = week_pub[week_pub["年級"] == cls_grade] if cls_grade else week_pub
                            if not grade_pub.empty:
                                st.markdown(f"##### {cls_grade} 完整排名")
                                st.dataframe(
                                    grade_pub[["排名","班級","總扣分","優良次數","總成績"]].reset_index(drop=True),
                                    hide_index=True
                                )
                        if pub_time:
                            st.caption(f"發布時間：{pub_time}")

    # --- Mode 3: 晨掃志工隊🧹 ---
    elif app_mode == "晨掃志工隊🧹":
        st.title("🧹 晨掃志工回報專區")
        
        cutoff_hour = 24 if sys_env == "DEV" else 16
        
        if now_tw.hour >= cutoff_hour: 
            st.error("🚫 今日回報已截止 (16:00)")
        else:
            if sys_env == "DEV" and now_tw.hour >= 16:
                st.info("🔧 **[測試機特權開啟]** 目前已超過 16:00，但因為是 DEV 環境，允許繼續測試！")
                
            my_cls = st.selectbox("選擇班級", all_classes, key="m3_cls_select")
            main_df = load_main_data()
            # 晨掃填報只需近兩週資料
            _now_week = get_week_num(today_tw)
            if _now_week >= 3:
                main_df = main_df[main_df["週次"] >= _now_week - 2]
            
            # [新增防呆] 建立一個本地暫存，記住剛送出的班級
            if "just_submitted_morning" not in st.session_state:
                st.session_state.just_submitted_morning = []
                
            is_in_sheet = not main_df[(main_df["日期"].astype(str)==str(today_tw)) & (main_df["班級"]==my_cls) & (main_df["評分項目"].astype(str).str.contains("晨間打掃"))].empty
            is_just_submitted = f"{today_tw}_{my_cls}" in st.session_state.just_submitted_morning
            
            if is_in_sheet or is_just_submitted: 
                st.warning(f"⚠️ {my_cls} 今日已回報，或資料正在系統排隊處理中囉！")
            else:
                duty_df, _ = get_daily_duty(today_tw)
                
                has_duty = False 
                area_name_str = ""
                n_std = 4
                
                # 1. 先查今天有沒有排班
                if not duty_df.empty:
                    m_d = duty_df[duty_df["負責班級"]==my_cls]
                    if not m_d.empty:
                        has_duty = True
                        area_name_str = str(m_d.iloc[0].get('掃地區域', '未指定區域'))
                        try: n_std = int(m_d.iloc[0].get('標準人數', 4))
                        except Exception: n_std = 4
                
                is_makeup = False
                found_duty = has_duty

                # 2. 如果今天沒排班，且是一、二年級 -> 往前翻找本週班表
                if not has_duty and ("一" in my_cls or "二" in my_cls):
                    from datetime import timedelta
                    
                    # [V5.30 Patch 1] 防呆：先檢查本週是不是已經交過晨掃了！(不管有沒有被核可)
                    start_of_week = today_tw - timedelta(days=today_tw.weekday())
                    already_done = False
                    for _, r in main_df[main_df["班級"] == my_cls].iterrows():
                        if "晨間打掃" in str(r["評分項目"]):
                            try:
                                r_date = pd.to_datetime(str(r["日期"])).date()
                                if start_of_week <= r_date <= today_tw:
                                    already_done = True
                                    break
                            except Exception: pass  # 日期解析失敗忽略

                    # 如果本週「還沒交過」，才開啟時光機去查哪一天缺交
                    if not already_done:
                        for d in range(1, 7):
                            past_date = today_tw - timedelta(days=d)
                            if past_date < start_of_week: 
                                break # 只找本週的紀錄，超過本週就不補了
                                
                            p_duty, _ = get_daily_duty(past_date)
                            if not p_duty.empty and "負責班級" in p_duty.columns:
                                m_p = p_duty[p_duty["負責班級"].astype(str)==my_cls]
                                if not m_p.empty:
                                    area_name_str = str(m_p.iloc[0].get('掃地區域', '未指定區域'))
                                    try: n_std = int(m_p.iloc[0].get('標準人數', 4))
                                    except Exception: n_std = 4
                                    is_makeup = True # 判定為跨日補掃
                                    found_duty = True
                                    break
                                
                # 3. 如果今天有排班，但超過 15:00 -> 當日遲交補掃
                if has_duty and now_tw.hour >= 15:
                    is_makeup = True

                # 若完全找不到班表 (例如三年級沒排班，或一二年級連上週都沒排)
                if not found_duty:
                    st.success(f"🎉 恭喜！系統顯示 **{my_cls}** 近期沒有被分配到晨掃任務，好好休息吧！")
                    st.balloons()
                else:
                    areas = [a.strip() for a in area_name_str.split('、') if a.strip()]
                    if not areas: areas = ["打掃區域"]
                    
                    # 依據是否為補掃，顯示不同的上方提示
                    if is_makeup:
                        st.info(f"💡 **{my_cls}** 進行補打掃任務。本班任務總應到: {n_std} 人\n\n*(補掃通過將給予學期總分 +1，並核發志工時數)*")
                    else:
                        st.info(f"📍 本班任務總應到: {n_std} 人")
                    
                    # 顯示每日廣播大聲公
                    daily_task = SYSTEM_CONFIG.get("daily_morning_task", "")
                    if daily_task:
                        formatted_task = daily_task.replace('\n', '<br>')
                        mascot_url = "https://drive.google.com/thumbnail?id=128ITPXtpGNuI-wLIt6p-qd4ZNNhCGbhd"
                        
                        bubble_html = f"""
                        <style>
                        .mascot-container {{ display: flex; align-items: flex-start; margin-bottom: 20px; gap: 15px; }}
                        .mascot-img {{ width: 160px; flex-shrink: 0; }}
                        .speech-bubble {{
                            position: relative; background: #FFF3CD; border-radius: 15px; padding: 15px 20px;
                            color: #664d03; font-size: 16px; box-shadow: 2px 4px 10px rgba(0,0,0,0.1); border: 2px solid #ffecb5; flex-grow: 1; 
                        }}
                        .speech-bubble::before {{ content: ''; position: absolute; left: -20px; top: 30px; width: 0; height: 0; border: 10px solid transparent; border-right-color: #ffecb5; }}
                        .speech-bubble::after {{ content: ''; position: absolute; left: -16px; top: 30px; width: 0; height: 0; border: 10px solid transparent; border-right-color: #FFF3CD; }}
                        @media (max-width: 500px) {{
                            .mascot-img {{ width: 120px; }}
                            .speech-bubble {{ font-size: 14px; padding: 10px 15px; }}
                        }}
                        </style>
                        <div class="mascot-container">
                            <img src="{mascot_url}" class="mascot-img" />
                            <div class="speech-bubble">
                                <strong>📢 組長廣播 / 今日任務：</strong><br>
                                {formatted_task}
                            </div>
                        </div>
                        """
                        st.markdown(bubble_html, unsafe_allow_html=True)
                    
                    with st.form("vol_form"):
                        st.write("請依照下方分配的區域，分別填寫打掃同學並上傳照片：")
                        
                        present_dict = {}
                        files_dict = {}
                        class_roster = [s for s, c in ROSTER_DICT.items() if c == my_cls]
                        
                        for idx, area in enumerate(areas):
                            with st.container(border=True):
                                st.markdown(f"#### 🏷️ 區域 {idx+1}: **{area}**")
                                col1, col2 = st.columns(2)
                                with col1:
                                    present_dict[area] = st.multiselect(f"✅ 負責此區同學", class_roster, key=f"ms_{idx}")
                                with col2:
                                    files_dict[area] = st.file_uploader(f"📸 {area} 成果照片", accept_multiple_files=True, type=['jpg','png'], key=f"fu_{idx}")
                                    
                        # 依據 is_makeup 動態變更按鈕文字
                        btn_text = "🚀 我們完成補打掃了喔" if is_makeup else "🚀 確認送出全部回報"
                        
                        if st.form_submit_button(btn_text):
                            # [Patch 12] 晨掃也套用「先標記再執行」防重複提交
                            _morning_key = f"{today_tw}_{my_cls}"
                            if _morning_key in st.session_state.just_submitted_morning:
                                st.warning("⚠️ 此筆已送出，請勿重複提交！")
                            elif time.time() - st.session_state.last_action_time < 3:
                                st.warning("⚠️ 系統處理中，請勿連續點擊！")
                            else:
                                # 先標記，再執行
                                st.session_state.just_submitted_morning.append(_morning_key)
                                st.session_state.last_action_time = time.time()
                                
                                all_present = []
                                all_files = []
                                note_parts = []
                                
                                for area in areas:
                                    if present_dict[area]:
                                        all_present.extend(present_dict[area])
                                        note_parts.append(f"[{area}]負責:{','.join(present_dict[area])}")
                                    if files_dict[area]:
                                        all_files.extend(files_dict[area])
                                        
                                all_present = list(set(all_present))
                                final_note = " | ".join(note_parts)
                                
                                if not all_present or not all_files:
                                    st.error("❌ 請至少選擇一位打掃同學，並上傳至少一張照片！")
                                    # 驗證失敗 → 移除標記允許重試
                                    st.session_state.just_submitted_morning.remove(_morning_key)
                                else:
                                    # 依據 is_makeup 變更存入的任務名稱
                                    task_name = "晨間打掃(補掃)" if is_makeup else "晨間打掃"
                                    # [給分自動計算] 把應到人數也存進備註，供組長審核時自動算分
                                    score_note = f"[應到:{n_std}人 實到:{len(all_present)}人 {'補掃' if is_makeup else '準時'}] {final_note}"

                                    ok = save_entry(
                                        {
                                            "日期": str(today_tw), 
                                            "週次": get_week_num(today_tw),
                                            "班級": my_cls, 
                                            "評分項目": task_name, 
                                            "檢查人員": f"志工(實到:{len(all_present)})", 
                                            "登錄時間": now_tw.strftime("%Y-%m-%d %H:%M:%S"), 
                                            "晨間打掃原始分": 0, 
                                            "備註": score_note
                                        }, 
                                        uploaded_files=all_files, 
                                        student_list=all_present, 
                                        custom_hours=0.5, 
                                        custom_category="晨掃志工"
                                    )
                                    if ok:
                                        st.success("✅ 回報成功！所有區域皆已記錄，辛苦了！")
                                        time.sleep(1.5)
                                        st.rerun()
                                    else:
                                        # 送出失敗 → 移除標記允許重試
                                        st.session_state.just_submitted_morning.remove(_morning_key)
    # --- Mode 4: 組長後台 ---
    elif app_mode == "組長ㄉ窩💃":
        st.title("⚙️ 管理後台")
        metrics = get_queue_metrics()
        hb_sec = get_worker_heartbeat_sec()
        ls_sec = get_last_success_sec()
        
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("待處理", metrics.get("pending", 0))
        col2.metric("失敗", metrics.get("failed", 0))
        col3.metric("延遲(s)", int(metrics.get("oldest_pending_sec", 0)))
        
        hb_status = "🟢 正常運作" if hb_sec < 300 else "🔴 已休眠/停止"
        is_dry_run = str(st.secrets.get("system_config", {}).get("dry_run", "false")).lower() in ["true", "1"]
        
        if is_dry_run: hb_status = "🟡 演習模式 (Dry Run)"
        
        if ls_sec == 999999: ls_text = "無紀錄"
        elif ls_sec < 120: ls_text = f"✅ {int(ls_sec)}秒前"
        else: ls_text = f"⚠️ {int(ls_sec//60)}分鐘前 (API可能卡住)"
            
        col4.metric("背景 Worker", f"{hb_status}", f"心跳: {int(hb_sec)}秒前 | 成功: {ls_text}")
        
        last_err = get_last_error_summary()
        if last_err != "無紀錄":
            st.error(f"🚨 **最後錯誤紀錄:** {last_err}")

        pwd_input = st.text_input("管理密碼", type="password", key="admin_pwd")
        if pwd_input == st.secrets["system_config"]["admin_password"]:

            # Worker 狀態、Log、重啟按鈕（密碼驗證後才顯示）
            _ws = _get_worker_state()
            _t = _ws.get("thread")
            _alive = _t.is_alive() if _t else False
            _started = _ws.get("started_at", "未知")
            _all_bw = [t for t in threading.enumerate() if "background_worker" in t.name]
            st.caption(f"Worker：{'🟢 alive' if _alive else '🔴 dead'}　啟動時間：{_started}　threads：{len(_all_bw)} 個")
            if not _alive:
                st.warning("⚠️ Worker 執行緒未存活")
            if st.button("🔄 強制重啟 Worker", type="primary", key="restart_worker_btn"):
                _start_fresh_worker()
                st.success("✅ 新 Worker 已啟動！")
                st.rerun()
            if _WORKER_LOG:
                with st.expander("📋 Worker Log", expanded=False):
                    st.code("\n".join(reversed(list(_WORKER_LOG))), language=None)
            st.divider()

            t_dash, t_rollcall, t_appeal, t_excellent, t2, t_settings, t3, t_debt = st.tabs([
                "📈 儀表板與明細", "🙋 出勤點名", "📣 申訴", "⭐ 優良審核", "📊 成績總表", 
                "⚙️ 設定", "🎖️ 服務時數發放", "🤝 愛校與欠時管理"  # [V6.5] 移除監控與晨掃頁
            ])
            
            with t_dash:
                # [V6 新增] 扣分儀表板：每天開起來掃一眼，問題當天浮現，不用等週三結報
                st.subheader("📈 扣分儀表板")
                st.caption("💡 即時概況（含待審核與修正前的原始資料），正式成績仍以「📊 成績總表」的結算為準。")
                df_dash = load_main_data()
                if df_dash.empty:
                    st.info("目前尚無評分資料。")
                else:
                    d = df_dash.copy()
                    _score_cols = ["內掃原始分", "外掃原始分", "垃圾原始分", "垃圾內掃原始分", "垃圾外掃原始分", "晨間打掃原始分"]
                    for _c in _score_cols:
                        d[_c] = pd.to_numeric(d[_c], errors="coerce").fillna(0) if _c in d.columns else 0
                    d["扣分合計"] = d[_score_cols].sum(axis=1)
                    # [V6.4] 異常值隔離：單筆合計 > 20 視為資料異常（誤填學號/電話等），排除統計並提示
                    _bad = d[d["扣分合計"] > 20]
                    if not _bad.empty:
                        st.error("🚑 偵測到 " + str(len(_bad)) + " 筆分數異常的紀錄（單筆扣分 > 20，疑似誤填），已自動排除於統計之外。請至「📣 申訴與撤分」將其撤銷，或直接修正 main_data 分頁的分數欄：")
                        st.dataframe(_bad[["日期", "班級", "評分項目", "檢查人員", "扣分合計", "紀錄ID"]], hide_index=True, use_container_width=True)
                        d = d[d["扣分合計"] <= 20]
                    d["週次"] = pd.to_numeric(d.get("週次"), errors="coerce").fillna(0).astype(int)
                    _cur_w = get_week_num(today_tw)
                    _weeks = sorted([int(w) for w in d["週次"].unique() if w > 0], reverse=True)
                    if not _weeks:
                        st.info("目前尚無有效週次資料。")
                    else:
                        _sel_w = st.selectbox("檢視週次", _weeks, index=_weeks.index(_cur_w) if _cur_w in _weeks else 0, key="dash_week")
                        dw = d[(d["週次"] == _sel_w) & (d["扣分合計"] > 0)]
                        d4 = d[(d["週次"] >= max(1, _sel_w - 3)) & (d["週次"] <= _sel_w) & (d["扣分合計"] > 0)]

                        m1, m2, m3 = st.columns(3)
                        m1.metric(f"第 {_sel_w} 週扣分筆數", len(dw))
                        m2.metric(f"第 {_sel_w} 週扣分總點數", int(dw["扣分合計"].sum()))
                        m3.metric("被扣分班級數", int(dw["班級"].nunique()))

                        col_rank, col_item = st.columns(2)
                        with col_rank:
                            st.markdown(f"#### 🔻 各班扣分排行（第 {_sel_w} 週）")
                            if dw.empty:
                                st.success("🎉 本週目前無扣分紀錄！")
                            else:
                                import altair as alt  # [V6.4] 中文標籤直式顯示
                                _rank_df = dw.groupby("班級")["扣分合計"].sum().sort_values(ascending=False).reset_index()
                                _rank_df.columns = ["班級", "扣分"]
                                st.altair_chart(
                                    alt.Chart(_rank_df).mark_bar().encode(
                                        x=alt.X("班級:N", sort="-y", axis=alt.Axis(labelAngle=0, labelExpr="join(split(datum.label, ''), '\\n')", title=None, labelFontSize=13)),
                                        y=alt.Y("扣分:Q", title=None),
                                        tooltip=["班級", "扣分"]
                                    ).properties(height=320),
                                    use_container_width=True
                                )
                        with col_item:
                            st.markdown("#### 🧾 常見違規細項 Top 10（近四週）")
                            _items = []
                            for _v in d4.get("違規細項", pd.Series(dtype=str)).fillna(""):
                                _items += [x.strip() for x in str(_v).replace(",", "、").split("、") if x.strip()]
                            if _items:
                                import altair as alt  # [V6.4] 中文標籤直式顯示
                                _icnt_df = pd.Series(_items).value_counts().head(10).reset_index()
                                _icnt_df.columns = ["違規項目", "次數"]
                                st.altair_chart(
                                    alt.Chart(_icnt_df).mark_bar().encode(
                                        x=alt.X("違規項目:N", sort="-y", axis=alt.Axis(labelAngle=0, labelExpr="join(split(datum.label, ''), '\\n')", title=None, labelFontSize=13)),
                                        y=alt.Y("次數:Q", title=None),
                                        tooltip=["違規項目", "次數"]
                                    ).properties(height=320),
                                    use_container_width=True
                                )
                                st.caption("👉 從這裡看出全校最常發生的問題：是該「教方法」，還是該「加強盯場」。")
                            else:
                                st.info("近四週沒有違規細項紀錄。")

                        st.markdown("#### 📉 近四週各班扣分趨勢")
                        if d4.empty:
                            st.info("近四週沒有扣分紀錄。")
                        else:
                            _trend = d4.pivot_table(index="週次", columns="班級", values="扣分合計", aggfunc="sum").fillna(0)
                            _trend = _trend[_trend.sum().sort_values(ascending=False).head(8).index]
                            st.line_chart(_trend)
                            st.caption("僅顯示近四週扣分最多的前 8 個班級，避免線條過多。")

                        # 連續上榜警示：連續兩週進入扣分前五名
                        _warn_cls = []
                        try:
                            _w_list = [w for w in sorted(d["週次"].unique()) if 0 < w <= _sel_w][-2:]
                            if len(_w_list) == 2:
                                _tops = []
                                for _w in _w_list:
                                    _t = d[(d["週次"] == _w) & (d["扣分合計"] > 0)].groupby("班級")["扣分合計"].sum().sort_values(ascending=False).head(5)
                                    _tops.append(set(_t.index))
                                _warn_cls = sorted(_tops[0] & _tops[1])
                        except Exception:
                            _warn_cls = []
                        if _warn_cls:
                            st.warning("🚨 **連續兩週進入扣分前五名：** " + "、".join(_warn_cls) + "　→ 建議私下與導師聊聊，或安排糾察示範正確打掃方法。")


                # [V6.1] 扣分明細與督核單併入儀表板分頁
                st.markdown("---")
                st.subheader("📝 扣分明細")
                df = load_main_data()
                if not df.empty:
                    st.dataframe(df[["登錄時間", "日期", "班級", "評分項目", "檢查人員", "備註", "違規細項", "紀錄ID"]].sort_values("登錄時間", ascending=False))

                # [V6 新增] 督核單批次列印：選日期 → 產出可列印 HTML（純讀取，不寫入任何資料）
                st.markdown("---")
                with st.expander("🖨️ 督核單批次列印", expanded=False):
                    st.caption("選擇日期後，系統將該日所有扣分紀錄依班級彙整，每班一張督核單（A4 一頁兩張）。下載後用瀏覽器開啟 → Ctrl+P 列印。")
                    dk_date = st.date_input("督核單日期", today_tw, key="dk_date")
                    if df.empty:
                        st.info("目前尚無評分資料。")
                    else:
                        dkd = df.copy()
                        _dk_cols = ["內掃原始分", "外掃原始分", "垃圾原始分", "垃圾內掃原始分", "垃圾外掃原始分", "晨間打掃原始分"]
                        for _c in _dk_cols:
                            dkd[_c] = pd.to_numeric(dkd[_c], errors="coerce").fillna(0) if _c in dkd.columns else 0
                        dkd["扣分合計"] = dkd[_dk_cols].sum(axis=1)
                        dkd = dkd[(dkd["日期"].astype(str) == str(dk_date)) & (dkd["扣分合計"] > 0) & (dkd["扣分合計"] <= 20)]  # [V6.4] 排除異常值
                        if dkd.empty:
                            st.success(f"🎉 {dk_date} 沒有扣分紀錄，不需要印督核單。")
                        else:
                            import html as _html_esc  # [V6.2 資安] 學生輸入寫進 HTML 前先跳脫，防 XSS
                            _slips = []
                            for _cls, _g in dkd.groupby("班級"):
                                _rows_html = ""
                                for _, _r in _g.iterrows():
                                    _detail = str(_r.get("違規細項", "") or "").strip()
                                    _note = str(_r.get("備註", "") or "").strip()
                                    _desc = "、".join([x for x in [_detail, _note] if x]) or "（見評分紀錄）"
                                    _rows_html += f"<tr><td>{_html_esc.escape(str(_r.get('評分項目','')))}</td><td>{_html_esc.escape(_desc)}</td><td class='pt'>{_r['扣分合計']:g}</td><td>{_html_esc.escape(str(_r.get('檢查人員','')))}</td></tr>"
                                _slips.append(f"""
<div class='slip'>
  <h2>中壢家商 整潔缺失督核單</h2>
  <p class='meta'>班級：<b>{_html_esc.escape(str(_cls))}</b>　　日期：{dk_date}</p>
  <table>
    <thead><tr><th style='width:18%'>檢查項目</th><th>缺失事項</th><th style='width:10%'>扣分</th><th style='width:20%'>檢查糾察</th></tr></thead>
    <tbody>{_rows_html}</tbody>
  </table>
  <p class='sign'>隊長確認：＿＿＿＿＿＿＿　　衛生股長：＿＿＿＿＿＿＿　　<b>導師簽名：＿＿＿＿＿＿＿</b></p>
  <p class='appeal'>※ 對扣分不服者，請於 <b>3 日內</b>之中午時間至衛生組提出申訴。本單經導師簽名後，請衛生股長繳回衛生組。</p>
</div>""")
                            _dk_html = f"""<!DOCTYPE html><html lang='zh-Hant'><head><meta charset='UTF-8'><title>督核單 {dk_date}</title>
<style>
body{{font-family:'Noto Sans TC','Microsoft JhengHei',sans-serif;font-size:12pt;color:#2b2823;margin:0;background:#fff;}}
.slip{{box-sizing:border-box;height:138mm;padding:10mm 12mm;border-bottom:1px dashed #999;page-break-inside:avoid;}}
.slip:nth-child(2n){{page-break-after:always;}}
h2{{font-size:16pt;text-align:center;margin:0 0 4mm;letter-spacing:.1em;}}
.meta{{margin:0 0 2mm;}}
table{{width:100%;border-collapse:collapse;font-size:12pt;}}
th{{background:#fff;font-weight:800;text-align:left;padding:1.5mm 2mm;border-bottom:2.5px solid #2b2823;color:#1e6b50;}}
td{{padding:1.5mm 2mm;border-bottom:1px solid #d8d3c8;vertical-align:top;}}
td.pt{{font-weight:800;text-align:center;}}
.sign{{margin-top:6mm;}}
.appeal{{font-size:11pt;color:#6b6558;margin-top:2mm;}}
@media print{{@page{{size:A4;margin:0;}}}}
</style></head><body>{''.join(_slips)}</body></html>"""
                            st.download_button(
                                f"📥 下載督核單（{len(_slips)} 班，{dk_date}）",
                                _dk_html.encode("utf-8"),
                                file_name=f"督核單_{dk_date}.html",
                                mime="text/html",
                                key="dk_dl"
                            )

            with t_rollcall:
                # [V6.1] 衛生／環保點名合併為單一分頁，由使用者切換隊別
                st.subheader("🙋 出勤點名")
                rc_target = st.radio("點名隊別", ["🧹 衛生糾察", "♻️ 環保糾察 (資收場)"], horizontal=True, key="rc_target")
                st.markdown("---")
                if rc_target == "🧹 衛生糾察":
                    # [V6 新增] 衛生糾察出勤點名：出勤與評分紀錄脫鉤，以現場點名為準
                    st.info("💡 說明：衛生糾察每天全員應勤（機動人員除外）。一般糾察採【扣除法】勾缺席者；機動人員採【加入法】勾有到者。送出即發放當日時數。")

                    hyg_rc_date = st.date_input("出勤日期", today_tw, key="hyg_rc_date")
                    hyg_rc_hours = st.number_input("每人發放時數 (小時)", min_value=0.0, max_value=8.0, value=0.25, step=0.25, key="hyg_rc_hours")

                    def _is_hyg_member(p):
                        r = p.get("raw_role", "")
                        if "組長" in r or "環保" in r:
                            return False
                        return any(x in r for x in ["內掃", "外掃", "機動", "隊長"])

                    _hyg_all = [p for p in INSPECTOR_LIST if _is_hyg_member(p)]
                    _hyg_regular = [p for p in _hyg_all if "機動" not in p.get("raw_role", "")]
                    _hyg_mobile = [p for p in _hyg_all if "機動" in p.get("raw_role", "")]
                    reg_names = [p["label"] for p in _hyg_regular]
                    mob_names = [p["label"] for p in _hyg_mobile]

                    if not reg_names and not mob_names:
                        st.warning("⚠️ 目前名單中沒有衛生糾察成員，請確認 inspectors 分頁的「負責項目」欄位。")
                    else:
                        with st.form("hyg_rc_form"):
                            st.write(f"📋 每日應勤名單共 {len(reg_names)} 人")
                            hyg_absent = st.multiselect("❌ 勾選【請假 / 未到】的糾察 (扣除法)", reg_names)
                            hyg_present_reg = [n for n in reg_names if n not in hyg_absent]

                            hyg_present_mob = []
                            if mob_names:
                                hyg_present_mob = st.multiselect("🟠 機動人員【今日有出勤】者 (加入法)", mob_names)

                            _hyg_final = hyg_present_reg + hyg_present_mob
                            st.write(f"✅ 預計發放對象：共 {len(_hyg_final)} 人 (每人 {hyg_rc_hours} 小時)")

                            if st.form_submit_button("🚀 發放衛生糾察時數"):
                                if _block_future_date(hyg_rc_date, "出勤日期"):
                                    pass  # 未來日期被擋下，不繼續
                                elif time.time() - st.session_state.last_action_time < 3:
                                    st.warning("⚠️ 系統處理中，請勿連續點擊！")
                                else:
                                    st.session_state.last_action_time = time.time()
                                    hyg_ids = [name.split("學號:")[1].strip() for name in _hyg_final if "學號:" in name]
                                    if hyg_ids:
                                        payload = {
                                            "student_list": hyg_ids,
                                            "date": str(hyg_rc_date),
                                            "class_name": "糾察隊",
                                            "category": "衛生糾察值勤",
                                            "hours": float(hyg_rc_hours)
                                        }
                                        enqueue_task("service_hours_only", payload)
                                        st.success(f"✅ 已排程發放 {len(hyg_ids)} 人的出勤時數！(系統會自動阻擋同一天的重複發放)")
                                        time.sleep(1.5)
                                        st.rerun()
                                    else:
                                        st.warning("沒有可發放時數的對象")

                else:
                    st.info("💡 說明：此區專為資收場的環保糾察設計。先選班別，再勾選沒來的人，系統會自動幫有來的人發放時數。")
                
                    rc_date = st.date_input("出勤日期", today_tw, key="insp_rc_date")
                    # [V6 新增] 環保糾察拆分中午班 / 下午班，兩班人員不同、分別點名
                    rc_shift = st.radio("點名班別", ["🕛 中午班 (資收場整理)", "🕒 下午班 (垃圾管制)"], horizontal=True, key="insp_rc_shift")
                    shift_tag = "中午" if "中午" in rc_shift else "下午"
                    rc_hours = st.number_input("每人發放時數 (小時)", min_value=0.0, max_value=8.0, value=0.25, step=0.25, key="insp_rc_hours")
                
                    trash_inspectors = [p for p in INSPECTOR_LIST if "垃圾" in p.get("raw_role", "") or "回收" in p.get("raw_role", "") or "環保" in p.get("raw_role", "")]
                    _shift_tagged = [p for p in trash_inspectors if shift_tag in p.get("raw_role", "")]
                    if _shift_tagged:
                        trash_inspectors = _shift_tagged
                    else:
                        st.warning(f"⚠️ 名單中沒有人標註「{shift_tag}」班別，目前顯示全部環保糾察。建議至 Google Sheet 的 inspectors 分頁，在「負責項目」欄加註「中午」或「下午」，即可自動分班。")
                    insp_names = [p["label"] for p in trash_inspectors]
                
                    if not insp_names:
                        st.warning("⚠️ 目前名單中沒有負責「環保/垃圾/回收」的糾察。")
                    else:
                        with st.form("insp_rc_form"):
                            st.write(f"資收場糾察名單共 {len(insp_names)} 人")
                            absent_insps = st.multiselect("❌ 勾選【請假 / 未到】的糾察 (扣除法)", insp_names)
                            present_insps = [n for n in insp_names if n not in absent_insps]
                        
                            st.write(f"✅ 預計發放對象：共 {len(present_insps)} 人 (每人 {rc_hours} 小時)")
                        
                            if st.form_submit_button("🚀 發放環保糾察時數"):
                                if _block_future_date(rc_date, "出勤日期"):
                                    pass  # 未來日期被擋下，不繼續
                                elif time.time() - st.session_state.last_action_time < 3:
                                    st.warning("⚠️ 系統處理中，請勿連續點擊！")
                                else:
                                    st.session_state.last_action_time = time.time()
                                    present_ids = [name.split("學號:")[1].strip() for name in present_insps if "學號:" in name]
                                    if present_ids:
                                        payload = {
                                            "student_list": present_ids,
                                            "date": str(rc_date),
                                            "class_name": "糾察隊",
                                            "category": f"資源回收糾察({shift_tag})",  # [V6] 班別入類別，同日兩班可各自發放
                                            "hours": float(rc_hours)
                                        }
                                        enqueue_task("service_hours_only", payload)
                                        st.success(f"✅ 已排程發放 {len(present_ids)} 人的出勤時數！(系統會自動阻擋同一天的重複發放)")
                                        time.sleep(1.5)
                                        st.rerun()
                                    else:
                                        st.warning("沒有可發放時數的對象")

            with t_appeal:
                st.subheader("📣 申訴與撤分")
                st.caption("線上申訴管道已關閉：申訴一律採現場制（扣分後 3 日內之中午，由衛生股長至衛生組提出）。申訴成立時，在下方直接撤銷該筆扣分。")

                # [V6.1] 組長直接撤分工具
                with st.container(border=True):
                    st.markdown("#### ↩️ 撤銷扣分")
                    _rv_df = load_main_data()
                    if _rv_df.empty:
                        st.info("目前尚無評分資料。")
                    else:
                        _rvd = _rv_df.copy()
                        _rv_cols = ["內掃原始分", "外掃原始分", "垃圾原始分", "垃圾內掃原始分", "垃圾外掃原始分", "晨間打掃原始分"]
                        for _c in _rv_cols:
                            _rvd[_c] = pd.to_numeric(_rvd[_c], errors="coerce").fillna(0) if _c in _rvd.columns else 0
                        _rvd["扣分合計"] = _rvd[_rv_cols].sum(axis=1)
                        _rvd["_dt"] = pd.to_datetime(_rvd["日期"], errors="coerce").dt.date
                        _rvd = _rvd[(_rvd["扣分合計"] > 0) & (_rvd["修正"] != True) & (_rvd["_dt"] >= today_tw - timedelta(days=14))]
                        if _rvd.empty:
                            st.info("近 14 天沒有可撤銷的扣分紀錄。")
                        else:
                            _rvd = _rvd.sort_values(["日期", "登錄時間"], ascending=False)
                            _rv_opts = {}
                            for _, _r in _rvd.iterrows():
                                _hint = str(_r.get("違規細項", "") or _r.get("備註", "") or "")[:20]
                                _rv_opts[f"{_r['日期']}｜{_r['班級']}｜{_r['評分項目']}｜扣 {_r['扣分合計']:g} 分｜{_hint}｜ID:{_r['紀錄ID']}"] = str(_r["紀錄ID"])
                            _rv_sel = st.selectbox("選擇要撤銷的扣分紀錄（近 14 天內）", list(_rv_opts.keys()), key="rv_sel")
                            _rv_reason = st.text_input("撤分原因（必填，會寫入該筆備註留下紀錄）", key="rv_reason", placeholder="例如：現場申訴成立／糾察誤評")
                            _rv_confirm = st.checkbox("我確認要撤銷這筆扣分（撤銷後成績自動重算）", key="rv_confirm")
                            if st.button("↩️ 確認撤分", key="rv_go"):
                                if not _rv_reason.strip():
                                    st.error("❌ 請填寫撤分原因")
                                elif not _rv_confirm:
                                    st.warning("⚠️ 請先勾選確認")
                                elif time.time() - st.session_state.last_action_time < 3:
                                    st.warning("⚠️ 系統處理中，請勿連續點擊！")
                                else:
                                    st.session_state.last_action_time = time.time()
                                    _tid = enqueue_task("revoke_record", {"record_id": _rv_opts[_rv_sel], "reason": _rv_reason.strip()})
                                    if _tid:
                                        st.success("✅ 已排入撤分佇列，約 10–30 秒後生效，儀表板與成績將自動重算。")
                                    else:
                                        st.error("❌ 排入佇列失敗，請重試或檢查網路連線。")

                st.markdown("---")
                st.markdown("#### 🗂️ 線上申訴審核（管道已關閉，此區僅處理先前遺留的案件）")

                # [Fix] 用 session_state 紀錄本地已排入佇列的申訴 ID，避免重複顯示
                if "queued_appeal_ids" not in st.session_state:
                    st.session_state.queued_appeal_ids = set()

                ap_df = load_appeals()
                pending_aps = ap_df[ap_df["處理狀態"]=="待處理"]
                # 過濾掉已排入佇列的
                if not pending_aps.empty and st.session_state.queued_appeal_ids:
                    pending_aps = pending_aps[~pending_aps["對應紀錄ID"].astype(str).isin(st.session_state.queued_appeal_ids)]
                
                if pending_aps.empty: 
                    st.success("目前無待審核的申訴案件。")
                else:
                    st.caption(f"共 {len(pending_aps)} 筆待審核，審核後系統將背景處理，不需等待。")
                    for i, r in pending_aps.iterrows():
                        with st.container(border=True):
                            c1, c2 = st.columns([3,2])
                            c1.write(f"### {r['班級']} | {r['違規項目']} (扣 {r['原始扣分']} 分)")
                            c1.write(f"**申訴理由**: {r['申訴理由']}")
                            c1.caption(f"違規日期: {r['違規日期']} | 申訴時間: {r['登錄時間']}")
                            
                            img_urls = str(r.get('佐證照片', ''))
                            if img_urls and "http" in img_urls:
                                c2.image([p for p in img_urls.split(";") if "http" in p], width=250)
                            else:
                                c2.info("無佐證照片")
                                
                            reply_text = c1.text_input("💬 審核回覆 (填寫後學生將在查詢頁面看到此說明)", key=f"reply_{i}")
                            
                            col_btn1, col_btn2 = c1.columns(2)
                            if col_btn1.button("✅ 核可並撤銷扣分", key=f"ok_{i}"): 
                                _tid = enqueue_task("appeal_review", {
                                    "record_id": str(r["對應紀錄ID"]),
                                    "status": "已核可",
                                    "reply_text": reply_text
                                })
                                if _tid:
                                    st.session_state.queued_appeal_ids.add(str(r["對應紀錄ID"]))
                                    c1.success("✅ 已排入佇列，系統將背景處理核可與撤銷扣分。")
                                else:
                                    c1.error("❌ 排入佇列失敗，請重試或檢查網路連線。")
                            if col_btn2.button("🚫 駁回維持原判", key=f"ng_{i}"): 
                                _tid = enqueue_task("appeal_review", {
                                    "record_id": str(r["對應紀錄ID"]),
                                    "status": "已駁回",
                                    "reply_text": reply_text
                                })
                                if _tid:
                                    st.session_state.queued_appeal_ids.add(str(r["對應紀錄ID"]))
                                    c1.info("已排入佇列，系統將背景處理駁回。")
                                else:
                                    c1.error("❌ 排入佇列失敗，請重試或檢查網路連線。")

                    if st.button("🔄 重新整理申訴列表", key="refresh_appeals"):
                        st.session_state.queued_appeal_ids.clear()
                        load_appeals.clear()
                        load_main_data.clear()
                        st.rerun()

            with t_excellent:
                st.subheader("⭐ 優良審核")
                ex_df = load_main_data()
                pending_ex = ex_df[ex_df["評分項目"].astype(str).str.contains("待審優良")] if not ex_df.empty else pd.DataFrame()

                if pending_ex.empty:
                    st.success("🎉 目前沒有待審核的優良紀錄！")
                else:
                    if "approved_excellent_ids" not in st.session_state:
                        st.session_state.approved_excellent_ids = set()

                    # [修正] 依 班級+日期+評分項目 群組，避免多張照片變多筆審核卡
                    pending_ex = pending_ex[~pending_ex["紀錄ID"].astype(str).isin(st.session_state.approved_excellent_ids)]
                    groups = list(pending_ex.groupby(["班級", "日期", "評分項目"]))
                    st.caption(f"共 {len(groups)} 筆待審核，審核後不會跳頁，可以繼續審核其他筆。")

                    for (cls_name, date_val, item_val), grp in groups:
                        r = grp.iloc[0]
                        all_rids = grp["紀錄ID"].astype(str).tolist()
                        all_photos = []
                        for _, gr in grp.iterrows():
                            if "http" in str(gr.get("照片路徑", "")):
                                all_photos += [p for p in str(gr["照片路徑"]).split(";") if "http" in p]

                        with st.container(border=True):
                            c1, c2, c3 = st.columns([2, 2, 1])
                            c1.write(f"⭐ **{cls_name}** | {r['檢查人員']}")
                            c1.caption(f"{date_val} | {item_val}")
                            c1.write(f"📝 {r['備註']}")
                            if grp.shape[0] > 1:
                                c1.caption(f"（共 {grp.shape[0]} 筆紀錄合併，{len(all_photos)} 張照片）")
                            if all_photos:
                                c2.image(all_photos[:6], width=120)

                            gid = all_rids[0]  # 用第一筆 ID 當按鈕 key

                            def _approve_group(rids, eval_suffix):
                                try:
                                    ws = get_worksheet(SHEET_TABS["main"])
                                    if not ws:
                                        st.error("❌ 無法連線至 Google Sheets，請稍後再試")
                                        return False
                                    id_list = ws.col_values(EXPECTED_COLUMNS.index("紀錄ID") + 1)
                                    # [Fix] 將 id_list 統一轉為 str 並 strip，避免型別不一致
                                    id_list = [str(v).strip() for v in id_list]
                                    success_count = 0
                                    for rid in rids:
                                        rid = str(rid).strip()
                                        if rid in id_list:
                                            ridx = id_list.index(rid) + 1
                                            matched = ex_df.loc[ex_df["紀錄ID"].astype(str).str.strip() == rid, "評分項目"]
                                            if not matched.empty:
                                                new_item = str(matched.iloc[0]).replace("(待審優良)", eval_suffix)
                                                ws.update_cell(ridx, EXPECTED_COLUMNS.index("評分項目") + 1, new_item)
                                            # [Fix] 只有成功更新 Sheet 後才標記為已處理
                                            st.session_state.approved_excellent_ids.add(rid)
                                            success_count += 1
                                        else:
                                            st.warning(f"⚠️ 紀錄 {rid[:12]}... 在 Sheet 中找不到，可能資料尚在佇列中，請稍後重新整理再試。")
                                    load_main_data.clear()
                                    return success_count > 0
                                except Exception as e:
                                    st.error(f"❌ 審核寫入失敗: {e}")
                                    return False

                            if c3.button("✅ 核可優良", key=f"ex_ok_{gid}"):
                                if _approve_group(all_rids, "(優良)"):
                                    c1.success("✅ 已核可為優良！")
                            if c3.button("🚫 駁回(改普通)", key=f"ex_ng_{gid}"):
                                if _approve_group(all_rids, "(普通)"):
                                    c1.info("已改為普通。")

                    if st.button("🔄 重新整理列表", key="refresh_excellent"):
                        st.session_state.approved_excellent_ids.clear()
                        load_main_data.clear()
                        st.rerun()

            with t2:
                st.subheader("📊 成績總表")
                full = load_full_semester_data_for_export()

                # ── 共用計算函式 ──────────────────────────────────────────
                def calc_scores(df_raw):
                    """回傳含結算欄位的 DataFrame（已排除優良/普通紀錄，以及申訴核可的修正紀錄）"""
                    # [Fix] 加括號修正運算子優先順序（& 優先於 ~，不加括號邏輯會錯）
                    df = df_raw[
                        (~df_raw["評分項目"].astype(str).str.contains("優良|普通")) &
                        (~df_raw["修正"].astype(str).str.upper().eq("TRUE"))   # 排除申訴已核可的紀錄
                    ].copy()
                    df["內掃結算"] = df["內掃原始分"].clip(upper=2)
                    df["外掃結算"] = df["外掃原始分"].clip(upper=2)
                    trash = df["垃圾內掃原始分"] + df["垃圾外掃原始分"]
                    trash = trash.where(trash > 0, df["垃圾原始分"])
                    df["垃圾結算"] = trash.clip(upper=2)
                    df["總扣分"] = df["內掃結算"] + df["外掃結算"] + df["垃圾結算"] + df["晨間打掃原始分"].clip(lower=0) + df["手機人數"]
                    return df

                def build_ranking(scored_df, df_all, classes_struct, base_score=90):
                    """依班級彙總扣分，合併完整班級清單，加上總成績與優良次數"""
                    rep = scored_df.groupby("班級")["總扣分"].sum().reset_index()
                    # [Fix] 計算優良次數：排除「待審優良」，只算正式核可的優良
                    excellent_mask = (
                        df_all["評分項目"].astype(str).str.contains("優良") &
                        ~df_all["評分項目"].astype(str).str.contains("待審")
                    )
                    excellent_counts = df_all[excellent_mask].groupby("班級").size().reset_index(name="優良次數")
                    cls_df = pd.DataFrame(classes_struct).rename(columns={"grade": "年級", "name": "班級"})
                    fin = pd.merge(cls_df, rep, on="班級", how="left").fillna(0)
                    fin = pd.merge(fin, excellent_counts, on="班級", how="left").fillna(0)
                    fin["總成績"] = base_score - fin["總扣分"]
                    fin["優良次數"] = fin["優良次數"].astype(int)
                    return fin

                def add_rank_and_label(fin_df, by_grade=False, threshold_good=3):
                    """排序並加上排名、評等欄位；同分以優良次數排名，使用標準競賽排名（並列同名次，下一名跳號）"""
                    def label(s):
                        if s == 0:   return "⭐ 優良"
                        if s <= threshold_good: return "✅ 普通"
                        return "⚠️ 需加強"

                    def apply_competition_rank(df):
                        # 先依「總成績高→優良次數多」排序
                        df = df.sort_values(["總成績", "優良次數"], ascending=[False, False]).reset_index(drop=True)
                        # 標準競賽排名：同名次並列，下一名跳號（例：1,1,1,4,5）
                        ranks = []
                        rank = 1
                        for i, row in df.iterrows():
                            if i == 0:
                                ranks.append(rank)
                            else:
                                prev = df.iloc[i-1]
                                if row["總成績"] == prev["總成績"] and row["優良次數"] == prev["優良次數"]:
                                    ranks.append(ranks[-1])  # 並列同名次
                                else:
                                    rank = i + 1  # 跳到實際位置號
                                    ranks.append(rank)
                        df.insert(0, "排名", ranks)
                        df["評等"] = df["總扣分"].apply(label)
                        return df

                    if not by_grade:
                        return apply_competition_rank(fin_df.copy())
                    else:
                        pieces = []
                        for g in sorted(fin_df["年級"].unique()):
                            if g == "其他": continue
                            pieces.append(apply_competition_rank(fin_df[fin_df["年級"]==g].copy()))
                        return pd.concat(pieces, ignore_index=True) if pieces else fin_df

                def build_detail(scored_df):
                    """各班扣分明細：每筆違規紀錄，含日期/週次/評分項目/扣分/備註"""
                    cols = ["日期", "週次", "班級", "評分項目", "檢查人員",
                            "內掃結算", "外掃結算", "垃圾結算", "手機人數", "總扣分", "備註", "違規細項"]
                    avail = [c for c in cols if c in scored_df.columns]
                    detail = scored_df[scored_df["總扣分"] > 0][avail].sort_values(["班級","日期"])
                    return detail

                def to_excel_bytes(sheets_dict):
                    """產生多分頁 Excel，sheets_dict = {分頁名稱: DataFrame}"""
                    buf = io.BytesIO()
                    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                        for sheet_name, df in sheets_dict.items():
                            safe_name = sheet_name[:31]  # Excel 分頁名稱上限 31 字
                            df.to_excel(writer, sheet_name=safe_name, index=False)
                            ws = writer.sheets[safe_name]
                            # 自動調整欄寬
                            for col_cells in ws.columns:
                                max_len = max((len(str(c.value)) if c.value else 0) for c in col_cells)
                                ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 40)
                    buf.seek(0)
                    return buf.getvalue()

                # ── UI ───────────────────────────────────────────────────
                if full.empty:
                    st.info("目前無評分資料")
                else:
                    tab_week, tab_semester = st.tabs(["📅 單週成績結算", "🏆 全學期總結算"])

                    with tab_week:
                        available_weeks = sorted([w for w in full["週次"].unique() if w > 0])
                        if not available_weeks:
                            st.warning("尚無有效的週次資料")
                        else:
                            sel_week = st.selectbox("請選擇結算週次", available_weeks, index=len(available_weeks)-1)
                            is_fall = (today_tw.month >= 8 or today_tw.month == 1)
                            default_mode = "年級 (上學期制)" if is_fall else "全校 (下學期制)"
                            st.info(f"💡 系統偵測目前為 **{'上' if is_fall else '下'}學期**，預設採用 **{default_mode}** 排名。")
                            rank_mode = st.radio("排名方式 (可手動更改)", ["年級", "全校"], index=0 if is_fall else 1, horizontal=True)
                            
                            col_calc, col_refresh = st.columns([3, 1])
                            if col_refresh.button("🔄 重新讀取資料", key="refresh_export", help="若有申訴剛核可，請先點此確保資料是最新的"):
                                load_main_data.clear()
                                st.success("✅ 資料已刷新！")
                                st.rerun()

                            if col_calc.button("🚀 計算並顯示當週成績"):
                                week_raw = full[full["週次"] == sel_week]  # 當週原始資料
                                scored = calc_scores(week_raw)
                                fin = build_ranking(scored, week_raw, structured_classes)  # 優良次數只算當週
                                by_grade = (rank_mode == "年級")
                                fin_ranked = add_rank_and_label(fin, by_grade=by_grade)
                                detail = build_detail(scored)
                                # 儲存供發布使用
                                st.session_state["last_computed_week"] = sel_week
                                st.session_state["last_computed_ranking"] = fin_ranked

                                # ── 畫面顯示 ──
                                if by_grade:
                                    for g in sorted(fin_ranked["年級"].unique()):
                                        st.write(f"#### {g} 排名")
                                        g_df = fin_ranked[fin_ranked["年級"]==g]
                                        st.dataframe(g_df[["排名","班級","總扣分","優良次數","總成績","評等"]], hide_index=True)
                                else:
                                    st.dataframe(fin_ranked[["排名","年級","班級","總扣分","優良次數","總成績","評等"]], hide_index=True)

                                st.markdown("---")
                                st.write(f"##### 📋 第 {sel_week} 週扣分明細（共 {len(detail)} 筆）")
                                if detail.empty:
                                    st.success("本週無任何扣分紀錄！")
                                else:
                                    st.dataframe(detail, hide_index=True)

                                # ── Excel 下載 ──
                                st.markdown("---")
                                sheets = {"排名總表": fin_ranked[["排名","年級","班級","總扣分","優良次數","總成績","評等"]]}
                                if by_grade:
                                    for g in sorted(fin_ranked["年級"].unique()):
                                        sheets[f"{g}排名"] = fin_ranked[fin_ranked["年級"]==g][["排名","班級","總扣分","優良次數","總成績","評等"]]
                                sheets["扣分明細"] = detail
                                excel_bytes = to_excel_bytes(sheets)
                                st.download_button(
                                    label=f"📥 下載第 {sel_week} 週成績報表 (Excel)",
                                    data=excel_bytes,
                                    file_name=f"衛生成績_第{sel_week}週_{today_tw.strftime('%Y%m%d')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )

                            # ── 發布按鈕（計算完後才會出現）──
                            if st.session_state.get("last_computed_week") == sel_week and \
                               st.session_state.get("last_computed_ranking") is not None:
                                st.markdown("---")
                                st.info(f"💡 計算完成後，可將第 {sel_week} 週成績發布給學生查詢。")
                                if st.button(f"📢 發布第 {sel_week} 週成績給學生", key="publish_week_btn"):
                                    ok, msg = publish_week_results(
                                        sel_week,
                                        st.session_state["last_computed_ranking"],
                                        rank_mode=rank_mode  # 把排名模式一起存進去
                                    )
                                    if ok:
                                        st.success(f"✅ {msg}")
                                    else:
                                        st.error(f"❌ 發布失敗：{msg}")

                    with tab_semester:
                        st.write("計算全學期累計總扣分與總成績")
                        sem_rank_mode = st.radio("學期排名方式", ["全校", "年級"], horizontal=True, key="sem_rank")

                        if st.button("🚀 計算並顯示全學期成績", key="sem_btn"):
                            scored = calc_scores(full)
                            fin = build_ranking(scored, full, structured_classes)
                            by_grade = (sem_rank_mode == "年級")
                            fin_ranked = add_rank_and_label(fin, by_grade=by_grade, threshold_good=10)
                            detail = build_detail(scored)

                            # ── 畫面顯示 ──
                            if by_grade:
                                for g in sorted(fin_ranked["年級"].unique()):
                                    st.write(f"#### {g} 排名")
                                    g_df = fin_ranked[fin_ranked["年級"]==g]
                                    st.dataframe(g_df[["排名","班級","總扣分","優良次數","總成績","評等"]], hide_index=True)
                            else:
                                st.dataframe(fin_ranked[["排名","年級","班級","總扣分","優良次數","總成績","評等"]], hide_index=True)

                            st.markdown("---")
                            st.write(f"##### 📋 全學期扣分明細（共 {len(detail)} 筆）")
                            if detail.empty:
                                st.success("學期內無任何扣分紀錄！")
                            else:
                                # 扣分明細太多時，提供依班級篩選
                                all_cls_options = ["全部班級"] + sorted(detail["班級"].unique().tolist())
                                sel_cls_filter = st.selectbox("篩選班級（可只看單班明細）", all_cls_options, key="sem_detail_filter")
                                detail_show = detail if sel_cls_filter == "全部班級" else detail[detail["班級"]==sel_cls_filter]
                                st.dataframe(detail_show, hide_index=True)

                            # ── Excel 下載（多分頁：總排名 + 各年級 + 各班明細） ──
                            st.markdown("---")
                            sheets = {"學期排名總表": fin_ranked[["排名","年級","班級","總扣分","優良次數","總成績","評等"]]}
                            if by_grade:
                                for g in sorted(fin_ranked["年級"].unique()):
                                    sheets[f"{g}排名"] = fin_ranked[fin_ranked["年級"]==g][["排名","班級","總扣分","優良次數","總成績","評等"]]
                            # 各班各週明細：樞紐表（班級 x 週次）
                            if not detail.empty:
                                pivot = detail.pivot_table(index="班級", columns="週次", values="總扣分", aggfunc="sum", fill_value=0)
                                pivot["學期總扣分"] = pivot.sum(axis=1)
                                pivot = pivot.reset_index()
                                sheets["各班週次扣分樞紐"] = pivot
                                sheets["全學期扣分明細"] = detail
                                # 每個年級各自一張扣分明細分頁
                                cls_df_map = pd.DataFrame(structured_classes).rename(columns={"grade":"年級","name":"班級"})
                                for g in sorted(cls_df_map["年級"].unique()):
                                    if g == "其他": continue
                                    g_classes = cls_df_map[cls_df_map["年級"]==g]["班級"].tolist()
                                    g_detail = detail[detail["班級"].isin(g_classes)]
                                    if not g_detail.empty:
                                        sheets[f"{g}扣分明細"] = g_detail
                            excel_bytes = to_excel_bytes(sheets)
                            st.download_button(
                                label="📥 下載全學期成績報表 (Excel，含多分頁)",
                                data=excel_bytes,
                                file_name=f"衛生成績_全學期_{today_tw.strftime('%Y%m%d')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
            with t_settings:
                st.subheader("⚙️ 系統設定與維護")
                curr = SYSTEM_CONFIG.get("semester_start")
                nd = st.date_input("開學日", datetime.strptime(curr, "%Y-%m-%d").date() if curr else today_tw)
                if st.button("更新開學日"): save_setting("semester_start", str(nd))
                
                st.markdown("---")
                st.write("📅 **週次手動對照表**（解決寒假跨週問題）")
                st.caption("只需填「錨點」：每個學期重置點的週一日期與週次號碼，用逗號分隔。\n\n例如：`2025-01-23:1,2025-02-23:2`\n\n這樣填即可：第1週從1/23起，第2週從2/23起，第3週之後系統會自動從2/23往後每7天累計，不需要填完所有週次。")
                curr_week_map = SYSTEM_CONFIG.get("week_map", "")
                new_week_map = st.text_area("週次對照表", value=curr_week_map, placeholder="2025-01-23:1,2025-02-23:2,2025-03-02:3,...")
                if st.button("💾 儲存週次對照表"):
                    if save_setting("week_map", new_week_map.strip()):
                        st.success("✅ 週次對照表已更新！")
                
                st.markdown("---")
                st.write("📢 晨掃志工每日廣播/任務")
                current_task = SYSTEM_CONFIG.get("daily_morning_task", "今日無特殊任務，請確實完成各區打掃即可！")
                new_task = st.text_area("請輸入想給志工看的話（例如：拍照請比 YA、今天請加強拖地等）", value=current_task, key="ta_morning_task")
                if st.button("💾 更新每日任務"):
                    if save_setting("daily_morning_task", new_task):
                        st.success("✅ 每日任務已更新！學生現在起會看到最新廣播。")

                st.markdown("---")

                st.write("📢 衛生糾察每日廣播/提醒")
                current_hygiene_task = SYSTEM_CONFIG.get("daily_hygiene_task", "今日無特殊任務，請確實完成各區檢查即可！")
                new_hygiene_task = st.text_area("請輸入想給糾察隊看的話（例如：今天重點檢查黑板、窗台）", value=current_hygiene_task, key="ta_hygiene_task")
                if st.button("💾 更新糾察任務"):
                    if save_setting("daily_hygiene_task", new_hygiene_task):
                        st.success("✅ 糾察任務已更新！糾察隊現在起會看到最新廣播。")

                st.markdown("---")
                st.write("🔧 系統連線狀態")
                if get_gspread_client(): st.success("✅ Google Sheets 連線正常")
                else: st.error("❌ Google Sheets 連線失敗")
                
                if NOTION_INSTALLED: st.success("✅ Notion 模組載入正常")
                else: st.warning("⚠️ 尚未安裝 Notion 模組")
                
                st.info("若需修改名單請直接至 Google Sheet 修改 inspectors / roster / office_areas 分頁")
                if st.button("🔄 重讀名單 (清除快取)"): st.cache_data.clear(); st.success("已清除快取！")

            with t3:
                st.subheader("🎖️ 服務時數發放")

                # ── 共用欄位 ──
                SVC_CATEGORIES = ["返校打掃", "校外服務", "社區服務", "班級義工", "其他（手動輸入）"]
                c_s1, c_s2 = st.columns(2)
                rd = c_s1.date_input("日期", today_tw, key="svc_date")
                if _block_future_date(rd, "服務時數發放日期"):
                    st.stop()
                sel_cat = c_s2.selectbox("活動類別", SVC_CATEGORIES, key="svc_cat")

                if sel_cat == "其他（手動輸入）":
                    custom_cat_input = st.text_input("請輸入活動名稱", key="svc_custom_cat", placeholder="例如：環境清潔日")
                    base_category = custom_cat_input.strip() if custom_cat_input.strip() else "其他"
                else:
                    base_category = sel_cat

                remark_input = st.text_input("備註（選填）", key="svc_remark", placeholder="例如：返校打掃第一梯次")
                # 備註塞入類別欄，格式：活動名稱｜備註內容（不改 Sheet 欄位結構）
                final_category = f"{base_category}｜{remark_input.strip()}" if remark_input.strip() else base_category

                st.markdown("---")
                target_mode = st.radio(
                    "發放對象模式",
                    ["🏫 班級模式", "🔢 直接輸入學號", "👮 衛生糾察隊全員", "🗑️ 資源回收糾察隊全員"],
                    horizontal=True, key="svc_mode"
                )
                st.markdown("")

                # ── 班級模式 ──
                if target_mode == "🏫 班級模式":
                    rc = st.selectbox("選擇班級", all_classes, key="svc_cls")
                    mems = [s for s, c_val in ROSTER_DICT.items() if c_val == rc]

                    if not mems:
                        st.warning("⚠️ 此班級在 Roster 中找不到成員，請確認 Google Sheet。")
                    else:
                        _is_return_clean = base_category in ("返校打掃",)  # 返校打掃啟用三類歸類

                        if _is_return_clean:
                            # ── [Patch 13] 返校打掃：三類歸類法 ──
                            st.info(f"📋 全班 {len(mems)} 人，請標記「免除」和「未到」的同學，其餘預設為出席。")

                            with st.form("svc_class_form"):
                                exempt = st.multiselect(
                                    f"🏃 免除名單（體育生等不需打掃的同學）",
                                    mems, key="svc_exempt"
                                )
                                remaining_after_exempt = [m for m in mems if m not in exempt]
                                absent_absent = st.multiselect(
                                    f"❌ 未到名單（需補時數 2hr 的同學）",
                                    remaining_after_exempt, key="svc_absent_debt"
                                )
                                pool = [m for m in remaining_after_exempt if m not in absent_absent]

                                _c1, _c2, _c3 = st.columns(3)
                                _c1.metric("✅ 出席", f"{len(pool)} 人")
                                _c2.metric("🏃 免除", f"{len(exempt)} 人")
                                _c3.metric("❌ 未到", f"{len(absent_absent)} 人")

                                st.markdown("---")
                                base_h = st.number_input("出席者時數", value=2.0, step=0.5, min_value=0.0, key="svc_base_h")
                                debt_h = st.number_input("未到者欠時數", value=2.0, step=0.5, min_value=0.0, key="svc_debt_h")

                                st.markdown("---")
                                spec = st.multiselect("⭐ 加強組（從出席者挑選，另給不同時數）", pool, key="svc_spec")
                                spec_h = st.number_input("加強組時數", value=3.0, step=0.5, min_value=0.0, key="svc_spec_h")
                                st.caption("（若無加強組，此欄留空即可，不影響結果）")

                                pf = st.file_uploader("📸 照片（選填）", type=['jpg', 'png', 'jpeg'], key="svc_pf_cls")

                                if st.form_submit_button("🚀 發放時數 + 登記未到欠時"):
                                    if time.time() - st.session_state.last_action_time < 3:
                                        st.warning("⚠️ 系統處理中，請勿連續點擊！")
                                    elif not pool and not absent_absent:
                                        st.error("❌ 出席和未到名單皆為空，請確認！")
                                    else:
                                        st.session_state.last_action_time = time.time()
                                        norm = [m for m in pool if m not in spec]
                                        ok_norm, ok_spec = True, True

                                        fb = None
                                        if pf:
                                            pf.seek(0)
                                            fb = pf.read()

                                        # 發放出席者時數
                                        if norm:
                                            files_norm = [io.BytesIO(fb)] if fb else None
                                            if files_norm: files_norm[0].name = "p.jpg"
                                            ok_norm = save_entry(
                                                {"日期": str(rd), "班級": rc, "評分項目": "服務時數發放",
                                                 "登錄時間": now_tw.strftime("%Y-%m-%d %H:%M:%S")},
                                                files_norm, norm, base_h, final_category
                                            )
                                        if spec:
                                            files_spec = [io.BytesIO(fb)] if fb else None
                                            if files_spec: files_spec[0].name = "p.jpg"
                                            ok_spec = save_entry(
                                                {"日期": str(rd), "班級": rc, "評分項目": "服務時數發放",
                                                 "登錄時間": now_tw.strftime("%Y-%m-%d %H:%M:%S")},
                                                files_spec, spec, spec_h, f"{final_category}(加強)"
                                            )

                                        # 自動寫入未到者欠時數
                                        debt_ok_count = 0
                                        debt_fail_sids = []
                                        if absent_absent and debt_h > 0:
                                            _debt_date_str = str(rd)
                                            _debt_reason = f"{_debt_date_str} 返校打掃未到"
                                            for _ab_sid in absent_absent:
                                                try:
                                                    _dok = update_student_debt(_ab_sid, debt_h, _debt_reason)
                                                    if _dok:
                                                        debt_ok_count += 1
                                                    else:
                                                        debt_fail_sids.append(_ab_sid)
                                                except Exception as _de:
                                                    print(f"[返校打掃欠時] {_ab_sid} 失敗: {_de}")
                                                    debt_fail_sids.append(_ab_sid)

                                        # 結果彙報
                                        result_parts = []
                                        if ok_norm and ok_spec:
                                            result_parts.append(f"✅ 已發放！出席 {len(norm)} 人 ({base_h}h)")
                                            if spec:
                                                result_parts.append(f"加強組 {len(spec)} 人 ({spec_h}h)")
                                        if exempt:
                                            result_parts.append(f"🏃 免除 {len(exempt)} 人（不處理）")
                                        if debt_ok_count > 0:
                                            result_parts.append(f"❌ 未到 {debt_ok_count} 人已登記欠時 {debt_h}hr")
                                        if debt_fail_sids:
                                            st.error(f"⚠️ 以下學號欠時寫入失敗，請手動處理：{', '.join(debt_fail_sids)}")
                                        st.success("　|　".join(result_parts))
                                        time.sleep(1.5)
                                        st.rerun()

                        else:
                            # ── 非返校打掃：保留原有的排除法 ──
                            with st.form("svc_class_form"):
                                absent = st.multiselect(f"❌ 缺席名單（共 {len(mems)} 人，扣除法）", mems, key="svc_absent")
                                pool = [m for m in mems if m not in absent]
                                st.caption(f"✅ 一般組：{len(pool)} 人")
                                base_h = st.number_input("基礎時數", value=2.0, step=0.5, min_value=0.0, key="svc_base_h")

                                st.markdown("---")
                                spec = st.multiselect("⭐ 加強組（從一般組挑選，另給不同時數）", pool, key="svc_spec")
                                spec_h = st.number_input("加強組時數", value=3.0, step=0.5, min_value=0.0, key="svc_spec_h")
                                st.caption("（若無加強組，此欄留空即可，不影響結果）")

                                pf = st.file_uploader("📸 照片（選填）", type=['jpg', 'png', 'jpeg'], key="svc_pf_cls")

                                if st.form_submit_button("🚀 發放"):
                                    if time.time() - st.session_state.last_action_time < 3:
                                        st.warning("⚠️ 系統處理中，請勿連續點擊！")
                                    elif not pool:
                                        st.error("❌ 發放名單為空，請確認名單設定！")
                                    else:
                                        st.session_state.last_action_time = time.time()
                                        norm = [m for m in pool if m not in spec]
                                        ok_norm, ok_spec = True, True

                                        fb = None
                                        if pf:
                                            pf.seek(0)
                                            fb = pf.read()

                                        if norm:
                                            files_norm = [io.BytesIO(fb)] if fb else None
                                            if files_norm: files_norm[0].name = "p.jpg"
                                            ok_norm = save_entry(
                                                {"日期": str(rd), "班級": rc, "評分項目": "服務時數發放",
                                                 "登錄時間": now_tw.strftime("%Y-%m-%d %H:%M:%S")},
                                                files_norm, norm, base_h, final_category
                                            )
                                        if spec:
                                            files_spec = [io.BytesIO(fb)] if fb else None
                                            if files_spec: files_spec[0].name = "p.jpg"
                                            ok_spec = save_entry(
                                                {"日期": str(rd), "班級": rc, "評分項目": "服務時數發放",
                                                 "登錄時間": now_tw.strftime("%Y-%m-%d %H:%M:%S")},
                                                files_spec, spec, spec_h, f"{final_category}(加強)"
                                            )

                                        if ok_norm and ok_spec:
                                            result_msg = f"✅ 已發放！一般組 {len(norm)} 人 ({base_h}h)"
                                            if spec:
                                                result_msg += f"，加強組 {len(spec)} 人 ({spec_h}h)"
                                            st.success(result_msg)
                                            time.sleep(1.5)
                                            st.rerun()

                # ── 直接輸入學號模式 ──
                elif target_mode == "🔢 直接輸入學號":
                    raw_sid_input = st.text_area(
                        "輸入學號（每行一個，或用逗號分隔）",
                        key="svc_sid_raw",
                        placeholder="例如：\n112001\n112002, 112003",
                        height=150
                    )

                    # 即時解析與驗證（放在 form 外，輸入時即時顯示結果）
                    valid_ids, invalid_ids = [], []
                    if raw_sid_input.strip():
                        raw_ids_list = [s.strip() for s in re.split(r'[\n,、，\s]+', raw_sid_input) if s.strip()]
                        for sid in raw_ids_list:
                            cleaned_sid = clean_id(sid)
                            if cleaned_sid in ROSTER_DICT:
                                if cleaned_sid not in valid_ids:
                                    valid_ids.append(cleaned_sid)
                            else:
                                if sid not in invalid_ids:
                                    invalid_ids.append(sid)

                    if valid_ids:
                        st.success(f"✅ 有效學號 {len(valid_ids)} 人，通過驗證，將發放服務時數。")
                    if invalid_ids:
                        st.error(f"❌ 以下學號不在 Roster 名單中，請確認後再送出：**{', '.join(invalid_ids)}**")

                    with st.form("svc_sid_form"):
                        hours_sid = st.number_input("時數（全體統一）", value=1.0, step=0.5, min_value=0.0, key="svc_sid_h")
                        pf_sid = st.file_uploader("📸 照片（選填）", type=['jpg', 'png', 'jpeg'], key="svc_pf_sid")

                        if st.form_submit_button("🚀 發放"):
                            if time.time() - st.session_state.last_action_time < 3:
                                st.warning("⚠️ 系統處理中，請勿連續點擊！")
                            elif invalid_ids:
                                st.error(f"❌ 尚有 {len(invalid_ids)} 個無效學號，請先修正後再送出！")
                            elif not valid_ids:
                                st.error("❌ 請先在上方輸入至少一個有效學號！")
                            else:
                                st.session_state.last_action_time = time.time()

                                # 按班級分組，讓 service_hours 的班級欄位正確對應
                                cls_groups = {}
                                for sid in valid_ids:
                                    cls_name = ROSTER_DICT.get(sid, "")
                                    if cls_name not in cls_groups:
                                        cls_groups[cls_name] = []
                                    cls_groups[cls_name].append(sid)

                                fb_sid = None
                                if pf_sid:
                                    pf_sid.seek(0)
                                    fb_sid = pf_sid.read()

                                all_ok = True
                                for i, (cls_name, sids) in enumerate(cls_groups.items()):
                                    # 照片只在第一個班級群組帶入，避免重複上傳到 Drive
                                    files_sid = [io.BytesIO(fb_sid)] if (fb_sid and i == 0) else None
                                    if files_sid: files_sid[0].name = "p.jpg"
                                    ok = save_entry(
                                        {"日期": str(rd), "班級": cls_name, "評分項目": "服務時數發放",
                                         "登錄時間": now_tw.strftime("%Y-%m-%d %H:%M:%S")},
                                        files_sid, sids, hours_sid, final_category
                                    )
                                    if not ok:
                                        all_ok = False

                                if all_ok:
                                    st.success(f"✅ 已為 {len(valid_ids)} 位同學發放 {hours_sid}h 服務時數！")
                                    time.sleep(1.5)
                                    st.rerun()

                # ── [V5.35] 衛生糾察隊全員 / 資源回收糾察隊全員 模式 ──
                elif target_mode in ("👮 衛生糾察隊全員", "🗑️ 資源回收糾察隊全員"):
                    _is_trash_team = (target_mode == "🗑️ 資源回收糾察隊全員")
                    _team_label = "資源回收糾察" if _is_trash_team else "衛生糾察"
                    _team_emoji = "🗑️" if _is_trash_team else "👮"
                    _key_suffix = "trash" if _is_trash_team else "clean"

                    # 篩選名單（參照第 3578 行環保點名同樣邏輯）
                    if _is_trash_team:
                        _team_inspectors = [
                            p for p in INSPECTOR_LIST
                            if any(k in p.get("raw_role", "") for k in ["垃圾", "回收", "環保"])
                        ]
                    else:
                        _team_inspectors = [
                            p for p in INSPECTOR_LIST
                            if any(k in p.get("raw_role", "") for k in ["內掃", "外掃", "組長", "機動", "衛生糾察"])
                        ]

                    _team_labels = [p["label"] for p in _team_inspectors]

                    if not _team_labels:
                        st.warning(f"⚠️ 目前 inspectors 名單中沒有「{_team_label}」的人員，無法發放。")
                    else:
                        st.info(
                            f"📋 {_team_emoji} **{_team_label}隊**共 **{len(_team_labels)}** 人。"
                            f"請勾選**未到 / 請假**的同學，其餘預設視為出席。"
                        )

                        with st.form(f"svc_team_form_{_key_suffix}"):
                            _absent_team = st.multiselect(
                                "❌ 未到 / 請假名單（扣除法）",
                                _team_labels,
                                key=f"svc_team_absent_{_key_suffix}"
                            )
                            _present_team = [n for n in _team_labels if n not in _absent_team]

                            st.markdown(f"✅ **預計發放對象：共 {len(_present_team)} 人**")
                            if _present_team:
                                with st.expander("查看預計發放對象名單", expanded=False):
                                    st.write("、".join(_present_team))

                            _team_hours = st.number_input(
                                "時數（全體統一）", value=2.0, step=0.5, min_value=0.0,
                                key=f"svc_team_h_{_key_suffix}"
                            )

                            st.caption(
                                f"💡 將寫入 service_hours：班級＝「糾察隊」、類別＝「**{final_category}**」（沿用上方類別設定）。"
                                f"系統會自動阻擋同一天同類別的重複發放。"
                            )

                            if st.form_submit_button(f"🚀 發放 {_team_emoji} {_team_label}隊時數"):
                                if _block_future_date(rd, "服務時數發放日期"):
                                    pass  # 未來日期被擋
                                elif time.time() - st.session_state.last_action_time < 3:
                                    st.warning("⚠️ 系統處理中，請勿連續點擊！")
                                elif not _present_team:
                                    st.error("❌ 發放名單為空，請確認名單設定！")
                                else:
                                    st.session_state.last_action_time = time.time()
                                    _present_ids = [
                                        n.split("學號:")[1].strip()
                                        for n in _present_team if "學號:" in n
                                    ]
                                    if _present_ids:
                                        _payload = {
                                            "student_list": _present_ids,
                                            "date": str(rd),
                                            "class_name": "糾察隊",
                                            "category": final_category,
                                            "hours": _team_hours,
                                        }
                                        enqueue_task("service_hours_only", _payload)
                                        st.success(
                                            f"✅ 已排程發放 {len(_present_ids)} 位 {_team_label} 各 {_team_hours}h 時數！"
                                            f"（類別：{final_category}）"
                                        )
                                        time.sleep(1.5)
                                        st.rerun()
                                    else:
                                        st.warning("沒有可發放時數的對象")

            # [Fix] _strip_notion_invisible / _parse_claimant_tag 已移至全域定義（約 line 674），
            # Worker 與 Admin UI 共用，不再重複定義。

            # [新增] 愛校服務 2.0：愛校與欠時管理 Tab
                # [V6 新增] 一鍵匯出學校正規時數檔（純讀取，不寫入任何資料）
                st.markdown("---")
                with st.expander("📥 匯出學校正規時數檔（期末交給幹事用）", expanded=False):
                    st.caption("彙總 service_hours 各學生時數 → 反查名冊 → 產出與學校格式相同的 Excel（證明單分頁＋服務時數整批輸入表）。產檔成功後會將本批來源資料蓋上「匯出批號」以利追蹤與封存；除批號欄外不修改任何數字。")
                    _exp_scope = st.radio("匯出範圍", ["🆕 僅未蓋章（尚未匯出過）的資料", "🔁 全部資料（重新匯出）"], horizontal=True, key="exp_scope")
                    _ec1, _ec2, _ec3 = st.columns(3)
                    _exp_default_start = date(today_tw.year, 2, 1) if today_tw.month < 8 else date(today_tw.year, 8, 1)
                    _exp_start = _ec1.date_input("統計起日", _exp_default_start, key="exp_start")
                    _exp_end = _ec2.date_input("統計迄日", today_tw, key="exp_end")
                    _roc = lambda dd: f"{dd.year-1911}{dd.month:02d}{dd.day:02d}"
                    _exp_sign_date = _ec3.text_input("簽呈日期（民國格式）", value=_roc(today_tw), key="exp_sign")
                    _date_span = f"{_roc(_exp_start)}~{_roc(_exp_end)}"
                    st.write(f"證明單日期區間將填為：**{_date_span}**")

                    _ros_up_exp = st.file_uploader("📇 名冊檔（選填：xls/xlsx，僅存本次瀏覽階段記憶體、不寫入雲端，用來帶入班級座號姓名）", type=["xls", "xlsx"], key="exp_roster_up")
                    if _ros_up_exp is not None:
                        try:
                            st.session_state["session_roster"] = parse_roster_upload(_ros_up_exp)
                            st.success(f"✅ 名冊已載入（{len(st.session_state['session_roster'])} 位）。僅存在本次瀏覽階段，關閉頁面即自動消失。")
                        except Exception as _pe:
                            st.error(f"❌ 名冊讀取失敗：{_pe}（.xls 舊格式讀不了時，請先用 Excel 另存為 .xlsx）")

                    if st.button("🚀 產生匯出檔", key="exp_go"):
                        _openpyxl_ok = True
                        try:
                            import openpyxl  # noqa: F401
                        except ImportError:
                            _openpyxl_ok = False
                            st.error("❌ 部署環境缺少 openpyxl 套件：請在 requirements.txt 加入一行 openpyxl 後重新部署，再使用本功能。")
                        if _openpyxl_ok:
                            with st.spinner("讀取 service_hours 與名冊資料中..."):
                                _svc_ws = get_worksheet(SHEET_TABS["service_hours"])
                                _svc_df = pd.DataFrame(_svc_ws.get_all_records()) if _svc_ws else pd.DataFrame()
                            if _svc_df.empty or "學號" not in _svc_df.columns:
                                st.warning("service_hours 分頁沒有資料。")
                            else:
                                _svc_df["_row"] = range(2, len(_svc_df) + 2)  # [V6.2] 對應試算表實際列號（蓋章用）
                                _svc_df["日期"] = pd.to_datetime(_svc_df["日期"], errors="coerce").dt.date
                                _svc_df["時數"] = pd.to_numeric(_svc_df["時數"], errors="coerce").fillna(0)
                                _svc_df["學號"] = _svc_df["學號"].apply(clean_id)
                                _svc_df = _svc_df[(_svc_df["日期"] >= _exp_start) & (_svc_df["日期"] <= _exp_end) & (_svc_df["時數"] > 0)]
                                _svc_df = _svc_df[~_svc_df["類別"].astype(str).str.contains("消警告", na=False)]  # [V6.3] 消警告=銷過用不計時數；補打掃照常計入
                                if "🆕" in _exp_scope and "匯出批號" in _svc_df.columns:
                                    _svc_df = _svc_df[_svc_df["匯出批號"].astype(str).str.strip() == ""]
                                if _svc_df.empty:
                                    st.warning("⚠️ 此範圍內沒有符合的資料。若選了「🆕 僅未蓋章」，代表這段期間的資料都已匯出過，可改選「🔁 全部資料」重新匯出。")

                                def _act_name(cat):
                                    _c0 = str(cat).split("｜")[0].strip()
                                    if "衛生糾察" in _c0 or "整潔評分" in _c0: return "衛生糾察值勤"
                                    if "回收" in _c0 or "環保" in _c0: return "資源回收場值勤"
                                    return _c0 or "其他服務"
                                _svc_df["活動"] = _svc_df["類別"].apply(_act_name)
                                _agg = _svc_df.groupby(["活動", "學號"])["時數"].sum().reset_index()

                                # 反查名冊（roster 分頁需有：學號 / 班級 / 座號 / 姓名）
                                _ros_ws = get_worksheet(SHEET_TABS["roster"])
                                _ros_df = pd.DataFrame(_ros_ws.get_all_records()) if _ros_ws else pd.DataFrame()
                                _colf = lambda df_, kw: next((c for c in df_.columns if kw in c), None)
                                _c_id, _c_cls, _c_seat, _c_name = (_colf(_ros_df, k) for k in ["學號", "班級", "座號", "姓名"])
                                _ros_map = {}
                                if _c_id is not None:
                                    for _, _rr in _ros_df.iterrows():
                                        _ros_map[clean_id(_rr[_c_id])] = (
                                            str(_rr[_c_cls]).strip() if _c_cls else "",
                                            _rr[_c_seat] if _c_seat else "",
                                            str(_rr[_c_name]).strip() if _c_name else "",
                                        )
                                for _s, _v in st.session_state.get("session_roster", {}).items():
                                    _ros_map[_s] = (_v[1], _v[2], _v[3])  # [V6.4 資安] 階段名冊優先
                                _miss_cols = [n for n, c in [("座號", _c_seat), ("姓名", _c_name)] if c is None and not st.session_state.get("session_roster")]
                                if _miss_cols:
                                    st.warning("⚠️ roster 分頁缺少欄位：" + "、".join(_miss_cols) + "，這些欄位將留白。建議至 Google Sheet 補齊後重新產生。")

                                _buf = io.BytesIO()
                                _hdr = ["班級", "座號", "學號\n(務必輸入)", "姓名", "服務活動名稱\n(事由：請簡略說明)", "服務類別\n(填入代號)", "地點", "日期\n(107年3月1日請填1070301)", "時數小時\n(超過8小請以另筆資料登錄)"]
                                _batch_rows = []
                                with pd.ExcelWriter(_buf, engine="openpyxl") as _xw:
                                    for _act, _g in _agg.groupby("活動"):
                                        _g = _g.sort_values("學號")
                                        _recs = []
                                        for _, _r in _g.iterrows():
                                            _cls_v, _seat_v, _name_v = _ros_map.get(str(_r["學號"]), ("", "", ""))
                                            if not _cls_v:
                                                _cls_v = ROSTER_DICT.get(str(_r["學號"]), "")
                                            _recs.append([_cls_v, _seat_v, str(_r["學號"]), _name_v, _act, "C043", "校內", _date_span, float(_r["時數"])])
                                            _batch_rows.append([str(_r["學號"]), _exp_sign_date, _date_span, _date_span, "A001", "C043", float(_r["時數"]), _act])
                                        for _i in range(0, len(_recs), 20):
                                            _sheet_df = pd.DataFrame([["市立中壢高級家事商業職業學校服務學習證明"] + [""] * 8, _hdr] + _recs[_i:_i + 20])
                                            _sheet_df.to_excel(_xw, sheet_name=f"{_act}{_i // 20 + 1}"[:31], index=False, header=False)
                                    pd.DataFrame(_batch_rows, columns=["學號", "簽呈日期", "開始日期", "結束日期", "單位代碼", "服務內容代碼", "時數", "備註"]).to_excel(_xw, sheet_name="服務時數整批輸入表", index=False)

                                # [V6.2] 匯出蓋章：本批來源資料標上匯出批號（只寫批號欄，不動任何數字）
                                _batch_id = f"{_roc(today_tw)}-{time.strftime('%H%M')}"
                                _stamped = 0
                                if not _svc_df.empty:
                                    try:
                                        _hdr_row = _svc_ws.row_values(1)
                                        if "匯出批號" in _hdr_row:
                                            _stamp_col = _hdr_row.index("匯出批號") + 1
                                        else:
                                            _stamp_col = len(_hdr_row) + 1
                                            _svc_ws.update_cell(1, _stamp_col, "匯出批號")
                                        _cells = [gspread.Cell(row=int(_r), col=_stamp_col, value=_batch_id) for _r in _svc_df["_row"].tolist()]
                                        if _cells:
                                            _svc_ws.update_cells(_cells)
                                            _stamped = len(_cells)
                                    except Exception as _stamp_e:
                                        st.warning(f"⚠️ 匯出檔已產生，但蓋章失敗（可重按一次或忽略）：{_stamp_e}")
                                st.success(f"✅ 完成！共 {len(_agg)} 筆學生時數、{_agg['活動'].nunique()} 種活動；已蓋章 {_stamped} 列，批號 **{_batch_id}**。核對無誤、交件完成後，可用試算表選單「🧹 衛生組工具 → 📦 依匯出批號封存」將此批封存。")
                                _no_name = [s for s in _agg["學號"].astype(str) if s not in _ros_map or not _ros_map[s][2]]
                                if _no_name:
                                    st.warning(f"⚠️ 有 {len(_no_name)} 個學號在名冊查不到姓名（欄位留白）：" + "、".join(_no_name[:15]) + ("…" if len(_no_name) > 15 else ""))
                                st.download_button(
                                    "📥 下載時數匯出檔 (.xlsx)",
                                    _buf.getvalue(),
                                    file_name=f"服務時數證明單_{_date_span}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key="exp_dl"
                                )

            with t_debt:
                st.subheader("🤝 愛校與欠時管理")

                # ── [V6.1] 區塊 A0：✅ 愛校完成登記（現場版，取代 Notion 認領流程） ──
                with st.expander("✅ 愛校完成登記（現場版）", expanded=True):
                    st.caption("學生現場完成愛校工作後，在此登記。系統將自動處理：🟢 還時數→扣抵欠時＋寫入補打掃時數；🔔 消警告→寫入消警告紀錄（之後到下方「🖨️ 銷過單核發」產製表單）。糾察懲罰性愛校為現場完成即可，不需登記。")
                    _cs_c1, _cs_c2 = st.columns(2)
                    _cs_date = _cs_c1.date_input("完成日期", today_tw, key="cs_date")
                    _cs_purpose = _cs_c2.radio("登記目的", ["🟢 還時數（補打掃）", "🔔 消警告"], horizontal=True, key="cs_purpose")
                    _cs_c3, _cs_c4 = st.columns(2)
                    _cs_hours = _cs_c3.number_input("時數 (小時)", min_value=0.25, max_value=8.0, value=1.0, step=0.25, key="cs_hours")
                    _cs_time = _cs_c4.text_input("開始時間（選填，會印在銷過單上）", key="cs_time", placeholder="例如：12:10")
                    _cs_area = st.text_input("工作內容（必填，例如：資收場整理、川堂拖地）", key="cs_area")
                    _cs_sids_raw = st.text_area("學號（可多位，用逗號、空白或換行分隔）", key="cs_sids", placeholder="例如：311005, 311024\n411001")
                    _cs_sids = re.findall(r"\d+", _cs_sids_raw or "")
                    if _cs_sids:
                        st.write(f"👥 解析到 {len(_cs_sids)} 位學生：{', '.join(_cs_sids)}")
                    _cs_confirm = st.checkbox("我確認以上學生已完成愛校工作", key="cs_confirm")
                    if st.button("📝 送出登記", key="cs_go"):
                        if not _cs_sids:
                            st.error("❌ 請輸入至少一位學號")
                        elif not _cs_area.strip():
                            st.error("❌ 請填寫工作內容")
                        elif _block_future_date(_cs_date, "完成日期"):
                            pass
                        elif not _cs_confirm:
                            st.warning("⚠️ 請先勾選確認")
                        elif time.time() - st.session_state.last_action_time < 3:
                            st.warning("⚠️ 系統處理中，請勿連續點擊！")
                        else:
                            st.session_state.last_action_time = time.time()
                            _cs_tag = "消警告" if "消警告" in _cs_purpose else "還時數"
                            # 重用 campus_service_verify Worker：格式與 Notion 認領完全相同，
                            # 含 SQLite dedup、扣抵欠時、service_hours 寫入；notion_page_id 留空即略過 Notion 更新
                            _tid = enqueue_task("campus_service_verify", {
                                "claimants": [f"{_s}({_cs_tag})" for _s in _cs_sids],
                                "task_title": _cs_area.strip(),
                                "task_hours": float(_cs_hours),
                                "task_date": str(_cs_date),
                                "notion_page_id": "",
                                "task_area": _cs_area.strip(),
                                "time_start": _cs_time.strip()
                            })
                            if _tid:
                                st.success(f"✅ 已排入處理佇列（{len(_cs_sids)} 人，{_cs_tag}）。約 10–30 秒後生效；消警告請稍後至「🖨️ 銷過單核發」產製表單。")
                            else:
                                st.error("❌ 排入佇列失敗，請重試或檢查網路連線。")

                # ── 區塊 B：⚠️ 欠時懲處結算報表 ──
                with st.expander("⚠️ 欠時懲處結算報表", expanded=True):
                    if st.button("🚀 結算滿 1 小時警告名單", key="debt_settle_btn"):
                        _all_debts = load_student_debts()
                        _warn_list = []
                        for _dsid, _dhours in _all_debts.items():
                            if _dhours >= 1.0:
                                _warnings = math.floor(_dhours)
                                _remaining = round(_dhours - _warnings, 2)
                                _cls_name = ROSTER_DICT.get(_dsid, "未知班級")
                                _warn_list.append({
                                    "學號": _dsid, "班級": _cls_name,
                                    "未完成時數": _dhours, "應記警告支數": _warnings,
                                    "結算後剩餘欠時": _remaining
                                })
                        if not _warn_list:
                            st.success("🎉 目前沒有學生達到記警告的門檻！")
                        else:
                            _warn_df = pd.DataFrame(_warn_list).sort_values(["班級", "學號"]).reset_index(drop=True)
                            st.warning(f"⚠️ 共有 **{len(_warn_df)}** 位學生達到記警告門檻：")
                            st.dataframe(_warn_df, hide_index=True)
                            _csv_data = _warn_df.to_csv(index=False).encode("utf-8-sig")
                            st.download_button(
                                label="📥 下載結算報表 (CSV)",
                                data=_csv_data,
                                file_name=f"欠時結算報表_{datetime.now(TW_TZ).strftime('%Y%m%d')}.csv",
                                mime="text/csv"
                            )
                            _excel_buf = io.BytesIO()
                            with pd.ExcelWriter(_excel_buf, engine="openpyxl") as _ew:
                                _warn_df.to_excel(_ew, sheet_name="欠時結算", index=False)
                            _excel_buf.seek(0)
                            st.download_button(
                                label="📥 下載結算報表 (Excel)",
                                data=_excel_buf.getvalue(),
                                file_name=f"欠時結算報表_{datetime.now(TW_TZ).strftime('%Y%m%d')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="debt_excel_dl"
                            )

                # ── 區塊 B3：📢 欠時數提醒通知 ──
                with st.expander("📢 欠時數提醒通知", expanded=False):
                    st.info("💡 一鍵產生通知文字，直接複製貼到 LINE 群組。")

                    if st.button("📢 產生提醒訊息", key="btn_gen_notify"):
                        _notify_debts = load_student_debts()
                        if not _notify_debts:
                            st.success("🎉 目前沒有學生有欠時數！")
                        else:
                            # 一次讀取備註
                            _notify_notes = {}
                            try:
                                _ws_nd = get_worksheet(SHEET_TABS["student_debts"])
                                if _ws_nd:
                                    for _r in _ws_nd.get_all_records():
                                        _nsid = clean_id(str(_r.get("學號", "")))
                                        _nnote = str(_r.get("備註", "")).strip()
                                        if _nsid and _nnote:
                                            if _nsid not in _notify_notes:
                                                _notify_notes[_nsid] = []
                                            if _nnote not in _notify_notes[_nsid]:
                                                _notify_notes[_nsid].append(_nnote)
                            except Exception:
                                pass

                            # 按班級分組
                            _cls_groups = {}
                            for _sid, _hrs in sorted(_notify_debts.items()):
                                if _hrs <= 0:
                                    continue
                                _cls = ROSTER_DICT.get(_sid, "未知班級")
                                if _cls not in _cls_groups:
                                    _cls_groups[_cls] = []
                                _note_str = "；".join(_notify_notes.get(_sid, []))
                                _cls_groups[_cls].append((_sid, _hrs, _note_str))

                            if not _cls_groups:
                                st.success("🎉 目前沒有學生有欠時數！")
                            else:
                                # 產生 LINE 通知文字
                                _lines = ["⚠️ 返校打掃欠時數提醒", ""]
                                _total_students = 0
                                for _cls_name in sorted(_cls_groups.keys()):
                                    _students = _cls_groups[_cls_name]
                                    _lines.append(f"📌 {_cls_name}")
                                    for _sid, _hrs, _note in _students:
                                        _note_part = f"　{_note}" if _note else ""
                                        _lines.append(f"  • {_sid}（{_hrs:g}hr）{_note_part}")
                                        _total_students += 1
                                    _lines.append("")
                                _lines.append(f"共 {_total_students} 位同學，請盡速安排補打掃。")
                                _lines.append("逾期未完成將依規定記警告處分。")

                                _notify_text = "\n".join(_lines)
                                st.text_area("📋 複製以下文字貼到 LINE 群組", _notify_text, height=300, key="notify_text")
                                st.caption("💡 點進文字框 → 全選 (Ctrl+A) → 複製 (Ctrl+C) → 貼到 LINE")

                                # 同時產生 email 清單（供 Apps Script 使用）
                                _email_lines = ["學號,班級,欠時數,原因"]
                                for _cls_name in sorted(_cls_groups.keys()):
                                    for _sid, _hrs, _note in _cls_groups[_cls_name]:
                                        _email_lines.append(f"{_sid},{_cls_name},{_hrs:g},{_note}")
                                _csv_text = "\n".join(_email_lines)
                                st.download_button(
                                    "📥 下載欠時名單 CSV（可搭配 Apps Script 寄信）",
                                    _csv_text.encode("utf-8-sig"),
                                    f"欠時提醒名單_{datetime.now(TW_TZ).strftime('%Y%m%d')}.csv",
                                    "text/csv",
                                    key="dl_notify_csv"
                                )

                # ── 區塊 C：🖨️ 銷過單核發 (一鍵批次) ──
                with st.expander("🖨️ 銷過單核發 (消警告單)", expanded=True):
                    st.info("💡 系統自動查詢所有「愛校服務(消警告)」紀錄，一鍵產製所有學生的《愛校服務申請單》。")
                    # [V6.4 資安] 階段性名冊：僅存瀏覽器工作階段記憶體，關頁即消失，不寫入雲端
                    if st.session_state.get("session_roster"):
                        st.success(f"📇 名冊已載入本階段（{len(st.session_state['session_roster'])} 位），銷過單將自動帶入姓名與班級。")
                    else:
                        _ros_up_ap = st.file_uploader("📇 名冊檔（選填：帶入姓名與班級用，僅存本次瀏覽階段、不寫入雲端）", type=["xls", "xlsx"], key="appeal_roster_up")
                        if _ros_up_ap is not None:
                            try:
                                st.session_state["session_roster"] = parse_roster_upload(_ros_up_ap)
                                st.success(f"✅ 名冊已載入（{len(st.session_state['session_roster'])} 位）。")
                                st.rerun()
                            except Exception as _pe:
                                st.error(f"❌ 名冊讀取失敗：{_pe}")
                    # [V6.4 Fix] 名冊對照表定義在區塊最上層，批次與單一下載路徑共用
                    _SESS_ROS = st.session_state.get("session_roster", {})
                    _NAME_MAP = load_roster_name_map()
                    _NAME_MAP.update({_s: _v[3] for _s, _v in _SESS_ROS.items()})

                    # [方案A] 顯示模式切換
                    _show_all_issued = st.checkbox(
                        "顯示全部紀錄（含已核發）",
                        value=False,
                        key="show_all_issued",
                        help="勾選後可查看已核發的歷史紀錄，或重新下載補印。"
                    )

                    # [Fix] 共用解析函式：將原始紀錄轉成申請單格式
                    def _parse_appeal_record(rec):
                        """解析 service_hours 中的消警告紀錄，回傳標準化 dict"""
                        _class_field = str(rec.get("班級", ""))
                        _parts = _class_field.split("|")
                        _work_content = _parts[0].strip() if _parts else ""
                        _start_time = _parts[1].strip() if len(_parts) > 1 else ""
                        _hours = 0.0
                        try:
                            _hours = float(rec.get("時數", 0))
                        except (ValueError, TypeError):
                            pass
                        _date_str = str(rec.get("日期", ""))
                        _end_time = ""
                        if _start_time:
                            try:
                                _st = datetime.strptime(_start_time, "%H:%M")
                                _et = _st + timedelta(hours=_hours)
                                _end_time = _et.strftime("%H:%M")
                            except Exception:
                                pass
                        return {
                            "work_content": _work_content,
                            "date": _date_str,
                            "start_time": _start_time,
                            "end_time": _end_time,
                            "hours": _hours,
                            "紀錄ID": str(rec.get("紀錄ID", ""))
                        }

                    if st.button("🔎 查詢消警告紀錄", key="btn_fetch_all_appeals"):
                        try:
                            def _fetch_all_appeal_records():
                                ws_svc = get_worksheet(SHEET_TABS["service_hours"])
                                if not ws_svc:
                                    return []
                                _svc_data = ws_svc.get_all_records()
                                _df_svc = pd.DataFrame(_svc_data)
                                if _df_svc.empty or "類別" not in _df_svc.columns:
                                    return []
                                # 確保核發狀態欄存在
                                if "核發狀態" not in _df_svc.columns:
                                    _df_svc["核發狀態"] = ""
                                _mask = _df_svc["類別"] == "愛校服務(消警告)"
                                return _df_svc[_mask].to_dict('records')

                            with st.spinner("正在查詢消警告紀錄..."):
                                _all_records = execute_with_retry(_fetch_all_appeal_records)

                            # [方案A] 依核發狀態篩選
                            if _show_all_issued:
                                _all_appeal_records = _all_records
                            else:
                                _all_appeal_records = [
                                    r for r in _all_records
                                    if str(r.get("核發狀態", "")).strip() != "已核發"
                                ]
                                _issued_count = len(_all_records) - len(_all_appeal_records)
                                if _issued_count > 0:
                                    st.info(f"ℹ️ 已隱藏 {_issued_count} 筆已核發紀錄。勾選上方「顯示全部」可查看。")

                            if not _all_appeal_records:
                                if _show_all_issued:
                                    st.warning("⚠️ 目前找不到任何消警告服務紀錄。請確認是否已有消警告驗收完成。")
                                else:
                                    st.success("✅ 所有消警告紀錄均已核發完畢！")
                            else:
                                # 依學號分組
                                _grouped = {}
                                for _rec in _all_appeal_records:
                                    _sid = str(_rec.get("學號", "")).strip()
                                    if not _sid:
                                        continue
                                    if _sid not in _grouped:
                                        _grouped[_sid] = []
                                    _grouped[_sid].append(_rec)

                                # 顯示摘要表格
                                _summary_rows = []
                                for _sid, _recs in sorted(_grouped.items()):
                                    _cls_name = (_SESS_ROS.get(clean_id(_sid), ["","","",""])[1] or ROSTER_DICT.get(clean_id(_sid), "未知班級"))
                                    _total_h = sum(float(r.get("時數", 0)) for r in _recs)
                                    _issued_tag = "✅ 已核發" if all(str(r.get("核發狀態","")).strip() == "已核發" for r in _recs) else "⏳ 未核發"
                                    _summary_rows.append({
                                        "學號": _sid,
                                        "班級": _cls_name,
                                        "服務筆數": len(_recs),
                                        "總時數": round(_total_h, 2),
                                        "狀態": _issued_tag
                                    })

                                st.success(f"✅ 共找到 **{len(_all_appeal_records)}** 筆消警告紀錄，涵蓋 **{len(_grouped)}** 位學生：")
                                _summary_df = pd.DataFrame(_summary_rows)
                                st.dataframe(_summary_df, hide_index=True)

                                # 存入 session_state 供後續下載
                                st.session_state["_appeal_grouped"] = _grouped
                                st.session_state["_appeal_summary"] = _summary_rows

                        except Exception as e:
                            st.error(f"❌ 查詢消警告紀錄時發生錯誤：{e}")

                    # ── 查詢結果存在 session_state 時，顯示下載區 ──
                    if st.session_state.get("_appeal_grouped"):
                        _grouped = st.session_state["_appeal_grouped"]
                        _summary_rows = st.session_state.get("_appeal_summary", [])

                        st.markdown("---")
                        st.markdown("#### 📥 下載銷過單")

                        # ── 一鍵下載全部（ZIP） ──
                        if st.button("📦 一鍵打包下載全部銷過單 (ZIP)", key="btn_dl_all_zip"):
                            import zipfile as _zipfile
                            _zip_buf = io.BytesIO()
                            _gen_count = 0
                            _all_rids_to_mark = []
                            with _zipfile.ZipFile(_zip_buf, 'w', _zipfile.ZIP_DEFLATED) as _zf:
                                for _sid, _recs in sorted(_grouped.items()):
                                    _cls_name = ROSTER_DICT.get(clean_id(_sid), "未知班級")
                                    _parsed = [_parse_appeal_record(r) for r in _recs[:8]]
                                    if _parsed:
                                        _excel_bytes = generate_appeal_form_excel(clean_id(_sid), _cls_name, _parsed, student_name=_NAME_MAP.get(clean_id(_sid), ""))
                                        _zf.writestr(f"愛校申請單_{_cls_name}_{_sid}.xlsx", _excel_bytes)
                                        _gen_count += 1
                                        # [方案A] 收集本次產製的紀錄ID
                                        _all_rids_to_mark += [p["紀錄ID"] for p in _parsed if p.get("紀錄ID")]
                            _zip_buf.seek(0)
                            st.download_button(
                                label=f"📦 下載 ZIP（共 {_gen_count} 份申請單）",
                                data=_zip_buf.getvalue(),
                                file_name=f"消警告銷過單_全部_{datetime.now(TW_TZ).strftime('%Y%m%d')}.zip",
                                mime="application/zip",
                                key="dl_all_appeal_zip"
                            )
                            # [方案A] 下載後標記已核發
                            if _all_rids_to_mark:
                                with st.spinner("📝 標記核發狀態中..."):
                                    _mark_service_hours_issued(_all_rids_to_mark)
                            st.success(f"✅ 已產製 {_gen_count} 份銷過單，並標記為已核發！下次查詢將不再顯示這些紀錄。")

                        # ── 個別學生下載 ──
                        st.markdown("##### 或選擇單一學生下載")
                        _sid_options = [f"{r['學號']} ({r['班級']}, {r['總時數']}h, {r['服務筆數']}筆)"
                                        for r in _summary_rows]
                        _sel_appeal_student = st.selectbox("選擇學生", _sid_options, key="sel_appeal_student")

                        if _sel_appeal_student:
                            _sel_sid = _sel_appeal_student.split(" ")[0]
                            _sel_recs = _grouped.get(_sel_sid, [])
                            _sel_cls = (st.session_state.get("session_roster", {}).get(clean_id(_sel_sid), ["","","",""])[1] or ROSTER_DICT.get(clean_id(_sel_sid), "未知班級"))

                            if _sel_recs:
                                _parsed_recs = [_parse_appeal_record(r) for r in _sel_recs[:8]]

                                # 顯示預覽
                                _preview_df = pd.DataFrame([{
                                    "序號": i + 1,
                                    "工作內容": r["work_content"],
                                    "服務日期": r["date"],
                                    "起始時間": r["start_time"] if r["start_time"] else "—",
                                    "結束時間": r["end_time"] if r["end_time"] else "—",
                                    "時數": r["hours"]
                                } for i, r in enumerate(_parsed_recs)])
                                st.dataframe(_preview_df, hide_index=True)

                                _total_h = sum(r["hours"] for r in _parsed_recs)
                                st.metric("總計服務時數", f"{_total_h:g} 小時")

                                # 產製個別 Excel
                                _excel_bytes = generate_appeal_form_excel(
                                    clean_id(_sel_sid), _sel_cls, _parsed_recs,
                                    student_name=_NAME_MAP.get(clean_id(_sel_sid), "")
                                )
                                if st.download_button(
                                    label=f"📥 下載 {_sel_cls} {_sel_sid} 的銷過單",
                                    data=_excel_bytes,
                                    file_name=f"愛校申請單_{_sel_cls}_{_sel_sid}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key="dl_single_appeal"
                                ):
                                    # [方案A] 個別下載後標記已核發
                                    _single_rids = [p["紀錄ID"] for p in _parsed_recs if p.get("紀錄ID")]
                                    if _single_rids:
                                        _mark_service_hours_issued(_single_rids)

                        st.caption("💡 下載後請用 Excel 開啟，確認格式無誤後列印 A4 紙張，再持表單至學務處完成後續流程。")

        elif pwd_input != "":
            st.error("密碼錯誤")

except Exception as e:
    st.error(f"❌ 系統發生錯誤: {str(e)}")
    st.code(traceback.format_exc())
